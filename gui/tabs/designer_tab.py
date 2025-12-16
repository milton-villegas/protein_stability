#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Experimental Modeling Suite v0.2.0
Comprehensive Experimental Design & Modeling Platform
Creates advanced experiment designs, fits models, checks assumptions,
and suggests next experiments. Exports to XLSX and Opentrons formats.

Milton F. Villegas

Enhanced with:
- pyDOE3: Design of experiments for Python (BSD-3-Clause License)
  https://github.com/relf/pyDOE3
- SMT: Surrogate Modeling Toolbox for optimized LHS (BSD-3-Clause License)
  https://github.com/SMTorg/smt

Design Types Available:
  • Full Factorial - All possible combinations
  • Latin Hypercube Sampling - Space-filling designs
  • 2-Level Fractional Factorial - Efficient screening (Resolution III, IV, V)
  • Plackett-Burman - Ultra-efficient screening for many factors
  • Central Composite Design - Response surface optimization
  • Box-Behnken - Response surface without extreme corners
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import itertools
import csv
import os
import re
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import numpy as np
from utils.constants import AVAILABLE_FACTORS

# Optional XLSX export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# Optional pyDOE3 for advanced designs
try:
    import pyDOE3
    HAS_PYDOE3 = True
except Exception:
    HAS_PYDOE3 = False

# Optional SMT for optimized LHS
try:
    from smt.sampling_methods import LHS
    HAS_SMT = True
except Exception:
    HAS_SMT = False

def validate_numeric_input(action, char, entry_value):
    """Validate numeric input - allows digits, decimal, minus, comma for multiple values"""
    # Allow all deletions
    if action == '0':
        return True
    # Allow empty
    if char == '':
        return True
    # Allow digits, decimal point, negative sign, and comma for multiple entries
    return char.isdigit() or char in '.-,'

def validate_single_numeric_input(action, char, entry_value):
    """Single numeric value only (no commas)"""

    if action == '0':
        return True

    if char == '':
        return True

    return char.isdigit() or char in '.-'

def validate_alphanumeric_input(action, char, entry_value):
    """Validate alphanumeric input - allows letters, digits, spaces, and common punctuation for names"""
    # Allow all deletions
    if action == '0':
        return True
    # Allow empty
    if char == '':
        return True
    # Allow letters, digits, spaces, hyphens, parentheses, and commas
    return char.isalnum() or char in ' -(),.'

class FactorModel:
    """Model with validation for factors and stock concentrations"""

    def __init__(self):
        """Initialize empty FactorModel with no factors or stock concentrations"""
        self._factors: Dict[str, List[str]] = {}
        self._stock_concs: Dict[str, float] = {}

    def add_factor(self, name: str, levels: List[str], stock_conc: Optional[float] = None):
        """
        Add a new factor to the model.

        Args:
            name: Factor name (e.g., 'nacl', 'buffer pH')
            levels: List of factor levels as strings
            stock_conc: Optional stock concentration for volume calculations

        Raises:
            ValueError: If name is empty or levels list is empty
        """
        name = name.strip()
        if not name:
            raise ValueError("Factor name cannot be empty.")
        if not levels:
            raise ValueError("At least one level is required.")
        self._factors[name] = list(levels)
        if stock_conc is not None:
            self._stock_concs[name] = stock_conc

    def update_factor(self, name: str, levels: List[str], stock_conc: Optional[float] = None):
        """
        Update an existing factor's levels and/or stock concentration.

        Args:
            name: Factor name to update
            levels: New list of factor levels
            stock_conc: Optional new stock concentration

        Raises:
            ValueError: If factor doesn't exist or levels list is empty
        """
        if name not in self._factors:
            raise ValueError(f"Factor '{name}' does not exist.")
        if not levels:
            raise ValueError("At least one level is required.")
        self._factors[name] = list(levels)
        if stock_conc is not None:
            self._stock_concs[name] = stock_conc

    def remove_factor(self, name: str):
        """
        Remove a factor and its stock concentration from the model.

        Args:
            name: Factor name to remove
        """
        if name in self._factors:
            del self._factors[name]
        if name in self._stock_concs:
            del self._stock_concs[name]

    def get_factors(self) -> Dict[str, List[str]]:
        """
        Get all factors and their levels.

        Returns:
            Dictionary mapping factor names to level lists (copy)
        """
        return {k: list(v) for k, v in self._factors.items()}

    def get_stock_conc(self, name: str) -> Optional[float]:
        """
        Get stock concentration for a specific factor.

        Args:
            name: Factor name

        Returns:
            Stock concentration or None if not set
        """
        return self._stock_concs.get(name)

    def get_all_stock_concs(self) -> Dict[str, float]:
        """
        Get all stock concentrations.

        Returns:
            Dictionary mapping factor names to stock concentrations (copy)
        """
        return dict(self._stock_concs)

    def clear(self):
        """Remove all factors and stock concentrations from the model"""
        self._factors.clear()
        self._stock_concs.clear()

    def total_combinations(self) -> int:
        """
        Calculate total number of full factorial combinations.

        Returns:
            Product of all factor level counts, or 0 if no factors
        """
        if not self._factors:
            return 0
        # Multiply all factor level counts (full factorial)
        result = 1
        for levels in self._factors.values():
            result *= len(levels)
        return result

class FactorEditDialog(tk.Toplevel):
    """Inline factor editor dialog with mandatory stock concentration"""
    def __init__(self, parent, factor_name: str, existing_levels: List[str] = None, 
                 stock_conc: Optional[float] = None):
        super().__init__(parent)

        display_name = AVAILABLE_FACTORS.get(factor_name, factor_name)
        
        self.title(f"Edit Factor: {display_name}")
        self.geometry("550x600")
        self.transient(parent)
        self.grab_set()
        
        self.factor_name = factor_name
        self.result_levels = None
        self.result_stock = None
        
        # Store entry widgets for key binding
        self.stock_entry_widget = None
        self.level_entry_widget = None
        
        # Validation commands with action code for proper backspace handling
        vcmd = (self.register(validate_numeric_input), '%d', '%S', '%P')  # For levels (allows commas)
        vcmd_stock = (self.register(validate_single_numeric_input), '%d', '%S', '%P')  # For stock (no commas)
        
        # For detergent and reducing_agent (categorical text), allow alphanumeric input
        # Buffer pH still uses numeric validation (values 1-14)
        if factor_name in ["detergent", "reducing_agent"]:
            vcmd_levels = (self.register(validate_alphanumeric_input), '%d', '%S', '%P')
        else:
            vcmd_levels = vcmd
        
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Factor info
        info_frame = ttk.LabelFrame(main_frame, text="Factor Information", padding=8)
        info_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(info_frame, text=f"Factor: {display_name}", 
                 font=("TkDefaultFont", 10, "bold")).pack(anchor="w")
        
        # Stock concentration
        if factor_name not in ["buffer pH", "detergent", "reducing_agent"]:  # Categorical factors don't need stock
            stock_frame = ttk.LabelFrame(info_frame, text="Stock Concentration (required)", padding=8)
            stock_frame.pack(fill=tk.X, pady=(8, 0))
            
            # Concentration entry and unit dropdown
            entry_frame = ttk.Frame(stock_frame)
            entry_frame.pack(fill=tk.X)
            
            ttk.Label(entry_frame, text="Value:").pack(side=tk.LEFT, padx=(0, 5))
            self.stock_var = tk.StringVar(value=str(stock_conc) if stock_conc else "")
            stock_entry = ttk.Entry(entry_frame, textvariable=self.stock_var, width=12,
                                   validate='key', validatecommand=vcmd_stock)
            stock_entry.pack(side=tk.LEFT, padx=(0, 10))
            stock_entry.bind('<Return>', lambda e: self._try_save())  # Enter key to proceed
            self.stock_entry_widget = stock_entry  # Store for key binding
            
            ttk.Label(entry_frame, text="Unit:").pack(side=tk.LEFT, padx=(0, 5))
            
            # Determine unit options based on factor name
            unit_options = self._get_unit_options(factor_name)
            self.unit_var = tk.StringVar(value=unit_options[0])
            unit_dropdown = ttk.Combobox(entry_frame, textvariable=self.unit_var, 
                                        values=unit_options, state="readonly", width=8)
            unit_dropdown.pack(side=tk.LEFT)
        
        # Levels editor
        levels_frame = ttk.LabelFrame(main_frame, text="Levels", padding=8)
        levels_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # Hints - different for categorical vs numeric factors
        if factor_name == "detergent":
            hint_text = "Tip: Enter detergent names (e.g., Tween-20, Triton X-100, None)"
        elif factor_name == "reducing_agent":
            hint_text = "Tip: Enter reducing agent names (e.g., DTT, TCEP, BME, None)"
        elif factor_name == "buffer pH":
            hint_text = "Tip: Enter pH values from 1-14 (e.g., 7.0, 7.5, 8.0)"
        else:
            hint_text = "Tip: Use commas to add multiple values at once (e.g., 2, 4, 6, 8)"
        
        hint_label = ttk.Label(levels_frame, 
                              text=hint_text,
                              foreground="gray", font=("TkDefaultFont", 9, "italic"))
        hint_label.pack(anchor="w", pady=(0, 5))
        
        # Frame
        quick_frame = ttk.Frame(levels_frame)
        quick_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(quick_frame, text="Add level:").pack(side=tk.LEFT)
        self.level_var = tk.StringVar()
        level_entry = ttk.Entry(quick_frame, textvariable=self.level_var, width=20,
                               validate='key', validatecommand=vcmd_levels)
        level_entry.pack(side=tk.LEFT, padx=5)
        level_entry.bind('<Return>', lambda e: self._on_level_entry_enter())
        self.level_entry_widget = level_entry  # Store for key binding
        
        ttk.Button(quick_frame, text="Add", 
                  command=self._add_level).pack(side=tk.LEFT, padx=2)
        
        # Listbox
        list_frame = ttk.Frame(levels_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.levels_listbox = tk.Listbox(list_frame, selectmode=tk.SINGLE, height=10)
        self.levels_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Bind Backspace and Delete keys to remove selected item
        self.levels_listbox.bind('<BackSpace>', lambda e: self._delete_level())
        self.levels_listbox.bind('<Delete>', lambda e: self._delete_level())
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", 
                                 command=self.levels_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.levels_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Load existing levels
        if existing_levels:
            for level in existing_levels:
                self.levels_listbox.insert(tk.END, level)
        
        # Buttons
        btn_frame = ttk.Frame(levels_frame)
        btn_frame.pack(fill=tk.X, pady=(8, 0))
        
        ttk.Button(btn_frame, text="Delete Selected", 
                  command=self._delete_level).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Clear All", 
                  command=self._clear_levels).pack(side=tk.LEFT, padx=2)
        
        # Generate sequence
        seq_frame = ttk.LabelFrame(main_frame, text="Generate Sequence", padding=8)
        seq_frame.pack(fill=tk.X, pady=(0, 8))
        
        seq_input_frame = ttk.Frame(seq_frame)
        seq_input_frame.pack(fill=tk.X)
        
        ttk.Label(seq_input_frame, text="From:").pack(side=tk.LEFT)
        self.seq_start = tk.StringVar()
        ttk.Entry(seq_input_frame, textvariable=self.seq_start, width=8,
                 validate='key', validatecommand=vcmd).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(seq_input_frame, text="To:").pack(side=tk.LEFT, padx=(10, 0))
        self.seq_end = tk.StringVar()
        ttk.Entry(seq_input_frame, textvariable=self.seq_end, width=8,
                 validate='key', validatecommand=vcmd).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(seq_input_frame, text="Step:").pack(side=tk.LEFT, padx=(10, 0))
        self.seq_step = tk.StringVar()
        ttk.Entry(seq_input_frame, textvariable=self.seq_step, width=8,
                 validate='key', validatecommand=vcmd).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(seq_input_frame, text="Generate", 
                  command=self._generate_sequence).pack(side=tk.LEFT, padx=5)
        
        # Final buttons
        final_btn_frame = ttk.Frame(main_frame)
        final_btn_frame.pack(fill=tk.X)
        
        ttk.Button(final_btn_frame, text="Save", 
                  command=self._save).pack(side=tk.RIGHT, padx=2)
        ttk.Button(final_btn_frame, text="Cancel", 
                  command=self.destroy).pack(side=tk.RIGHT, padx=2)
        
        # Bind Enter key globally for the dialog
        def on_enter_key(event):
            # If focus is on level entry, add the level
            if event.widget == self.level_entry_widget:
                return  # Let _on_level_entry_enter handle it
            # If focus is on stock entry, don't save yet
            if hasattr(self, 'stock_entry_widget') and event.widget == self.stock_entry_widget:
                return
            # Otherwise, if we have levels, save the dialog
            if self.levels_listbox.size() > 0:
                self._save()
        
        self.bind('<Return>', on_enter_key)
        
        # Set initial focus
        if self.stock_entry_widget:
            # Factor has stock concentration - focus there first
            self.stock_entry_widget.focus()
        else:
            # Buffer pH (no stock) - focus on level entry
            level_entry.focus()
    
    def _get_unit_options(self, factor_name: str) -> List[str]:
        """Get appropriate unit options for factor"""
        if factor_name in ["buffer_concentration", "nacl", "kcl", "zinc", "magnesium", "calcium", "reducing_agent_concentration"]:
            return ["mM", "M", "µM"]
        elif factor_name in ["glycerol", "dmso", "detergent_concentration"]:
            return ["%", "% v/v", "% w/v"]
        else:
            # Custom factor - allow common units
            return ["mM", "%", "M", "µM", "mg/mL", "µg/mL", "U/mL"]
    
    def _on_level_entry_enter(self):
        """Handle Enter key in level entry - add level or save if empty"""
        level_text = self.level_var.get().strip()
        
        if level_text:
            # Has text - add the level
            self._add_level()
        else:
            # Empty box - if we have levels, save the dialog
            if self.levels_listbox.size() > 0:
                self._save()
        
    def _add_level(self):
        level_text = self.level_var.get().strip()
        if not level_text:
            return
        
        # Check for comma-separated values
        if ',' in level_text:
            # Split and add multiple
            parts = [p.strip() for p in level_text.split(',') if p.strip()]
            for part in parts:
                # Parse and validate each part
                parsed_value = self._parse_numeric_value(part)
                if parsed_value is not None:
                    # Validate range (pH, percentage, etc.)
                    is_valid, error_msg = self._validate_range(parsed_value, self.factor_name)
                    if not is_valid:
                        messagebox.showerror("Invalid Value", error_msg)
                        return
                    level_clean = str(parsed_value)
                else:
                    level_clean = part
                
                if level_clean not in self.levels_listbox.get(0, tk.END):
                    self.levels_listbox.insert(tk.END, level_clean)
        else:
            # Single value - parse and validate
            parsed_value = self._parse_numeric_value(level_text)
            if parsed_value is not None:
                # Validate range (pH, percentage, etc.)
                is_valid, error_msg = self._validate_range(parsed_value, self.factor_name)
                if not is_valid:
                    messagebox.showerror("Invalid Value", error_msg)
                    return
                level_clean = str(parsed_value)
            else:
                level_clean = level_text
            
            if level_clean not in self.levels_listbox.get(0, tk.END):
                self.levels_listbox.insert(tk.END, level_clean)
        
        # Clear entry and keep focus
        self.level_var.set("")
        self.level_entry_widget.focus()
    
    def _delete_level(self):
        selection = self.levels_listbox.curselection()
        if selection:
            index = selection[0]
            self.levels_listbox.delete(index)
            
            # Auto-select the next available item for continuous deletion
            size = self.levels_listbox.size()
            if size > 0:
                # If we deleted the last item, select the new last item
                if index >= size:
                    self.levels_listbox.selection_set(size - 1)
                else:
                    # Select the item that took the deleted item's place
                    self.levels_listbox.selection_set(index)
    
    def _clear_levels(self):
        self.levels_listbox.delete(0, tk.END)
    
    def _generate_sequence(self):
        try:
            start = float(self.seq_start.get())
            end = float(self.seq_end.get())
            step = float(self.seq_step.get())
            
            if step <= 0:
                messagebox.showerror("Invalid Step", "Step must be positive.")
                return
            
            if start >= end:
                messagebox.showerror("Invalid Range", "Start must be less than End.")
                return
            
            # Generate sequence
            current = start
            while current <= end:
                level_str = str(current) if current % 1 != 0 else str(int(current))
                if level_str not in self.levels_listbox.get(0, tk.END):
                    self.levels_listbox.insert(tk.END, level_str)
                current += step
        
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter valid numeric values.")
    
    def _try_save(self):
        """Try to save - called by Enter key"""
        self._save()
    
    def _parse_numeric_value(self, value_str: str) -> Optional[float]:
        """Parse a numeric value, handling various formats"""
        try:
            return float(value_str)
        except ValueError:
            return None
    
    def _validate_range(self, value: float, factor_name: str) -> Tuple[bool, str]:
        """Validate that value is in reasonable range for factor"""
        # pH validation: 1-14
        if factor_name == "buffer pH":
            if value < 1 or value > 14:
                return False, (f"Invalid pH value: {value}\n\n"
                             f"pH must be between 1.0 and 14.0\n"
                             f"(Typical range: 2-12 for most buffers)")
        
        # Percentage validation: 0-100
        elif factor_name in ["glycerol", "dmso", "detergent_concentration"]:
            if value < 0 or value > 100:
                return False, (f"Invalid percentage: {value}%\n\n"
                             f"Percentage must be between 0 and 100%")
        
        # Concentration validation: 0-10000 mM
        elif factor_name in ["buffer_concentration", "nacl", "kcl", "zinc", "magnesium", "calcium", "reducing_agent_concentration"]:
            if value < 0:
                return False, f"Concentration cannot be negative: {value}"
            if value > 10000:
                return False, (f"Very high concentration: {value} mM\n\n"
                             f"Please verify this is correct.\n"
                             f"(Typical range: 10-1000 mM)")
        
        return True, ""
    
    def _save(self):
        # Get levels
        levels = list(self.levels_listbox.get(0, tk.END))
        if not levels:
            messagebox.showerror("No Levels", "Please add at least one level.")
            return
        
        # Get stock concentration only for non-categorical factors
        stock = None
        if self.factor_name not in ["buffer pH", "detergent", "reducing_agent"]:
            stock_text = self.stock_var.get().strip()
            if not stock_text:
                messagebox.showerror("Missing Stock Concentration", 
                    "Stock concentration is required for this factor.")
                return
            
            try:
                stock_base = float(stock_text)
                if stock_base <= 0:
                    messagebox.showerror("Invalid Stock", 
                        "Stock concentration must be positive.")
                    return
            except ValueError:
                messagebox.showerror("Invalid Stock", 
                    "Stock concentration must be a number.")
                return
            
            # Get unit and convert to base unit if needed
            unit = self.unit_var.get()
            
            # Validate based on expected range
            is_valid, error_msg = self._validate_range(stock_base, self.factor_name)
            if not is_valid:
                messagebox.showerror("Invalid Range", error_msg)
                return
            
            stock = stock_base
            
            # Check that no level exceeds stock concentration
            invalid_levels = []
            for level in levels:
                level_value = self._parse_numeric_value(level)
                if level_value is None:
                    try:
                        level_value = float(level)
                    except ValueError:
                        continue
                
                is_valid, error_msg = self._validate_range(level_value, self.factor_name)
                if not is_valid:
                    messagebox.showerror("Invalid Level Range", error_msg)
                    return
                
                # Check if exceeds stock
                if level_value > stock_base:
                    invalid_levels.append(f"{level} (parsed as {level_value})")
            
            if invalid_levels:
                messagebox.showerror("Invalid Concentrations", 
                    f"The following concentration(s) EXCEED the stock concentration:\n\n"
                    f"Stock: {stock_base} {unit}\n"
                    f"Invalid levels: {', '.join(invalid_levels)}\n\n"
                    f"⚠️ You cannot make a solution more concentrated than your stock!\n\n"
                    f"Either:\n"
                    f"• Increase stock concentration, OR\n"
                    f"• Reduce the level values")
                return
        
        # Sort levels numerically if possible
        try:
            levels = sorted(levels, key=lambda x: float(x))
        except ValueError:
            # fallback to string sort if not all numeric
            levels = sorted(levels)
        
        self.result_levels = levels
        self.result_stock = stock
        self.destroy()


class DesignerTab(ttk.Frame):
    """DoE Designer GUI tab for creating experimental designs"""

    # Define categorical factor pairings as class constant
    CATEGORICAL_PAIRS = {
        "buffer pH": "buffer_concentration",
        "detergent": "detergent_concentration",
        "reducing_agent": "reducing_agent_concentration"
    }

    def __init__(self, parent, project, main_window):
        """
        Initialize the Designer Tab.

        Args:
            parent: Parent tkinter widget
            project: Project instance for data persistence
            main_window: Main application window reference
        """
        super().__init__(parent)
        self.project = project
        self.main_window = main_window

        self.model = FactorModel()
        self._build_ui()
        self._update_display()

    def _build_ui(self):
        """Build the complete Designer Tab user interface"""
        # Configure grid
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        
        # Main container with reduced padding
        main_container = ttk.Frame(self, padding=8)
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.columnconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=2)
        main_container.rowconfigure(0, weight=1)
        main_container.rowconfigure(1, weight=0)
        
        # Left panel: Factor Selection
        left_panel = ttk.Frame(main_container)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(0, weight=0)
        left_panel.rowconfigure(1, weight=1)
        
        # Quick actions: Custom factor
        quick_frame = ttk.LabelFrame(left_panel, text="Quick Actions", padding=8)
        quick_frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        
        ttk.Button(quick_frame, text="Add Custom Factor",
                  command=self._add_custom_factor).pack(fill=tk.X, pady=2)
        ttk.Button(quick_frame, text="Clear All",
                  command=self._clear_all).pack(fill=tk.X, pady=2)
        
        # Available factors
        factors_frame = ttk.LabelFrame(left_panel, text="Available Factors", padding=8)
        factors_frame.grid(row=1, column=0, sticky="nsew")
        factors_frame.columnconfigure(0, weight=1)
        factors_frame.rowconfigure(1, weight=1)
        
        ttk.Label(factors_frame, text="Double-click to add:", 
                 font=("TkDefaultFont", 9, "italic")).grid(row=0, column=0, sticky="w", pady=(0, 4))
        
        self.available_listbox = tk.Listbox(factors_frame, exportselection=False)
        self.available_listbox.grid(row=1, column=0, sticky="nsew")
        self.available_listbox.bind('<Double-Button-1>', lambda e: self._quick_add_factor())
        
        scrollbar = ttk.Scrollbar(factors_frame, orient="vertical", 
                                 command=self.available_listbox.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.available_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Populate available factors - organized by category
        # Define categories and their factors in order
        factor_categories = [
            ("--- BUFFER SYSTEM ---", [
                "buffer pH",
                "buffer_concentration"
            ]),
            ("--- DETERGENTS ---", [
                "detergent",
                "detergent_concentration"
            ]),
            ("--- REDUCING AGENTS ---", [
                "reducing_agent",
                "reducing_agent_concentration"
            ]),
            ("--- SALTS ---", [
                "nacl",
                "kcl"
            ]),
            ("--- METALS ---", [
                "zinc",
                "magnesium",
                "calcium"
            ]),
            ("--- ADDITIVES ---", [
                "glycerol",
                "dmso"
            ])
        ]
        
        # Populate listbox with categories
        for category_name, factor_keys in factor_categories:
            # Add category header (disabled/unselectable)
            self.available_listbox.insert(tk.END, category_name)
            # Make category header visually distinct (we'll handle selection later)
            
            # Add factors in this category
            for key in factor_keys:
                display_name = AVAILABLE_FACTORS.get(key, key)
                self.available_listbox.insert(tk.END, f"  {display_name}")
            
            # Add empty line for spacing between categories
            self.available_listbox.insert(tk.END, "")
        
        # Right panel: Current Design
        right_panel = ttk.Frame(main_container)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=1)
        right_panel.rowconfigure(1, weight=0)
        
        # Design table
        design_frame = ttk.LabelFrame(right_panel, text="Current Design", padding=8)
        design_frame.grid(row=0, column=0, sticky="nsew")
        design_frame.columnconfigure(0, weight=1)
        design_frame.rowconfigure(0, weight=1)
        
        # Treeview for factors
        self.tree = ttk.Treeview(design_frame, 
                                columns=("factor", "levels", "count", "stock"),
                                show="headings", height=10, selectmode="browse")
        self.tree.heading("factor", text="Factor")
        self.tree.heading("levels", text="Levels")
        self.tree.heading("count", text="# Levels")
        self.tree.heading("stock", text="Stock Conc")
        
        self.tree.column("factor", width=130, anchor="w")
        self.tree.column("levels", width=250, anchor="w")
        self.tree.column("count", width=70, anchor="center")
        self.tree.column("stock", width=90, anchor="center")
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind('<Double-Button-1>', lambda e: self._edit_factor())
        
        tree_scroll = ttk.Scrollbar(design_frame, orient="vertical", command=self.tree.yview)
        tree_scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        # Instructions
        ttk.Label(design_frame, text="Double-click a factor to edit",
                 font=("TkDefaultFont", 8, "italic")).grid(row=1, column=0, sticky="w", pady=(4, 0))
        
        # Factor controls
        ctrl_frame = ttk.Frame(design_frame)
        ctrl_frame.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        
        ttk.Button(ctrl_frame, text="Edit",
                  command=self._edit_factor).pack(side=tk.LEFT, padx=2)
        ttk.Button(ctrl_frame, text="Delete",
                  command=self._delete_factor).pack(side=tk.LEFT, padx=2)
        
        # ========== NEW: Design Type Selection ==========
        design_type_frame = ttk.LabelFrame(right_panel, text="Design Type", padding=8)
        design_type_frame.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        design_type_frame.columnconfigure(0, weight=1)
        
        # Dropdown for design type
        dropdown_frame = ttk.Frame(design_type_frame)
        dropdown_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(dropdown_frame, text="Select Design:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.design_type_var = tk.StringVar(value="full_factorial")
        
        design_options = [
            ("full_factorial", "Full Factorial (all combinations)"),
            ("lhs", "Latin Hypercube (space-filling)"),
            ("fractional", "2-Level Fractional Factorial (screening)"),
            ("plackett_burman", "Plackett-Burman (efficient screening)"),
            ("central_composite", "Central Composite (optimization)"),
            ("box_behnken", "Box-Behnken (optimization)")
        ]
        
        self.design_dropdown = ttk.Combobox(dropdown_frame, 
                                           textvariable=self.design_type_var,
                                           values=[desc for _, desc in design_options],
                                           state="readonly", width=40)
        self.design_dropdown.pack(side=tk.LEFT)
        self.design_dropdown.current(0)
        self.design_dropdown.bind('<<ComboboxSelected>>', lambda e: self._on_design_type_changed())
        
        # Map display names back to internal values
        self.design_map = {desc: val for val, desc in design_options}
        self.design_map_reverse = {val: desc for val, desc in design_options}
        
        if not HAS_PYDOE3:
            warning_label = ttk.Label(dropdown_frame, text="⚠ pyDOE3 required for advanced designs", 
                                     foreground="red", font=("TkDefaultFont", 8))
            warning_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # Container for design-specific controls
        self.design_controls_frame = ttk.Frame(design_type_frame)
        self.design_controls_frame.pack(fill=tk.X, pady=(5, 0))
        
        # === LHS-specific controls ===
        self.lhs_controls = ttk.Frame(self.design_controls_frame)
        
        sample_frame = ttk.Frame(self.lhs_controls)
        sample_frame.pack(fill=tk.X)
        
        ttk.Label(sample_frame, text="Sample Size:").pack(side=tk.LEFT, padx=(0, 5))
        self.sample_size_var = tk.IntVar(value=96)
        vcmd = (self.register(validate_single_numeric_input), '%d', '%S', '%P')
        ttk.Entry(sample_frame, textvariable=self.sample_size_var, 
                 width=8, validate='key', validatecommand=vcmd).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(sample_frame, text="(8-384)", 
                 font=("TkDefaultFont", 8, "italic")).pack(side=tk.LEFT)
        
        optimize_frame = ttk.Frame(self.lhs_controls)
        optimize_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.optimize_lhs_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(optimize_frame, 
                       text="Optimize LHS (SMT - better space filling)",
                       variable=self.optimize_lhs_var).pack(side=tk.LEFT)
        
        if not HAS_SMT:
            ttk.Label(optimize_frame, text="(requires SMT)", 
                     foreground="orange", font=("TkDefaultFont", 8)).pack(side=tk.LEFT, padx=5)
        
        # === Fractional Factorial controls ===
        self.fractional_controls = ttk.Frame(self.design_controls_frame)
        
        resolution_frame = ttk.Frame(self.fractional_controls)
        resolution_frame.pack(fill=tk.X)
        
        ttk.Label(resolution_frame, text="Resolution:").pack(side=tk.LEFT, padx=(0, 5))
        self.resolution_var = tk.StringVar(value="IV")
        resolution_combo = ttk.Combobox(resolution_frame, textvariable=self.resolution_var,
                                       values=["III", "IV", "V"], state="readonly", width=8)
        resolution_combo.pack(side=tk.LEFT, padx=(0, 10))
        resolution_combo.current(1)
        
        ttk.Label(resolution_frame, text="(III=screening, IV=interactions, V=full)", 
                 font=("TkDefaultFont", 8, "italic"), foreground="gray").pack(side=tk.LEFT)
        
        # === Plackett-Burman controls ===
        self.pb_controls = ttk.Frame(self.design_controls_frame)
        
        pb_info = ttk.Label(self.pb_controls, 
                           text="Automatically determines optimal run count based on number of factors",
                           font=("TkDefaultFont", 8, "italic"), foreground="gray")
        pb_info.pack(anchor="w")
        
        # === Central Composite controls ===
        self.ccd_controls = ttk.Frame(self.design_controls_frame)
        
        alpha_frame = ttk.Frame(self.ccd_controls)
        alpha_frame.pack(fill=tk.X)
        
        ttk.Label(alpha_frame, text="Design Type:").pack(side=tk.LEFT, padx=(0, 5))
        self.ccd_type_var = tk.StringVar(value="faced")
        ccd_combo = ttk.Combobox(alpha_frame, textvariable=self.ccd_type_var,
                                values=["faced", "inscribed", "circumscribed"], 
                                state="readonly", width=15)
        ccd_combo.pack(side=tk.LEFT, padx=(0, 10))
        ccd_combo.current(0)
        
        ttk.Label(alpha_frame, text="(faced=practical, inscribed=scaled, circumscribed=extended)", 
                 font=("TkDefaultFont", 8, "italic"), foreground="gray").pack(side=tk.LEFT)
        
        # === Box-Behnken controls ===
        self.bb_controls = ttk.Frame(self.design_controls_frame)
        
        bb_info = ttk.Label(self.bb_controls, 
                           text="Requires 3+ factors. Does not test extreme corner combinations.",
                           font=("TkDefaultFont", 8, "italic"), foreground="gray")
        bb_info.pack(anchor="w")
        
        # Hide all controls initially
        self._hide_all_design_controls()
        
        # ========== END NEW SECTION ==========
        
        # Bottom panel: Export Settings
        bottom_panel = ttk.LabelFrame(main_container, text="Export", padding=8)
        bottom_panel.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        bottom_panel.columnconfigure(1, weight=1)
        
        # Final volume
        ttk.Label(bottom_panel, text="Final Volume (µL):").grid(row=0, column=0, sticky="w")
        self.final_vol_var = tk.StringVar(value="100")
        vcmd = (self.register(validate_single_numeric_input), '%d', '%S', '%P')
        ttk.Entry(bottom_panel, textvariable=self.final_vol_var, width=12,
                 validate='key', validatecommand=vcmd).grid(row=0, column=1, sticky="w", padx=5)
        
        # Status display
        status_frame = ttk.Frame(bottom_panel)
        status_frame.grid(row=0, column=2, sticky="e")
        
        ttk.Label(status_frame, text="Combinations:").pack(side=tk.LEFT, padx=(0, 3))
        self.combo_var = tk.StringVar(value="0")
        combo_label = ttk.Label(status_frame, textvariable=self.combo_var, 
                               font=("TkDefaultFont", 10, "bold"), foreground="#2196F3")
        combo_label.pack(side=tk.LEFT, padx=(0, 8))
        
        ttk.Label(status_frame, text="Plates:").pack(side=tk.LEFT, padx=(0, 3))
        self.plates_var = tk.StringVar(value="0")
        plates_label = ttk.Label(status_frame, textvariable=self.plates_var, 
                                font=("TkDefaultFont", 10, "bold"), foreground="#4CAF50")
        plates_label.pack(side=tk.LEFT)
        
        # Export button
        ttk.Button(bottom_panel, text="Export Design",
                  command=self._export_both).grid(row=1, column=0, columnspan=3, pady=(8, 0))
    
    def _hide_all_design_controls(self):
        """Hide all design-specific control frames"""
        self.lhs_controls.pack_forget()
        self.fractional_controls.pack_forget()
        self.pb_controls.pack_forget()
        self.ccd_controls.pack_forget()
        self.bb_controls.pack_forget()
    
    def _on_design_type_changed(self):
        """Handle design type dropdown change"""
        # Get the internal value from the display text
        display_text = self.design_type_var.get()
        if display_text in self.design_map:
            design_type = self.design_map[display_text]
        else:
            design_type = display_text
        
        # Hide all controls first
        self._hide_all_design_controls()
        
        # Show relevant controls
        if design_type == "lhs":
            self.lhs_controls.pack(fill=tk.X, pady=(5, 0))
        elif design_type == "fractional":
            self.fractional_controls.pack(fill=tk.X, pady=(5, 0))
        elif design_type == "plackett_burman":
            self.pb_controls.pack(fill=tk.X, pady=(5, 0))
        elif design_type == "central_composite":
            self.ccd_controls.pack(fill=tk.X, pady=(5, 0))
        elif design_type == "box_behnken":
            self.bb_controls.pack(fill=tk.X, pady=(5, 0))
        
        # Update combination count
        self._update_display()
    
    def _add_custom_factor(self):
        """Add a completely custom factor (e.g., KCl, MgCl2, etc.) with units"""
        # Ask for custom factor name with units
        dialog = tk.Toplevel(self)
        dialog.title("Add Custom Factor")
        dialog.geometry("450x250")
        dialog.transient(self)
        dialog.grab_set()
        
        result = {"name": None}
        
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Enter factor name:", 
                 font=("TkDefaultFont", 10, "bold")).pack(anchor="w", pady=(0, 5))
        
        ttk.Label(frame, text="Examples: KCl, MgCl₂, Enzyme X, Temperature, etc.",
                 font=("TkDefaultFont", 9, "italic"),
                 foreground="gray").pack(anchor="w", pady=(0, 10))
        
        name_var = tk.StringVar()
        name_entry = ttk.Entry(frame, textvariable=name_var, font=("TkDefaultFont", 11))
        name_entry.pack(fill=tk.X, pady=(0, 15))
        name_entry.focus()
        
        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showerror("Invalid Name", "Please enter a factor name.")
                return
            
            # Check if already exists
            factors = self.model.get_factors()
            if name in factors:
                messagebox.showerror("Duplicate", f"Factor '{name}' already exists.")
                return
            
            result["name"] = name
            dialog.destroy()
        
        name_entry.bind('<Return>', lambda e: save())
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(20, 0))
        
        ttk.Button(btn_frame, text="Cancel", 
                  command=dialog.destroy).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Continue", 
                  command=save).pack(side=tk.RIGHT, padx=2)
        
        self.wait_window(dialog)
        
        # If a name was provided, open factor editor
        if result["name"]:
            factor_name = result["name"]
            editor = FactorEditDialog(self, factor_name)
            self.wait_window(editor)
            
            if editor.result_levels:
                try:
                    self.model.add_factor(factor_name, editor.result_levels, editor.result_stock)
                    self._update_display()
                except ValueError as e:
                    messagebox.showerror("Invalid Factor",
                        f"Could not add factor '{factor_name}'.\n\n"
                        f"Error: {str(e)}\n\n"
                        f"Check that the factor name and levels are valid.")
    
    def _quick_add_factor(self):
        """Quick add from available factors"""
        selection = self.available_listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        display_name = self.available_listbox.get(idx)
        
        # Skip category headers (start with ---) and empty lines
        if display_name.startswith("---") or display_name.strip() == "":
            return
        
        # Remove leading spaces from indented items
        display_name = display_name.strip()
        
        # Find the key
        factor_key = None
        for key, disp in AVAILABLE_FACTORS.items():
            if disp == display_name:
                factor_key = key
                break
        
        if not factor_key:
            return
        
        # Check if already added
        factors = self.model.get_factors()
        if factor_key in factors:
            messagebox.showinfo("Already Added", 
                f"Factor '{display_name}' is already in your design.")
            return
        
        # Open editor
        editor = FactorEditDialog(self, factor_key)
        self.wait_window(editor)
        
        if editor.result_levels:
            try:
                self.model.add_factor(factor_key, editor.result_levels, editor.result_stock)
                self._update_display()
            except ValueError as e:
                messagebox.showerror("Invalid Factor",
                    f"Could not add factor.\n\n"
                    f"Error: {str(e)}\n\n"
                    f"Check that the factor levels are valid and within the stock concentration.")
    
    def _edit_factor(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("No Selection", "Please select a factor to edit.")
            return
        
        item = selection[0]
        factor_name = self.tree.item(item, "values")[0]
        
        # Find original key
        factor_key = None
        for key, disp in AVAILABLE_FACTORS.items():
            if disp == factor_name:
                factor_key = key
                break
        
        if not factor_key:
            # Custom factor
            factor_key = factor_name
        
        factors = self.model.get_factors()
        if factor_key not in factors:
            return
        
        existing_levels = factors[factor_key]
        stock_conc = self.model.get_stock_conc(factor_key)
        
        editor = FactorEditDialog(self, factor_key, existing_levels, stock_conc)
        self.wait_window(editor)
        
        if editor.result_levels:
            try:
                self.model.update_factor(factor_key, editor.result_levels, editor.result_stock)
                self._update_display()
            except ValueError as e:
                messagebox.showerror("Invalid Factor Update",
                    f"Could not update factor '{factor_key}'.\n\n"
                    f"Error: {str(e)}\n\n"
                    f"Check that the factor levels are valid and within the stock concentration.")
    
    def _delete_factor(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("No Selection", "Please select a factor to delete.")
            return
        
        item = selection[0]
        factor_name = self.tree.item(item, "values")[0]
        
        # Find original key
        factor_key = None
        for key, disp in AVAILABLE_FACTORS.items():
            if disp == factor_name:
                factor_key = key
                break
        
        if not factor_key:
            factor_key = factor_name
        
        result = messagebox.askyesno("Confirm Delete", 
            f"Are you sure you want to delete '{factor_name}'?")
        
        if result:
            self.model.remove_factor(factor_key)
            self._update_display()
    
    def _clear_all(self):
        if not self.model.get_factors():
            messagebox.showinfo("No Factors", "There are no factors to clear.")
            return
        
        result = messagebox.askyesno("Confirm Clear All", 
            "Are you sure you want to clear all factors?")
        
        if result:
            self.model.clear()
            self._update_display()
    
    def _update_display(self):
        """Update the treeview and combination counter"""
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Populate tree
        factors = self.model.get_factors()
        for factor_key, levels in factors.items():
            display_name = AVAILABLE_FACTORS.get(factor_key, factor_key)
            levels_str = ", ".join(str(l) for l in levels[:5])
            if len(levels) > 5:
                levels_str += "..."
            
            count = len(levels)
            stock = self.model.get_stock_conc(factor_key)
            stock_str = f"{stock}" if stock else "N/A"
            
            self.tree.insert("", tk.END, values=(display_name, levels_str, count, stock_str))
        
        # Update combination counter based on design type
        display_text = self.design_type_var.get()
        if display_text in self.design_map:
            design_type = self.design_map[display_text]
        else:
            design_type = display_text
        
        if not factors:
            self.combo_var.set("0")
            self.plates_var.set("0")
            return
        
        n_factors = len(factors)
        
        try:
            if design_type == "full_factorial":
                total = self.model.total_combinations()
                self.combo_var.set(str(total))
            
            elif design_type == "lhs":
                sample_size = self.sample_size_var.get()
                self.combo_var.set(f"{sample_size} (LHS)")
                total = sample_size
            
            elif design_type == "fractional":
                resolution = self.resolution_var.get()
                if resolution == "III":
                    runs = 2 ** (n_factors - max(1, n_factors - 4))
                elif resolution == "IV":
                    runs = 2 ** (n_factors - max(0, n_factors - 5))
                else:  # V
                    runs = 2 ** (n_factors - max(0, n_factors - 6))
                self.combo_var.set(f"~{runs} (Frac)")
                total = runs
            
            elif design_type == "plackett_burman":
                runs = ((n_factors // 4) + 1) * 4
                if runs < 12:
                    runs = 12
                self.combo_var.set(f"~{runs} (PB)")
                total = runs
            
            elif design_type == "central_composite":
                runs = (2 ** n_factors) + (2 * n_factors) + 4
                self.combo_var.set(f"~{runs} (CCD)")
                total = runs
            
            elif design_type == "box_behnken":
                if n_factors >= 3:
                    runs = 2 * n_factors * (n_factors - 1) + 3
                    self.combo_var.set(f"~{runs} (BB)")
                    total = runs
                else:
                    self.combo_var.set("N/A (need 3+)")
                    total = 0
            
            else:
                total = 0
                self.combo_var.set("?")
            
            # Calculate plates
            if total > 0:
                plates = (total + 95) // 96
                self.plates_var.set(str(plates))
            else:
                self.plates_var.set("0")
        
        except Exception:
            self.combo_var.set("?")
            self.plates_var.set("?")
    
    def _generate_well_position(self, idx: int) -> Tuple[int, str]:
        """Generate 96-well plate and well position from index (row-major order)"""
        plate_num = (idx // 96) + 1
        well_idx = idx % 96

        # Row-major order: A1, A2, A3...A12, B1, B2...B12
        row = chr(ord('A') + (well_idx // 12))  # 0-7 for A-H
        col = (well_idx % 12) + 1               # 1-12
        well_pos = f"{row}{col}"

        return plate_num, well_pos
    
    def _convert_96_to_384_well(self, plate_num: int, well_96: str) -> str:
        """Convert 96-well to 384-well position using WellMapper"""
        from core.well_mapper import WellMapper

        # Use WellMapper to convert 96-well to 384-well position
        return WellMapper.convert_96_to_384_well(plate_num, well_96)
    
    def _filter_categorical_combinations(self, combinations: List[Tuple], factor_names: List[str]) -> List[Tuple]:
        """Filter out illogical categorical-concentration pairings

        Rules:
        - If detergent is None/0, detergent_concentration must be 0
        - If detergent has a value (Triton, etc.), detergent_concentration must be > 0
        - Same logic for reducing_agent and buffer pH
        """
        filtered = []

        for combo in combinations:
            row_dict = {factor_names[i]: combo[i] for i in range(len(factor_names))}
            valid = True

            # Check detergent-concentration pairing
            if "detergent" in row_dict and "detergent_concentration" in row_dict:
                det = str(row_dict["detergent"]).strip()
                det_conc = float(row_dict["detergent_concentration"])

                # If detergent is None/empty, concentration must be 0
                if det.lower() in ['none', '0', '', 'nan']:
                    if det_conc != 0:
                        valid = False
                # If detergent has a value, concentration must be > 0
                else:
                    if det_conc == 0:
                        valid = False

            # Check reducing_agent-concentration pairing
            if "reducing_agent" in row_dict and "reducing_agent_concentration" in row_dict:
                agent = str(row_dict["reducing_agent"]).strip()
                agent_conc = float(row_dict["reducing_agent_concentration"])

                # If agent is None/empty, concentration must be 0
                if agent.lower() in ['none', '0', '', 'nan']:
                    if agent_conc != 0:
                        valid = False
                # If agent has a value, concentration must be > 0
                else:
                    if agent_conc == 0:
                        valid = False

            # Check buffer pH-concentration pairing
            if "buffer pH" in row_dict and "buffer_concentration" in row_dict:
                # pH is always defined if present, so concentration should always be > 0
                buffer_conc = float(row_dict["buffer_concentration"])
                if buffer_conc == 0:
                    valid = False

            if valid:
                filtered.append(combo)

        return filtered

    def _generate_lhs_design(self, factors: Dict[str, List[str]], n_samples: int) -> List[Tuple]:
        """Generate Latin Hypercube Sampling design using pyDOE3 or SMT

        Args:
            factors: Dictionary of factor names to level lists
            n_samples: Number of samples to generate

        Returns:
            List of tuples representing combinations
        """
        factor_names = list(factors.keys())
        n_factors = len(factor_names)

        # Identify categorical factors (non-numeric)
        categorical_factors = ["buffer pH", "detergent", "reducing_agent"]

        # Separate numeric and categorical factors
        numeric_factor_names = []
        categorical_factor_names = []
        for fn in factor_names:
            if fn in categorical_factors:
                categorical_factor_names.append(fn)
            else:
                numeric_factor_names.append(fn)

        n_numeric = len(numeric_factor_names)

        # Check if user wants optimized LHS with SMT
        use_smt = self.optimize_lhs_var.get() and HAS_SMT

        if n_numeric > 0:
            if use_smt:
                # Use SMT for optimized LHS with maximin criterion (numeric factors only)
                xlimits = []
                for factor_name in numeric_factor_names:
                    levels = factors[factor_name]
                    # Convert to float and get min/max range
                    numeric_levels = [float(lv) for lv in levels]
                    xlimits.append([min(numeric_levels), max(numeric_levels)])

                xlimits = np.array(xlimits)

                # Generate optimized LHS
                sampling = LHS(xlimits=xlimits, criterion='maximin')
                lhs_design = sampling(n_samples)

                # Map continuous values to discrete levels
                numeric_combinations = []
                for sample in lhs_design:
                    combo = []
                    for i, factor_name in enumerate(numeric_factor_names):
                        levels = factors[factor_name]
                        numeric_levels = [float(lv) for lv in levels]
                        # Find closest level to the continuous value
                        value = sample[i]
                        closest_idx = min(range(len(numeric_levels)),
                                        key=lambda j: abs(numeric_levels[j] - value))
                        combo.append(levels[closest_idx])
                    numeric_combinations.append(combo)

            else:
                # Use standard pyDOE3 LHS (numeric factors only)
                if not HAS_PYDOE3:
                    raise ImportError("pyDOE3 is required for Latin Hypercube Sampling. "
                                    "Install with: pip install pyDOE3")

                # Generate LHS design in [0,1] hypercube
                lhs_design = pyDOE3.lhs(n=n_numeric, samples=n_samples, criterion='center')

                # Map to actual factor levels
                numeric_combinations = []
                for sample in lhs_design:
                    combo = []
                    for i, factor_name in enumerate(numeric_factor_names):
                        levels = factors[factor_name]
                        # Map [0,1] to level index
                        level_idx = int(sample[i] * len(levels))
                        level_idx = min(level_idx, len(levels) - 1)  # Ensure within bounds
                        combo.append(levels[level_idx])
                    numeric_combinations.append(combo)
        else:
            # No numeric factors, create empty combinations
            numeric_combinations = [[] for _ in range(n_samples)]

        # Handle categorical factors - cycle through combinations evenly
        if categorical_factor_names:
            # Get all combinations of categorical factors
            categorical_level_lists = [factors[fn] for fn in categorical_factor_names]
            all_cat_combos = list(itertools.product(*categorical_level_lists))

            # Distribute evenly across samples
            categorical_combinations = []
            for i in range(n_samples):
                # Cycle through categorical combinations
                cat_idx = i % len(all_cat_combos)
                categorical_combinations.append(list(all_cat_combos[cat_idx]))
        else:
            categorical_combinations = [[] for _ in range(n_samples)]

        # Combine numeric and categorical factors in original order
        combinations = []
        for i in range(n_samples):
            combo = []
            numeric_idx = 0
            categorical_idx = 0
            for fn in factor_names:
                if fn in categorical_factors:
                    combo.append(categorical_combinations[i][categorical_idx])
                    categorical_idx += 1
                else:
                    combo.append(numeric_combinations[i][numeric_idx])
                    numeric_idx += 1
            combinations.append(tuple(combo))

        return combinations
    
    def _generate_fractional_factorial(self, factors: Dict[str, List[str]], resolution: str) -> List[Tuple]:
        """Generate 2-level fractional factorial design using pyDOE3
        
        Args:
            factors: Dictionary of factor names to level lists (must have exactly 2 levels each)
            resolution: Resolution level ('III', 'IV', or 'V')
        
        Returns:
            List of tuples representing combinations
        """
        if not HAS_PYDOE3:
            raise ImportError("pyDOE3 is required for Fractional Factorial designs. "
                            "Install with: pip install pyDOE3")
        
        factor_names = list(factors.keys())
        n_factors = len(factor_names)
        
        # Validate that all factors have exactly 2 levels
        for fn in factor_names:
            if len(factors[fn]) != 2:
                raise ValueError(f"Fractional Factorial requires exactly 2 levels per factor. "
                               f"Factor '{AVAILABLE_FACTORS.get(fn, fn)}' has {len(factors[fn])} levels.")
        
        # Build generator string based on resolution
        # For pyDOE3.fracfact, we use notation like "a b c ab" for 4 factors
        if resolution == "III":
            # Resolution III: Main effects aliased with 2-factor interactions
            gen_string = " ".join([chr(97 + i) for i in range(n_factors)])  # "a b c d..."
        elif resolution == "IV":
            # Resolution IV: Main effects clear, 2-factor interactions may be aliased
            if n_factors <= 4:
                gen_string = " ".join([chr(97 + i) for i in range(n_factors)])
            else:
                # Add some interaction terms
                gen_string = " ".join([chr(97 + i) for i in range(min(n_factors, 6))])
        elif resolution == "V":
            # Resolution V: Main effects and 2-factor interactions clear
            gen_string = " ".join([chr(97 + i) for i in range(n_factors)])
        else:
            raise ValueError(f"Invalid resolution: {resolution}. Must be III, IV, or V.")
        
        # Generate fractional factorial design
        design = pyDOE3.fracfact(gen_string)
        
        # Convert from -1/+1 coding to actual factor levels
        combinations = []
        for row in design:
            combo = []
            for i, factor_name in enumerate(factor_names):
                levels = factors[factor_name]
                # Map -1 to low level (index 0), +1 to high level (index 1)
                level_idx = 0 if row[i] < 0 else 1
                combo.append(levels[level_idx])
            combinations.append(tuple(combo))
        
        return combinations
    
    def _generate_plackett_burman(self, factors: Dict[str, List[str]]) -> List[Tuple]:
        """Generate Plackett-Burman screening design using pyDOE3
        
        Args:
            factors: Dictionary of factor names to level lists (must have exactly 2 levels each)
        
        Returns:
            List of tuples representing combinations
        """
        if not HAS_PYDOE3:
            raise ImportError("pyDOE3 is required for Plackett-Burman designs. "
                            "Install with: pip install pyDOE3")
        
        factor_names = list(factors.keys())
        n_factors = len(factor_names)
        
        # Validate that all factors have exactly 2 levels
        for fn in factor_names:
            if len(factors[fn]) != 2:
                raise ValueError(f"Plackett-Burman requires exactly 2 levels per factor. "
                               f"Factor '{AVAILABLE_FACTORS.get(fn, fn)}' has {len(factors[fn])} levels.")
        
        # Generate Plackett-Burman design
        design = pyDOE3.pbdesign(n_factors)
        
        # Convert from -1/+1 coding to actual factor levels
        combinations = []
        for row in design:
            combo = []
            for i, factor_name in enumerate(factor_names):
                levels = factors[factor_name]
                # Map -1 to low level (index 0), +1 to high level (index 1)
                level_idx = 0 if row[i] < 0 else 1
                combo.append(levels[level_idx])
            combinations.append(tuple(combo))
        
        return combinations
    
    def _generate_central_composite(self, factors: Dict[str, List[str]], ccd_type: str) -> List[Tuple]:
        """Generate Central Composite Design using pyDOE3

        Args:
            factors: Dictionary of factor names to level lists
            ccd_type: Type of CCD ('faced', 'inscribed', or 'circumscribed')

        Returns:
            List of tuples representing combinations
        """
        if not HAS_PYDOE3:
            raise ImportError("pyDOE3 is required for Central Composite designs. "
                            "Install with: pip install pyDOE3")

        factor_names = list(factors.keys())

        # Identify categorical factors
        categorical_factors = ["buffer pH", "detergent", "reducing_agent"]

        # Separate numeric and categorical factors
        numeric_factor_names = []
        categorical_factor_names = []
        for fn in factor_names:
            if fn in categorical_factors:
                categorical_factor_names.append(fn)
            else:
                numeric_factor_names.append(fn)

        n_numeric = len(numeric_factor_names)

        if n_numeric == 0:
            raise ValueError("Central Composite Design requires at least one numeric factor.")

        # Determine face parameter based on type
        if ccd_type == "faced":
            face = "faced"
        elif ccd_type == "inscribed":
            face = "inscribed"
        elif ccd_type == "circumscribed":
            face = "circumscribed"
        else:
            face = "faced"  # Default

        # Generate CCD design for numeric factors only
        design = pyDOE3.ccdesign(n_numeric, center=(4, 4), alpha='orthogonal', face=face)

        # Map coded values to actual factor levels (numeric only)
        numeric_combinations = []
        for row in design:
            combo = []
            for i, factor_name in enumerate(numeric_factor_names):
                levels = factors[factor_name]
                numeric_levels = sorted([float(lv) for lv in levels])

                # Map coded value to actual level
                if len(numeric_levels) >= 3:
                    min_val = numeric_levels[0]
                    max_val = numeric_levels[-1]
                    center_val = numeric_levels[len(numeric_levels)//2]
                else:
                    min_val = numeric_levels[0]
                    max_val = numeric_levels[-1]
                    center_val = (min_val + max_val) / 2

                coded_val = row[i]

                # Map coded value to actual value
                if abs(coded_val) < 0.1:  # Center point (0)
                    actual_val = center_val
                elif coded_val < -0.5:  # Low level (-1)
                    actual_val = min_val
                elif coded_val > 0.5:  # High level (+1)
                    actual_val = max_val
                else:
                    # Axial point (alpha)
                    actual_val = center_val + coded_val * (max_val - min_val) / 2

                # Find closest existing level or use computed value
                if len(levels) > 2:
                    closest = min(levels, key=lambda x: abs(float(x) - actual_val))
                    combo.append(closest)
                else:
                    combo.append(str(round(actual_val, 2)))

            numeric_combinations.append(combo)

        # Handle categorical factors - use all combinations
        if categorical_factor_names:
            categorical_level_lists = [factors[fn] for fn in categorical_factor_names]
            categorical_combos = list(itertools.product(*categorical_level_lists))

            # Combine numeric CCD with all categorical combinations
            all_combinations = []
            for num_combo in numeric_combinations:
                for cat_combo in categorical_combos:
                    # Merge in original factor order
                    combo = []
                    num_idx = 0
                    cat_idx = 0
                    for fn in factor_names:
                        if fn in categorical_factors:
                            combo.append(cat_combo[cat_idx])
                            cat_idx += 1
                        else:
                            combo.append(num_combo[num_idx])
                            num_idx += 1
                    all_combinations.append(tuple(combo))
            return all_combinations
        else:
            # No categorical factors, return numeric combinations in order
            return [tuple(combo) for combo in numeric_combinations]
    
    def _generate_box_behnken(self, factors: Dict[str, List[str]]) -> List[Tuple]:
        """Generate Box-Behnken design using pyDOE3

        Args:
            factors: Dictionary of factor names to level lists (requires 3+ factors)

        Returns:
            List of tuples representing combinations
        """
        if not HAS_PYDOE3:
            raise ImportError("pyDOE3 is required for Box-Behnken designs. "
                            "Install with: pip install pyDOE3")

        factor_names = list(factors.keys())

        # Identify categorical factors
        categorical_factors = ["buffer pH", "detergent", "reducing_agent"]

        # Separate numeric and categorical factors
        numeric_factor_names = []
        categorical_factor_names = []
        for fn in factor_names:
            if fn in categorical_factors:
                categorical_factor_names.append(fn)
            else:
                numeric_factor_names.append(fn)

        n_numeric = len(numeric_factor_names)
        n_total = len(factor_names)

        if n_total < 3:
            raise ValueError("Box-Behnken design requires at least 3 factors. "
                           f"You have {n_total} factor(s).")

        if n_numeric < 3:
            raise ValueError("Box-Behnken design requires at least 3 numeric factors. "
                           f"You have {n_numeric} numeric factor(s).")

        # Generate Box-Behnken design for numeric factors only
        design = pyDOE3.bbdesign(n_numeric, center=3)

        # Map coded values to actual factor levels (numeric only)
        numeric_combinations = []
        for row in design:
            combo = []
            for i, factor_name in enumerate(numeric_factor_names):
                levels = factors[factor_name]
                numeric_levels = sorted([float(lv) for lv in levels])

                # Map coded value to actual level
                if len(numeric_levels) >= 3:
                    min_val = numeric_levels[0]
                    max_val = numeric_levels[-1]
                    center_val = numeric_levels[len(numeric_levels)//2]
                else:
                    min_val = numeric_levels[0]
                    max_val = numeric_levels[-1]
                    center_val = (min_val + max_val) / 2

                coded_val = row[i]

                # Map to closest level
                if abs(coded_val) < 0.1:  # Center (0)
                    actual_val = center_val
                elif coded_val < -0.5:  # Low (-1)
                    actual_val = min_val
                else:  # High (+1)
                    actual_val = max_val

                # Find closest existing level or use computed value
                if len(levels) >= 3:
                    closest = min(levels, key=lambda x: abs(float(x) - actual_val))
                    combo.append(closest)
                else:
                    combo.append(str(round(actual_val, 2)))

            numeric_combinations.append(combo)

        # Handle categorical factors - use all combinations
        if categorical_factor_names:
            categorical_level_lists = [factors[fn] for fn in categorical_factor_names]
            categorical_combos = list(itertools.product(*categorical_level_lists))

            # Combine numeric BB with all categorical combinations
            all_combinations = []
            for num_combo in numeric_combinations:
                for cat_combo in categorical_combos:
                    # Merge in original factor order
                    combo = []
                    num_idx = 0
                    cat_idx = 0
                    for fn in factor_names:
                        if fn in categorical_factors:
                            combo.append(cat_combo[cat_idx])
                            cat_idx += 1
                        else:
                            combo.append(num_combo[num_idx])
                            num_idx += 1
                    all_combinations.append(tuple(combo))
            return all_combinations
        else:
            # No categorical factors, return numeric combinations in order
            return [tuple(combo) for combo in numeric_combinations]

    def _extract_unique_categorical_values(self, factors: Dict[str, List[str]],
                                           factor_name: str,
                                           skip_none: bool = False) -> List[str]:
        """
        Extract unique values for categorical factors.

        Args:
            factors: Dictionary of factor names to levels
            factor_name: Name of the categorical factor
            skip_none: If True, skip None/0/empty values

        Returns:
            Sorted list of unique values
        """
        if factor_name not in factors:
            return []

        unique_values = set()
        for val in factors[factor_name]:
            val_str = str(val).strip()

            if skip_none:
                # Skip empty, None, or 0 values
                if val_str and val_str.lower() not in ['none', '0', 'nan', '']:
                    unique_values.add(val_str)
            else:
                unique_values.add(val_str)

        return sorted(unique_values)

    def _build_excel_headers(self, factor_names: List[str]) -> List[str]:
        """
        Build Excel headers with proper ordering of categorical factors.

        Args:
            factor_names: List of factor names in the design

        Returns:
            List of column headers for Excel export
        """
        excel_headers = ["ID", "Plate_96", "Well_96", "Well_384", "Source", "Batch"]

        # Track which factors we've already added
        added_factors = set()

        # Add factors in order, but group categorical with their concentrations
        for fn in factor_names:
            if fn in added_factors:
                continue

            # Add the factor
            excel_headers.append(AVAILABLE_FACTORS.get(fn, fn))
            added_factors.add(fn)

            # If it's a categorical factor with a paired concentration, add that next
            if fn in self.CATEGORICAL_PAIRS:
                paired_factor = self.CATEGORICAL_PAIRS[fn]
                if paired_factor in factor_names and paired_factor not in added_factors:
                    excel_headers.append(AVAILABLE_FACTORS.get(paired_factor, paired_factor))
                    added_factors.add(paired_factor)

        # Add Response column for TM data entry
        excel_headers.append("Response")

        return excel_headers

    def _build_volume_headers(self, factor_names: List[str],
                             buffer_ph_values: List[str],
                             detergent_values: List[str],
                             reducing_agent_values: List[str]) -> List[str]:
        """
        Build volume headers for Opentrons CSV.

        Args:
            factor_names: List of factor names
            buffer_ph_values: List of unique buffer pH values
            detergent_values: List of unique detergent types
            reducing_agent_values: List of unique reducing agent types

        Returns:
            List of column headers for volume CSV
        """
        volume_headers = ["ID"]

        # Add buffer pH columns (one column per pH value)
        for ph in buffer_ph_values:
            volume_headers.append(f"buffer_{ph}")

        # Add detergent columns (one column per detergent type)
        for det in detergent_values:
            # Clean up name for column header (lowercase, replace spaces/hyphens with underscores)
            det_clean = det.replace(' ', '_').replace('-', '_').lower()
            volume_headers.append(det_clean)

        # Add reducing agent columns (one column per reducing agent type)
        for agent in reducing_agent_values:
            # Clean up name for column header (lowercase, replace spaces/hyphens with underscores)
            agent_clean = agent.replace(' ', '_').replace('-', '_').lower()
            volume_headers.append(agent_clean)

        # Add other volume headers (skip categorical factors and their concentrations)
        for factor in factor_names:
            if factor in ["buffer pH", "buffer_concentration", "detergent", "detergent_concentration",
                         "reducing_agent", "reducing_agent_concentration"]:
                continue
            volume_headers.append(factor)

        # Add water column at the end
        volume_headers.append("water")

        return volume_headers

    def _build_factorial_with_volumes(self) -> Tuple[List[str], List[List], List[str], List[List]]:
        """Build factorial design (full or LHS) and calculate volumes"""
        factors = self.model.get_factors()
        if not factors:
            raise ValueError("No factors defined.")
        
        # Get final volume
        try:
            final_vol = float(self.final_vol_var.get())
        except ValueError:
            raise ValueError("Invalid final volume value.")
        
        stock_concs = self.model.get_all_stock_concs()
        
        # Get design type and generate combinations
        display_text = self.design_type_var.get()
        if display_text in self.design_map:
            design_type = self.design_map[display_text]
        else:
            design_type = display_text
        
        factor_names = list(factors.keys())

        if design_type == "full_factorial":
            # Original full factorial logic
            level_lists = [factors[f] for f in factor_names]
            combinations = list(itertools.product(*level_lists))
            # Filter out illogical categorical-concentration pairings
            combinations = self._filter_categorical_combinations(combinations, factor_names)
        
        elif design_type == "lhs":
            # Latin Hypercube Sampling
            n_samples = self.sample_size_var.get()
            
            # Validate sample size
            if n_samples > 384:
                raise ValueError("Sample size cannot exceed 384 (4 plates of 96 wells).")
            if n_samples < 1:
                raise ValueError("Sample size must be at least 1.")
            
            combinations = self._generate_lhs_design(factors, n_samples)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "fractional":
            # 2-Level Fractional Factorial
            resolution = self.resolution_var.get()
            combinations = self._generate_fractional_factorial(factors, resolution)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "plackett_burman":
            # Plackett-Burman screening design
            combinations = self._generate_plackett_burman(factors)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "central_composite":
            # Central Composite Design
            ccd_type = self.ccd_type_var.get()
            combinations = self._generate_central_composite(factors, ccd_type)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "box_behnken":
            # Box-Behnken design
            combinations = self._generate_box_behnken(factors)
            combinations = self._filter_categorical_combinations(combinations, factor_names)
        
        else:
            raise ValueError(f"Unknown design type: {design_type}")
        
        # Extract unique values for categorical factors
        buffer_ph_values = self._extract_unique_categorical_values(factors, "buffer pH")
        detergent_values = self._extract_unique_categorical_values(factors, "detergent", skip_none=True)
        reducing_agent_values = self._extract_unique_categorical_values(factors, "reducing_agent", skip_none=True)

        # Build headers
        excel_headers = self._build_excel_headers(factor_names)
        volume_headers = self._build_volume_headers(factor_names, buffer_ph_values,
                                                     detergent_values, reducing_agent_values)
        
        # Calculate volumes
        excel_rows = []
        volume_rows = []
        
        for idx, combo in enumerate(combinations):
            row_dict = {factor_names[i]: combo[i] for i in range(len(factor_names))}
            
            plate_num, well_pos = self._generate_well_position(idx)
            well_384 = self._convert_96_to_384_well(plate_num, well_pos)
            
            # Excel row - use same order as headers
            excel_row = [idx + 1, plate_num, well_pos, well_384, design_type.upper(), 0]
            
            # Track which factors we've already added
            added_factors_row = set()
            
            # Add factor values in same order as headers
            for fn in factor_names:
                if fn in added_factors_row:
                    continue
                    
                # Add the factor value
                excel_row.append(row_dict.get(fn, ""))
                added_factors_row.add(fn)

                # If it's a categorical factor with a paired concentration, add that next
                if fn in self.CATEGORICAL_PAIRS:
                    paired_factor = self.CATEGORICAL_PAIRS[fn]
                    if paired_factor in factor_names and paired_factor not in added_factors_row:
                        excel_row.append(row_dict.get(paired_factor, ""))
                        added_factors_row.add(paired_factor)
            
            # Add empty Response column for manual data entry
            excel_row.append("")
            excel_rows.append(excel_row)
            
            # Volume calculations
            volumes = {}
            total_volume_used = 0
            
            # Handle buffer pH (categorical - one column per pH value)
            if "buffer pH" in row_dict:
                buffer_ph = str(row_dict["buffer pH"])
                for ph in buffer_ph_values:
                    volumes[f"buffer_{ph}"] = 0
                
                if "buffer_concentration" in row_dict and "buffer_concentration" in stock_concs:
                    try:
                        desired_conc = float(row_dict["buffer_concentration"])
                        buffer_stock = stock_concs["buffer_concentration"]
                        # C1*V1 = C2*V2 → V1 = (C2*V2)/C1
                        volume = (desired_conc * final_vol) / buffer_stock
                        volumes[f"buffer_{buffer_ph}"] = round(volume, 2)
                        total_volume_used += volumes[f"buffer_{buffer_ph}"]
                    except (ValueError, ZeroDivisionError):
                        volumes[f"buffer_{buffer_ph}"] = 0
            
            # Handle detergent (categorical - one column per detergent type)
            if "detergent" in row_dict:
                detergent_type = str(row_dict["detergent"]).strip()
                
                # Initialize all detergent columns to 0
                for det in detergent_values:
                    det_clean = det.replace(' ', '_').replace('-', '_').lower()
                    volumes[det_clean] = 0
                
                # Only add volume if detergent is not None/empty
                if detergent_type and detergent_type.lower() not in ['none', '0', 'nan', '']:
                    if "detergent_concentration" in row_dict and "detergent_concentration" in stock_concs:
                        try:
                            desired_conc = float(row_dict["detergent_concentration"])
                            detergent_stock = stock_concs["detergent_concentration"]
                            # C1*V1 = C2*V2 → V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / detergent_stock
                            det_clean = detergent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[det_clean] = round(volume, 2)
                            total_volume_used += volumes[det_clean]
                        except (ValueError, ZeroDivisionError):
                            pass
            
            # Handle reducing_agent (categorical - one column per reducing agent type)
            if "reducing_agent" in row_dict:
                agent_type = str(row_dict["reducing_agent"]).strip()
                
                # Initialize all reducing agent columns to 0
                for agent in reducing_agent_values:
                    agent_clean = agent.replace(' ', '_').replace('-', '_').lower()
                    volumes[agent_clean] = 0
                
                # Only add volume if reducing agent is not None/empty
                if agent_type and agent_type.lower() not in ['none', '0', 'nan', '']:
                    if "reducing_agent_concentration" in row_dict and "reducing_agent_concentration" in stock_concs:
                        try:
                            desired_conc = float(row_dict["reducing_agent_concentration"])
                            agent_stock = stock_concs["reducing_agent_concentration"]
                            # C1*V1 = C2*V2 → V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / agent_stock
                            agent_clean = agent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[agent_clean] = round(volume, 2)
                            total_volume_used += volumes[agent_clean]
                        except (ValueError, ZeroDivisionError):
                            pass
            
            # Calculate volumes for other factors (NaCl, Zinc, Glycerol, etc.)
            for factor in factor_names:
                if factor in ["buffer pH", "buffer_concentration", "detergent", "detergent_concentration",
                             "reducing_agent", "reducing_agent_concentration"]:
                    continue
                if factor in row_dict and factor in stock_concs:
                    try:
                        desired_conc = float(row_dict[factor])
                        stock_conc = stock_concs[factor]
                        volume = (desired_conc * final_vol) / stock_conc
                        volumes[factor] = round(volume, 2)
                        total_volume_used += volumes[factor]
                    except (ValueError, ZeroDivisionError):
                        volumes[factor] = 0
            
            # Calculate water to reach final volume
            water_volume = round(final_vol - total_volume_used, 2)
            volumes["water"] = water_volume
            
            # Build volume row in correct order matching headers
            volume_row = [idx + 1]  # ID first
            for h in volume_headers[1:]:  # Skip ID column in headers
                volume_row.append(volumes.get(h, 0))
            volume_rows.append(volume_row)
        
        # Check for negative water volumes
        negative_water_wells = []
        for idx, volume_row in enumerate(volume_rows):
            water_idx = volume_headers.index("water")
            water_vol = volume_row[water_idx]
            if water_vol < 0:
                well_id = excel_rows[idx][0]  # ID
                well_pos = excel_rows[idx][2]  # Well position
                negative_water_wells.append((well_id, well_pos, water_vol))
        
        if negative_water_wells:
            # Build error message for impossible designs
            error_msg = "⚠️ IMPOSSIBLE DESIGN DETECTED ⚠️\n\n"
            error_msg += f"The following wells require NEGATIVE water volumes:\n\n"
            
            # Show problematic wells
            for well_id, well_pos, water_vol in negative_water_wells[:5]:
                error_msg += f"  • Well {well_pos} (ID {well_id}): {water_vol} µL water\n"
            
            if len(negative_water_wells) > 5:
                error_msg += f"  ... and {len(negative_water_wells) - 5} more wells\n"
            
            error_msg += f"\nTotal problematic wells: {len(negative_water_wells)}\n\n"
            error_msg += "This means the sum of component volumes EXCEEDS the final volume!\n\n"
            error_msg += "Solutions:\n"
            error_msg += "  1. INCREASE stock concentrations (recommended)\n"
            error_msg += "  2. INCREASE final volume\n"
            error_msg += "  3. REDUCE desired concentration levels\n\n"
            error_msg += "Example: If stock is 50 mM and you want 100 mM,\n"
            error_msg += "you'd need to add 200 µL of stock to make 100 µL final volume.\n"
            error_msg += "This is physically impossible!"
            
            raise ValueError(error_msg)
        
        return excel_headers, excel_rows, volume_headers, volume_rows
    
    def _export_both(self):
        """Export XLSX and CSV files with single-step file dialog"""
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library", 
                               "openpyxl is required. Install with: pip install openpyxl")
            return
        
        try:
            # Validate stock concentrations
            factors = self.model.get_factors()
            stock_concs = self.model.get_all_stock_concs()
            
            missing_stocks = []
            for factor in factors.keys():
                # Skip categorical factors that don't need stock concentrations
                if factor in ["buffer pH", "detergent", "reducing_agent"]:
                    continue
                if factor not in stock_concs:
                    missing_stocks.append(AVAILABLE_FACTORS.get(factor, factor))
            
            if missing_stocks:
                messagebox.showerror("Missing Stock Concentrations",
                    f"The following factors need stock concentrations:\n\n" +
                    "\n".join(f"• {f}" for f in missing_stocks) +
                    "\n\nEdit each factor to add stock concentrations.")
                return
            
            # Build design
            excel_headers, excel_rows, volume_headers, volume_rows = self._build_factorial_with_volumes()
            
            total = len(excel_rows)
            if total > 384:
                messagebox.showerror("Too Many Combinations",
                    f"Design has {total} combinations.\n\n"
                    f"Maximum: 384 (4 plates of 96 wells)\n\n"
                    f"Please reduce factors/levels or sample size.")
                return
            
            # Single-step file save dialog with suggested name
            date_str = datetime.now().strftime('%Y%m%d')

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Factorial Design",
                initialfile="Design1.xlsx"
            )

            if not path:
                return  # User cancelled

            # Generate paths with naming convention: [UserName]_Design_[Date]
            base_path = os.path.splitext(path)[0]

            # Add standardized suffix
            xlsx_path = f"{base_path}_Design_{date_str}.xlsx"
            csv_path = f"{base_path}_Design_{date_str}_Opentron.csv"
            
            # Export XLSX
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample Tracking"
            
            # Headers
            for col_idx, header in enumerate(excel_headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # Data - convert numeric strings to numbers
            for row_idx, row_data in enumerate(excel_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    # Try to convert to number if possible
                    try:
                        numeric_value = float(value)
                        ws.cell(row=row_idx, column=col_idx, value=numeric_value)
                    except (ValueError, TypeError):
                        # Keep as string if not numeric
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # CREATE STOCK CONCENTRATIONS METADATA SHEET
            stock_sheet = wb.create_sheet(title="Stock_Concentrations")
            
            # Headers for stock concentrations sheet
            stock_headers = ["Factor Name", "Stock Value", "Unit"]
            for col_idx, header in enumerate(stock_headers, start=1):
                cell = stock_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            
            # Write stock concentration data
            row_idx = 2
            for factor_name, stock_value in stock_concs.items():
                # Get display name
                display_name = AVAILABLE_FACTORS.get(factor_name, factor_name)
                
                # Determine unit based on factor name
                if "pH" in factor_name:
                    unit = "pH"
                elif "conc" in factor_name.lower() or "salt" in factor_name.lower():
                    unit = "mM"
                elif "glycerol" in factor_name.lower() or "dmso" in factor_name.lower() or "detergent" in factor_name.lower():
                    unit = "%"
                else:
                    unit = ""  # Unknown unit
                
                # Write row
                stock_sheet.cell(row=row_idx, column=1, value=display_name)
                stock_sheet.cell(row=row_idx, column=2, value=stock_value)
                stock_sheet.cell(row=row_idx, column=3, value=unit)
                row_idx += 1
            
            # Auto-adjust stock sheet columns
            for col in stock_sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                stock_sheet.column_dimensions[col_letter].width = adjusted_width
            
            wb.save(xlsx_path)
            
            # Export CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(volume_headers)
                for vol_row in volume_rows:
                    writer.writerow(vol_row)
            
            plates = (total + 95) // 96
            
            # Get design name for display
            display_text = self.design_type_var.get()
            if display_text in self.design_map:
                design_type = self.design_map[display_text]
            else:
                design_type = display_text
            
            # Build design name with details
            if design_type == "full_factorial":
                design_name = "Full Factorial"
            elif design_type == "lhs":
                if self.optimize_lhs_var.get() and HAS_SMT:
                    design_name = "Latin Hypercube (Optimized - SMT)"
                else:
                    design_name = "Latin Hypercube (Standard - pyDOE3)"
            elif design_type == "fractional":
                resolution = self.resolution_var.get()
                design_name = f"2-Level Fractional Factorial (Resolution {resolution})"
            elif design_type == "plackett_burman":
                design_name = "Plackett-Burman Screening"
            elif design_type == "central_composite":
                ccd_type = self.ccd_type_var.get()
                design_name = f"Central Composite Design ({ccd_type})"
            elif design_type == "box_behnken":
                design_name = "Box-Behnken Design"
            else:
                design_name = display_text
            
            # Extract filenames and directory for clean message
            xlsx_filename = os.path.basename(xlsx_path)
            csv_filename = os.path.basename(csv_path)
            directory = os.path.dirname(xlsx_path)

            messagebox.showinfo("Export Complete",
                f"Files saved:\n\n"
                f"    {xlsx_filename}\n"
                f"    {csv_filename}\n\n"
                f"Location:\n"
                f"    {directory}")
        
        except Exception as e:
            messagebox.showerror("Export Failed", f"Error during export:\n\n{str(e)}")
    
    def export_excel(self):
        """Export design to Excel (called from main window menu)"""
        self._export_both()

    def export_csv(self):
        """Export design to CSV (called from main window menu)"""
        self._export_both()

    def refresh(self):
        """Refresh UI from project data (called when switching tabs)"""
        self._update_display()
