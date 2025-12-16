"""
Export Panel Mixin for Designer Tab.

This module provides export functionality for factorial designs,
including Excel and CSV export capabilities for Opentrons.
"""

import csv
import itertools
import os
from datetime import datetime
from tkinter import filedialog, messagebox
from typing import Dict, List, Tuple

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import AVAILABLE_FACTORS from utils
from utils.constants import AVAILABLE_FACTORS


class ExportPanelMixin:
    """
    Mixin class providing export functionality for factorial designs.

    This mixin provides methods for exporting experimental designs to
    Excel (.xlsx) and CSV formats suitable for Opentrons liquid handling.

    Attributes expected from parent class:
        self.project: DoEProject instance
        self.main_window: Main application window
        self.final_vol_var: tkinter variable for final volume
        self.design_type_var: tkinter variable for design type
        self.design_map: Dict mapping display names to design types
        self.sample_size_var: tkinter variable for LHS sample size
        self.resolution_var: tkinter variable for fractional factorial resolution
        self.ccd_type_var: tkinter variable for CCD type
        self.optimize_lhs_var: tkinter variable for LHS optimization
        self.CATEGORICAL_PAIRS: Dict mapping categorical factors to concentration factors

    Methods expected from parent class:
        self._filter_categorical_combinations()
        self._generate_lhs_design()
        self._generate_fractional_factorial()
        self._generate_plackett_burman()
        self._generate_central_composite()
        self._generate_box_behnken()
        self._generate_well_position()
        self._convert_96_to_384_well()
        self._update_display()
    """

    def _create_reagent_setup_guide(self, workbook, volume_headers: List[str],
                                    volume_rows: List[List], stock_concs: Dict[str, float],
                                    per_level_concs: Dict) -> None:
        """
        Create a reagent setup guide sheet showing where to place each reagent in the reservoir.

        Args:
            workbook: openpyxl Workbook object
            volume_headers: List of reagent column names from CSV
            volume_rows: List of volume data rows
            stock_concs: Dict of stock concentrations for numeric factors
            per_level_concs: Dict of per-level concentrations for categorical factors
        """
        guide_sheet = workbook.create_sheet(title="Reagent Setup Guide")

        # Cytiva 24 reservoir positions (column-wise: A1, B1, C1, D1, A2, B2...)
        reservoir_positions = []
        for col in range(1, 7):  # 6 columns
            for row in ['A', 'B', 'C', 'D']:  # 4 rows
                reservoir_positions.append(f"{row}{col}")

        # Headers - simple style matching other sheets
        headers = ["Position", "Reagent", "Stock Concentration", "Volume Needed", "Add to Reservoir (with 20% overage)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = guide_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

        # Calculate total volumes for each reagent
        row_idx = 2
        position_idx = 0  # Separate counter for reservoir positions

        for reagent_idx, reagent_name in enumerate(volume_headers):
            # Skip ID column if present
            if reagent_name.upper() == "ID":
                continue

            # Calculate total volume needed (sum across all samples)
            total_volume = 0.0
            for vol_row in volume_rows:
                try:
                    total_volume += float(vol_row[reagent_idx])
                except (ValueError, TypeError, IndexError):
                    pass

            # Add 20% overage
            volume_with_overage = round(total_volume * 1.20, 1)

            # Determine stock concentration display
            stock_display = "N/A"

            # Check if it's a buffer pH column (buffer_7.5, buffer_8.5, etc.)
            if reagent_name.startswith("buffer_"):
                # Get buffer concentration stock
                if "buffer_concentration" in stock_concs:
                    stock_display = f"{stock_concs['buffer_concentration']} mM"

            # Check if it's a categorical factor with per-level concentrations
            elif per_level_concs:
                # Detergents (c12e8, lmng, ddm, etc.)
                if "detergent" in per_level_concs:
                    for level, conc_data in per_level_concs["detergent"].items():
                        # Normalize level name for comparison
                        level_normalized = level.lower().replace(' ', '_').replace('-', '_')
                        reagent_normalized = reagent_name.lower().replace(' ', '_').replace('-', '_')
                        if level_normalized == reagent_normalized:
                            stock_display = f"{conc_data.get('stock', 'N/A')}%"
                            break

                # Reducing agents (dtt, tcep, etc.)
                if "reducing_agent" in per_level_concs:
                    for level, conc_data in per_level_concs["reducing_agent"].items():
                        level_normalized = level.lower().replace(' ', '_').replace('-', '_')
                        reagent_normalized = reagent_name.lower().replace(' ', '_').replace('-', '_')
                        if level_normalized == reagent_normalized:
                            stock_display = f"{conc_data.get('stock', 'N/A')} mM"
                            break

            # Check if it's a regular numeric factor
            if stock_display == "N/A" and reagent_name in stock_concs:
                stock_val = stock_concs[reagent_name]
                # Try to determine unit from AVAILABLE_FACTORS
                display_name = AVAILABLE_FACTORS.get(reagent_name, reagent_name)
                unit = ""
                if "(" in display_name and ")" in display_name:
                    start = display_name.rfind("(") + 1
                    end = display_name.rfind(")")
                    unit = display_name[start:end]
                stock_display = f"{stock_val} {unit}".strip()

            # Get reagent display name (capitalize first letter, replace underscores)
            reagent_display = reagent_name.replace('_', ' ').title()
            if reagent_name.startswith("buffer_"):
                ph_value = reagent_name.replace("buffer_", "")
                reagent_display = f"Buffer pH {ph_value}"

            # Get reservoir position using separate position counter
            position = reservoir_positions[position_idx] if position_idx < len(reservoir_positions) else f"Extra-{position_idx}"

            # Write row
            guide_sheet.cell(row=row_idx, column=1, value=position)
            guide_sheet.cell(row=row_idx, column=2, value=reagent_display)
            guide_sheet.cell(row=row_idx, column=3, value=stock_display)
            guide_sheet.cell(row=row_idx, column=4, value=f"{total_volume:.1f} µL")
            guide_sheet.cell(row=row_idx, column=5, value=f"{volume_with_overage:.1f} µL")

            row_idx += 1
            position_idx += 1  # Increment position counter only for actual reagents

        # Auto-adjust column widths
        for col in guide_sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            guide_sheet.column_dimensions[col_letter].width = adjusted_width

    def _extract_unique_categorical_values(self, factors: Dict[str, List[str]],
                                           factor_name: str,
                                           skip_none: bool = False) -> List[str]:
        """
        Extract unique values for categorical factors.

        Args:
            factors: Dictionary of factor names to levels
            factor_name: Name of the categorical factor
            skip_none: If True, skip None/0/empty values

        Returns:
            Sorted list of unique values
        """
        if factor_name not in factors:
            return []

        unique_values = set()
        for val in factors[factor_name]:
            val_str = str(val).strip()

            if skip_none:
                # Skip empty, None, or 0 values
                if val_str and val_str.lower() not in ['none', '0', 'nan', '']:
                    unique_values.add(val_str)
            else:
                unique_values.add(val_str)

        return sorted(unique_values)

    def _build_excel_headers(self, factor_names: List[str]) -> List[str]:
        """
        Build Excel headers with proper ordering of categorical factors.

        Args:
            factor_names: List of factor names in the design

        Returns:
            List of column headers for Excel export
        """
        excel_headers = ["ID", "Plate_96", "Well_96", "Well_384", "Source", "Batch"]

        # Track which factors we've already added
        added_factors = set()

        # Add factors in order, but group categorical with their concentrations
        for fn in factor_names:
            if fn in added_factors:
                continue

            # Add the factor
            excel_headers.append(AVAILABLE_FACTORS.get(fn, fn))
            added_factors.add(fn)

            # If it's a categorical factor with a paired concentration, add that next
            if fn in self.CATEGORICAL_PAIRS:
                paired_factor = self.CATEGORICAL_PAIRS[fn]
                if paired_factor in factor_names and paired_factor not in added_factors:
                    excel_headers.append(AVAILABLE_FACTORS.get(paired_factor, paired_factor))
                    added_factors.add(paired_factor)

        # Add Response column for TM data entry
        excel_headers.append("Response")

        return excel_headers

    def _build_volume_headers(self, factor_names: List[str],
                             buffer_ph_values: List[str],
                             detergent_values: List[str],
                             reducing_agent_values: List[str]) -> List[str]:
        """
        Build volume headers for Opentrons CSV.

        Args:
            factor_names: List of factor names
            buffer_ph_values: List of unique buffer pH values
            detergent_values: List of unique detergent types
            reducing_agent_values: List of unique reducing agent types

        Returns:
            List of column headers for volume CSV
        """
        volume_headers = ["ID"]

        # Add buffer pH columns (one column per pH value)
        for ph in buffer_ph_values:
            volume_headers.append(f"buffer_{ph}")

        # Add detergent columns (one column per detergent type)
        for det in detergent_values:
            # Clean up name for column header (lowercase, replace spaces/hyphens with underscores)
            det_clean = det.replace(' ', '_').replace('-', '_').lower()
            volume_headers.append(det_clean)

        # Add reducing agent columns (one column per reducing agent type)
        for agent in reducing_agent_values:
            # Clean up name for column header (lowercase, replace spaces/hyphens with underscores)
            agent_clean = agent.replace(' ', '_').replace('-', '_').lower()
            volume_headers.append(agent_clean)

        # Add other volume headers (skip categorical factors and their concentrations)
        for factor in factor_names:
            if factor in ["buffer pH", "buffer_concentration", "detergent", "detergent_concentration",
                         "reducing_agent", "reducing_agent_concentration"]:
                continue
            volume_headers.append(factor)

        # NOTE: Protein column is NOT included in Opentrons CSV since protein is added manually.
        # The protein volume is still calculated and subtracted from water volume.
        # To enable protein column in CSV, uncomment the line below:
        # volume_headers.append("protein")

        # Add water column at the end
        volume_headers.append("water")

        return volume_headers

    def _build_factorial_with_volumes(self) -> Tuple[List[str], List[List], List[str], List[List], List[Tuple]]:
        """
        Build factorial design (full or LHS) and calculate volumes.

        Returns:
            Tuple containing:
                - excel_headers: List of column headers for Excel export
                - excel_rows: List of data rows for Excel export
                - volume_headers: List of column headers for volume CSV
                - volume_rows: List of volume data rows for Opentrons CSV
                - negative_water_wells: List of (well_id, well_pos, water_vol) tuples for problematic wells

        Raises:
            ValueError: If no factors defined, invalid volume, or sample size issues
        """
        factors = self.project.get_factors()
        if not factors:
            raise ValueError("No factors defined.")

        # Get final volume
        try:
            final_vol = float(self.final_vol_var.get())
        except ValueError:
            raise ValueError("Invalid final volume value.")

        stock_concs = self.project.get_all_stock_concs()

        # Check for per-level concentrations (when used, detergent_concentration has placeholder levels)
        per_level_concs = self.project.get_all_per_level_concs()

        # Validate that detergent factor exists when per-level concs are used
        # Also, exclude concentration factors from design when per-level mode is active
        if "detergent" in per_level_concs and per_level_concs["detergent"]:
            if "detergent" not in factors:
                raise ValueError(
                    "Per-level concentrations are configured but 'Detergent' factor is missing.\n\n"
                    "Please add the 'Detergent' factor with the detergent types (DDM, LMNG, etc.)."
                )
            # Remove detergent_concentration from factors if it exists (per-level mode)
            if "detergent_concentration" in factors:
                factors = {k: v for k, v in factors.items() if k != "detergent_concentration"}

        if "reducing_agent" in per_level_concs and per_level_concs["reducing_agent"]:
            if "reducing_agent" not in factors:
                raise ValueError(
                    "Per-level concentrations are configured but 'Reducing Agent' factor is missing.\n\n"
                    "Please add the 'Reducing Agent' factor with the types (DTT, TCEP, etc.)."
                )
            # Remove reducing_agent_concentration from factors if it exists (per-level mode)
            if "reducing_agent_concentration" in factors:
                factors = {k: v for k, v in factors.items() if k != "reducing_agent_concentration"}

        # Get design type and generate combinations
        display_text = self.design_type_var.get()
        if display_text in self.design_map:
            design_type = self.design_map[display_text]
        else:
            design_type = display_text

        factor_names = list(factors.keys())

        if design_type == "full_factorial":
            # Original full factorial logic
            level_lists = [factors[f] for f in factor_names]
            combinations = list(itertools.product(*level_lists))
            # Filter out illogical categorical-concentration pairings
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "lhs":
            # Latin Hypercube Sampling
            n_samples = self.sample_size_var.get()

            # Validate sample size
            if n_samples > 384:
                raise ValueError("Sample size cannot exceed 384 (4 plates of 96 wells).")
            if n_samples < 1:
                raise ValueError("Sample size must be at least 1.")

            combinations = self._generate_lhs_design(factors, n_samples)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "d_optimal":
            # D-Optimal Design
            n_samples = self.d_optimal_sample_var.get()
            model_type = self.d_optimal_model_var.get()

            # Validate sample size
            if n_samples > 384:
                raise ValueError("Sample size cannot exceed 384 (4 plates of 96 wells).")
            if n_samples < 1:
                raise ValueError("Sample size must be at least 1.")

            combinations = self._generate_d_optimal_design(factors, n_samples, model_type)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "fractional":
            # 2-Level Fractional Factorial
            resolution = self.resolution_var.get()
            combinations = self._generate_fractional_factorial(factors, resolution)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "plackett_burman":
            # Plackett-Burman screening design
            combinations = self._generate_plackett_burman(factors)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "central_composite":
            # Central Composite Design
            ccd_type = self.ccd_type_var.get()
            combinations = self._generate_central_composite(factors, ccd_type)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        elif design_type == "box_behnken":
            # Box-Behnken design
            combinations = self._generate_box_behnken(factors)
            combinations = self._filter_categorical_combinations(combinations, factor_names)

        else:
            raise ValueError(f"Unknown design type: {design_type}")

        # Extract unique values for categorical factors
        buffer_ph_values = self._extract_unique_categorical_values(factors, "buffer pH")
        detergent_values = self._extract_unique_categorical_values(factors, "detergent", skip_none=True)
        reducing_agent_values = self._extract_unique_categorical_values(factors, "reducing_agent", skip_none=True)

        # Build headers
        excel_headers = self._build_excel_headers(factor_names)
        volume_headers = self._build_volume_headers(factor_names, buffer_ph_values,
                                                     detergent_values, reducing_agent_values)

        # Calculate volumes
        excel_rows = []
        volume_rows = []

        for idx, combo in enumerate(combinations):
            row_dict = {factor_names[i]: combo[i] for i in range(len(factor_names))}

            plate_num, well_pos = self._generate_well_position(idx)
            well_384 = self._convert_96_to_384_well(plate_num, well_pos)

            # Excel row - use same order as headers
            excel_row = [idx + 1, plate_num, well_pos, well_384, design_type.upper(), 0]

            # Track which factors we've already added
            added_factors_row = set()

            # Add factor values in same order as headers
            for fn in factor_names:
                if fn in added_factors_row:
                    continue

                # Add the factor value
                excel_row.append(row_dict.get(fn, ""))
                added_factors_row.add(fn)

                # If it's a categorical factor with a paired concentration, add that next
                if fn in self.CATEGORICAL_PAIRS:
                    paired_factor = self.CATEGORICAL_PAIRS[fn]
                    if paired_factor in factor_names and paired_factor not in added_factors_row:
                        excel_row.append(row_dict.get(paired_factor, ""))
                        added_factors_row.add(paired_factor)

            # Add empty Response column for manual data entry
            excel_row.append("")
            excel_rows.append(excel_row)

            # Volume calculations
            volumes = {}
            total_volume_used = 0

            # Handle buffer pH (categorical - one column per pH value)
            if "buffer pH" in row_dict:
                buffer_ph = str(row_dict["buffer pH"])
                for ph in buffer_ph_values:
                    volumes[f"buffer_{ph}"] = 0

                if "buffer_concentration" in row_dict and "buffer_concentration" in stock_concs:
                    try:
                        desired_conc = float(row_dict["buffer_concentration"])
                        buffer_stock = stock_concs["buffer_concentration"]
                        # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                        volume = (desired_conc * final_vol) / buffer_stock
                        volumes[f"buffer_{buffer_ph}"] = round(volume, 2)
                        total_volume_used += volumes[f"buffer_{buffer_ph}"]
                    except (ValueError, ZeroDivisionError):
                        volumes[f"buffer_{buffer_ph}"] = 0

            # Handle detergent (categorical - one column per detergent type)
            if "detergent" in row_dict:
                detergent_type = str(row_dict["detergent"]).strip()

                # Initialize all detergent columns to 0
                for det in detergent_values:
                    det_clean = det.replace(' ', '_').replace('-', '_').lower()
                    volumes[det_clean] = 0

                # Only add volume if detergent is not None/empty
                if detergent_type and detergent_type.lower() not in ['none', '0', 'nan', '']:
                    # Check for per-level concentrations first
                    per_level_concs = self.project.get_per_level_concs("detergent")
                    if per_level_concs and detergent_type in per_level_concs:
                        # Use per-level stock and final concentrations
                        try:
                            level_data = per_level_concs[detergent_type]
                            detergent_stock = level_data["stock"]
                            desired_conc = level_data["final"]
                            # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / detergent_stock
                            det_clean = detergent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[det_clean] = round(volume, 2)
                            total_volume_used += volumes[det_clean]
                        except (KeyError, ValueError, ZeroDivisionError):
                            pass
                    elif "detergent_concentration" in row_dict and "detergent_concentration" in stock_concs:
                        # Fall back to old behavior: single stock, desired conc from row
                        try:
                            desired_conc = float(row_dict["detergent_concentration"])
                            detergent_stock = stock_concs["detergent_concentration"]
                            # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / detergent_stock
                            det_clean = detergent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[det_clean] = round(volume, 2)
                            total_volume_used += volumes[det_clean]
                        except (ValueError, ZeroDivisionError):
                            pass

            # Handle reducing_agent (categorical - one column per reducing agent type)
            if "reducing_agent" in row_dict:
                agent_type = str(row_dict["reducing_agent"]).strip()

                # Initialize all reducing agent columns to 0
                for agent in reducing_agent_values:
                    agent_clean = agent.replace(' ', '_').replace('-', '_').lower()
                    volumes[agent_clean] = 0

                # Only add volume if reducing agent is not None/empty
                if agent_type and agent_type.lower() not in ['none', '0', 'nan', '']:
                    # Check for per-level concentrations first
                    per_level_concs = self.project.get_per_level_concs("reducing_agent")
                    if per_level_concs and agent_type in per_level_concs:
                        # Use per-level stock and final concentrations
                        try:
                            level_data = per_level_concs[agent_type]
                            agent_stock = level_data["stock"]
                            desired_conc = level_data["final"]
                            # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / agent_stock
                            agent_clean = agent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[agent_clean] = round(volume, 2)
                            total_volume_used += volumes[agent_clean]
                        except (KeyError, ValueError, ZeroDivisionError):
                            pass
                    elif "reducing_agent_concentration" in row_dict and "reducing_agent_concentration" in stock_concs:
                        # Fall back to old behavior: single stock, desired conc from row
                        try:
                            desired_conc = float(row_dict["reducing_agent_concentration"])
                            agent_stock = stock_concs["reducing_agent_concentration"]
                            # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                            volume = (desired_conc * final_vol) / agent_stock
                            agent_clean = agent_type.replace(' ', '_').replace('-', '_').lower()
                            volumes[agent_clean] = round(volume, 2)
                            total_volume_used += volumes[agent_clean]
                        except (ValueError, ZeroDivisionError):
                            pass

            # Calculate volumes for other factors (NaCl, Zinc, Glycerol, etc.)
            for factor in factor_names:
                if factor in ["buffer pH", "buffer_concentration", "detergent", "detergent_concentration",
                             "reducing_agent", "reducing_agent_concentration"]:
                    continue
                if factor in row_dict and factor in stock_concs:
                    try:
                        desired_conc = float(row_dict[factor])
                        stock_conc = stock_concs[factor]
                        volume = (desired_conc * final_vol) / stock_conc
                        volumes[factor] = round(volume, 2)
                        total_volume_used += volumes[factor]
                    except (ValueError, ZeroDivisionError):
                        volumes[factor] = 0

            # Calculate protein volume if concentrations are provided
            protein_volume = 0.0
            try:
                protein_stock_str = self.protein_stock_var.get().strip()
                protein_final_str = self.protein_final_var.get().strip()
                if protein_stock_str and protein_final_str:
                    protein_stock = float(protein_stock_str)
                    protein_final = float(protein_final_str)
                    if protein_stock > 0 and protein_final > 0:
                        # C1*V1 = C2*V2 -> V1 = (C2*V2)/C1
                        protein_volume = round((protein_final * final_vol) / protein_stock, 2)
                        total_volume_used += protein_volume
            except (ValueError, AttributeError):
                pass
            volumes["protein"] = protein_volume

            # Calculate water to reach final volume
            water_volume = round(final_vol - total_volume_used, 2)
            volumes["water"] = water_volume

            # Build volume row in correct order matching headers
            volume_row = [idx + 1]  # ID first
            for h in volume_headers[1:]:  # Skip ID column in headers
                volume_row.append(volumes.get(h, 0))
            volume_rows.append(volume_row)

        # Check for negative water volumes
        negative_water_wells = []
        for idx, volume_row in enumerate(volume_rows):
            water_idx = volume_headers.index("water")
            water_vol = volume_row[water_idx]
            if water_vol < 0:
                well_id = excel_rows[idx][0]  # ID
                well_pos = excel_rows[idx][2]  # Well position
                negative_water_wells.append((well_id, well_pos, water_vol))

        # Return all data including conflict information
        return excel_headers, excel_rows, volume_headers, volume_rows, negative_water_wells

    def _export_both(self):
        """
        Export XLSX and CSV files with single-step file dialog.

        Creates two files:
            - Excel file with sample tracking and stock concentrations
            - CSV file with volumes for Opentrons liquid handling

        The files are named with a standardized convention:
            [UserName]_Design_[Date].xlsx
            [UserName]_Design_[Date]_Opentron.csv
        """
        if not HAS_OPENPYXL:
            messagebox.showerror("Missing Library",
                               "openpyxl is required. Install with: pip install openpyxl")
            return

        try:
            # Validate stock concentrations
            factors = self.project.get_factors()
            stock_concs = self.project.get_all_stock_concs()

            missing_stocks = []
            per_level_concs = self.project.get_all_per_level_concs()

            for factor in factors.keys():
                # Skip categorical factors that don't need stock concentrations
                if factor in ["buffer pH", "detergent", "reducing_agent"]:
                    continue

                # Skip concentration factors that use per-level concentrations
                if factor == "detergent_concentration" and "detergent" in per_level_concs:
                    continue
                if factor == "reducing_agent_concentration" and "reducing_agent" in per_level_concs:
                    continue

                if factor not in stock_concs:
                    missing_stocks.append(AVAILABLE_FACTORS.get(factor, factor))

            if missing_stocks:
                messagebox.showerror("Missing Stock Concentrations",
                    f"The following factors need stock concentrations:\n\n" +
                    "\n".join(f"- {f}" for f in missing_stocks) +
                    "\n\nEdit each factor to add stock concentrations.")
                return

            # Build design
            excel_headers, excel_rows, volume_headers, volume_rows, negative_water_wells = self._build_factorial_with_volumes()

            # Check for volume conflicts
            has_conflicts = len(negative_water_wells) > 0

            total = len(excel_rows)
            if total > 384:
                messagebox.showerror("Too Many Combinations",
                    f"Design has {total} combinations.\n\n"
                    f"Maximum: 384 (4 plates of 96 wells)\n\n"
                    f"Please reduce factors/levels or sample size.")
                return

            # If there are volume conflicts, show warning BEFORE file dialog
            export_csv = True  # By default, export both files
            if has_conflicts:
                warning_msg = f"⚠️ WARNING: Volume Conflicts Detected\n\n"
                warning_msg += f"Problem: {len(negative_water_wells)} wells have negative water volumes\n\n"
                warning_msg += "What will be exported:\n"
                warning_msg += "  ✅ Excel file - WITH Warning sheet for review\n"
                warning_msg += "  ❌ CSV file - SKIPPED (cannot be used by robot)\n\n"
                warning_msg += "Fix the issues, then export again to get the CSV file.\n\n"
                warning_msg += "Do you want to export Excel with diagnostics?"

                response = messagebox.askyesno("Volume Conflicts", warning_msg, icon='warning')
                if not response:
                    return  # User cancelled
                export_csv = False  # Skip CSV export

            # Single-step file save dialog with suggested name
            date_str = datetime.now().strftime('%Y%m%d')

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Factorial Design",
                initialfile="Design1.xlsx"
            )

            if not path:
                return  # User cancelled

            # Generate paths with naming convention: [UserName]_Design_[Date]
            base_path = os.path.splitext(path)[0]

            # Add standardized suffix
            xlsx_path = f"{base_path}_Design_{date_str}.xlsx"
            csv_path = f"{base_path}_Design_{date_str}_Opentron.csv"

            # Export XLSX
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample Tracking"

            # Headers
            for col_idx, header in enumerate(excel_headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

            # Data - convert numeric strings to numbers
            for row_idx, row_data in enumerate(excel_rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    # Try to convert to number if possible
                    try:
                        numeric_value = float(value)
                        ws.cell(row=row_idx, column=col_idx, value=numeric_value)
                    except (ValueError, TypeError):
                        # Keep as string if not numeric
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            # CREATE STOCK CONCENTRATIONS METADATA SHEET
            stock_sheet = wb.create_sheet(title="Stock_Concentrations")

            # Headers for unified stock concentrations sheet
            stock_headers = ["Factor Name", "Level", "Stock Value", "Final Value", "Unit"]
            for col_idx, header in enumerate(stock_headers, start=1):
                cell = stock_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

            row_idx = 2

            # Write normal stock concentration data (Level column empty)
            for factor_name, stock_value in stock_concs.items():
                # Get display name
                display_name = AVAILABLE_FACTORS.get(factor_name, factor_name)

                # Extract unit from display name (e.g., "NaCl (mM)" -> "mM")
                unit = ""
                if "(" in display_name and ")" in display_name:
                    start = display_name.rfind("(") + 1
                    end = display_name.rfind(")")
                    unit = display_name[start:end]
                elif "pH" in display_name:
                    unit = ""  # pH is unitless

                # Write row: Factor Name | (empty) | Stock Value | (empty) | Unit
                stock_sheet.cell(row=row_idx, column=1, value=display_name)
                stock_sheet.cell(row=row_idx, column=2, value="")  # Level column empty
                stock_sheet.cell(row=row_idx, column=3, value=stock_value)
                stock_sheet.cell(row=row_idx, column=4, value="")  # Final Value empty
                stock_sheet.cell(row=row_idx, column=5, value=unit)
                row_idx += 1

            # Write per-level concentrations (same uniform format, just with Level filled in)
            all_per_level = self.project.get_all_per_level_concs()
            if all_per_level:
                # Debug: Print what we have
                print(f"[DEBUG] Per-level concentrations: {all_per_level}")

                for factor_name, level_concs in all_per_level.items():
                    if not level_concs:
                        continue

                    print(f"[DEBUG] Processing factor: {factor_name}")
                    print(f"[DEBUG] Level data: {level_concs}")

                    # Get concentration factor display name and unit
                    if factor_name == "detergent":
                        conc_factor = "detergent_concentration"
                        unit = "%"
                    elif factor_name == "reducing_agent":
                        conc_factor = "reducing_agent_concentration"
                        unit = "mM"
                    else:
                        print(f"[DEBUG] Skipping unknown factor: {factor_name}")
                        continue

                    display_name = AVAILABLE_FACTORS.get(conc_factor, conc_factor)
                    print(f"[DEBUG] Display name: {display_name}, Unit: {unit}")

                    # Write one row per level: Factor Name | Level | Stock Value | Final Value | Unit
                    for level, conc_data in level_concs.items():
                        print(f"[DEBUG] Writing level '{level}': {conc_data}")
                        stock_sheet.cell(row=row_idx, column=1, value=display_name)
                        stock_sheet.cell(row=row_idx, column=2, value=level)
                        stock_sheet.cell(row=row_idx, column=3, value=conc_data.get("stock", ""))
                        stock_sheet.cell(row=row_idx, column=4, value=conc_data.get("final", ""))
                        stock_sheet.cell(row=row_idx, column=5, value=unit)
                        row_idx += 1

            # Add protein information if provided (kept separate - added manually, not by robot)
            try:
                protein_stock_str = self.protein_stock_var.get().strip()
                protein_final_str = self.protein_final_var.get().strip()
                if protein_stock_str and protein_final_str:
                    protein_stock = float(protein_stock_str)
                    protein_final = float(protein_final_str)
                    if protein_stock > 0 and protein_final > 0:
                        final_vol = float(self.final_vol_var.get())
                        protein_vol = round((protein_final * final_vol) / protein_stock, 2)

                        # Add separator row
                        row_idx += 1

                        # Add protein info row (uniform format but highlighted)
                        # Factor Name | Level | Stock Value | Final Value | Unit
                        stock_sheet.cell(row=row_idx, column=1, value="Protein (added manually)")
                        stock_sheet.cell(row=row_idx, column=2, value="")
                        stock_sheet.cell(row=row_idx, column=3, value=protein_stock)
                        stock_sheet.cell(row=row_idx, column=4, value=protein_final)
                        stock_sheet.cell(row=row_idx, column=5, value="mg/mL")

                        # Highlight protein row in green
                        for col_idx in range(1, 6):
                            cell = stock_sheet.cell(row=row_idx, column=col_idx)
                            cell.fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")

                        row_idx += 1

                        # Add calculated volume info row
                        stock_sheet.cell(row=row_idx, column=1, value="→ Volume per well")
                        stock_sheet.cell(row=row_idx, column=2, value="")
                        stock_sheet.cell(row=row_idx, column=3, value=protein_vol)
                        stock_sheet.cell(row=row_idx, column=4, value="")
                        cell = stock_sheet.cell(row=row_idx, column=5, value="µL")

                        # Make volume row bold and slightly highlighted
                        for col_idx in range(1, 6):
                            cell = stock_sheet.cell(row=row_idx, column=col_idx)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                        cell.font = Font(bold=True)
            except (ValueError, AttributeError):
                pass

            # Auto-adjust stock sheet columns
            for col in stock_sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                stock_sheet.column_dimensions[col_letter].width = adjusted_width

            # Add Warning sheet if there are volume conflicts
            if has_conflicts:
                self._add_warning_sheet(wb, negative_water_wells, excel_rows, volume_headers, volume_rows)

            # Add Reagent Setup Guide sheet
            self._create_reagent_setup_guide(wb, volume_headers, volume_rows, stock_concs, per_level_concs)

            wb.save(xlsx_path)

            # Export CSV only if no conflicts
            if export_csv:
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(volume_headers)
                    for vol_row in volume_rows:
                        writer.writerow(vol_row)

            plates = (total + 95) // 96

            # Get design name for display
            display_text = self.design_type_var.get()
            if display_text in self.design_map:
                design_type = self.design_map[display_text]
            else:
                design_type = display_text

            # Build design name with details
            if design_type == "full_factorial":
                design_name = "Full Factorial"
            elif design_type == "lhs":
                if self.optimize_lhs_var.get() and HAS_SMT:
                    design_name = "Latin Hypercube (Optimized - SMT)"
                else:
                    design_name = "Latin Hypercube (Standard - pyDOE3)"
            elif design_type == "d_optimal":
                model_type = self.d_optimal_model_var.get()
                design_name = f"D-Optimal Design ({model_type})"
            elif design_type == "fractional":
                resolution = self.resolution_var.get()
                design_name = f"2-Level Fractional Factorial (Resolution {resolution})"
            elif design_type == "plackett_burman":
                design_name = "Plackett-Burman Screening"
            elif design_type == "central_composite":
                ccd_type = self.ccd_type_var.get()
                design_name = f"Central Composite Design ({ccd_type})"
            elif design_type == "box_behnken":
                design_name = "Box-Behnken Design"
            else:
                design_name = display_text

            # Extract filenames and directory for clean message
            xlsx_filename = os.path.basename(xlsx_path)
            directory = os.path.dirname(xlsx_path)

            if export_csv:
                csv_filename = os.path.basename(csv_path)
                messagebox.showinfo("Export Complete",
                    f"Files saved:\n\n"
                    f"    {xlsx_filename}\n"
                    f"    {csv_filename}\n\n"
                    f"Location:\n"
                    f"    {directory}")
            else:
                messagebox.showinfo("Export Complete",
                    f"Excel file saved with Warning sheet:\n\n"
                    f"    {xlsx_filename}\n\n"
                    f"Location:\n"
                    f"    {directory}\n\n"
                    f"⚠️ CSV not exported due to volume conflicts.\n"
                    f"Review the Warning sheet, fix issues, and export again.")

        except Exception as e:
            messagebox.showerror("Export Failed", f"Error during export:\n\n{str(e)}")

    def _add_warning_sheet(self, wb, negative_water_wells, excel_rows, volume_headers, volume_rows):
        """Add a Warning sheet to the workbook with details about volume conflicts"""
        ws = wb.create_sheet(title="⚠️ WARNINGS")

        # Title
        ws['A1'] = "DESIGN VALIDATION REPORT"
        ws['A1'].font = Font(bold=True, size=14, color="FF0000")
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Summary section
        ws['A4'] = "VOLUME CONFLICTS DETECTED"
        ws['A4'].font = Font(bold=True, size=12, color="FF6600")
        ws['A5'] = f"Total wells in design: {len(excel_rows)}"
        ws['A6'] = f"Problematic wells: {len(negative_water_wells)}"
        ws['A7'] = "Status: ❌ NOT READY FOR ROBOT"
        ws['A7'].font = Font(bold=True, color="FF0000")

        # Affected wells table
        ws['A9'] = "AFFECTED WELLS:"
        ws['A9'].font = Font(bold=True, size=11)

        # Table headers
        headers = ['Well', 'ID', 'Water (µL)'] + [h for h in volume_headers if h not in ['ID', 'Plate_96', 'Well_96', 'water']]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

        # Affected well data
        for row_idx, (well_id, well_pos, water_vol) in enumerate(negative_water_wells, start=11):
            # Find the corresponding row in excel_rows and volume_rows
            for idx, excel_row in enumerate(excel_rows):
                if excel_row[0] == well_id:  # Match by ID
                    volume_row = volume_rows[idx]

                    # Well position and ID
                    ws.cell(row=row_idx, column=1, value=well_pos)
                    ws.cell(row=row_idx, column=2, value=well_id)
                    ws.cell(row=row_idx, column=3, value=water_vol)

                    # Add volumes for each reagent
                    col_offset = 4
                    for vol_idx, vol_header in enumerate(volume_headers):
                        if vol_header not in ['ID', 'Plate_96', 'Well_96', 'water']:
                            ws.cell(row=row_idx, column=col_offset, value=volume_row[vol_idx])
                            col_offset += 1
                    break

        # Recommendations section
        rec_row = 11 + len(negative_water_wells) + 2
        ws.cell(row=rec_row, column=1, value="NEXT STEPS:")
        ws.cell(row=rec_row, column=1).font = Font(bold=True, size=11)

        ws.cell(row=rec_row+1, column=1, value="1. Return to Design tab")
        ws.cell(row=rec_row+2, column=1, value="2. Increase stock concentrations for high-volume reagents")
        ws.cell(row=rec_row+3, column=1, value="3. Or increase final volume")
        ws.cell(row=rec_row+4, column=1, value="4. Re-export design")
        ws.cell(row=rec_row+5, column=1, value="5. Verify this Warning sheet is gone")
        ws.cell(row=rec_row+6, column=1, value="6. CSV file will be exported when conflicts are resolved")

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_excel(self):
        """Export design to Excel (called from main window menu)."""
        self._export_both()

    def export_csv(self):
        """Export design to CSV (called from main window menu)."""
        self._export_both()

    def refresh(self):
        """Refresh UI from project data (called when switching tabs)."""
        self._update_display()


# Check for SMT optimization library (used in _export_both for display)
try:
    from smt.sampling_methods import LHS as SMT_LHS
    HAS_SMT = True
except ImportError:
    HAS_SMT = False
