#!/usr/bin/env python3
"""
DoE Data Analysis Tool - v0.4.1
Statistical analysis GUI for Design of Experiments data
Replicates MATLAB fitlm (EDA CAPKIN MATLAB code), main effects, and interaction plots
Includes Bayesian Optimization for intelligent experiment suggestions

v0.4.1 Changes:
- Fixed pH handling in Bayesian Optimization
- pH now treated as ordered categorical parameter (only suggests tested pH values)
- Prevents extrapolation to untested pH values
- Removed pH rounding logic (no longer needed)

Milton F. Villegas
"""

# GUI and system imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Data processing imports
import pandas as pd
import numpy as np

# Statistical analysis imports
import statsmodels.formula.api as smf

# Bayesian Optimization imports
try:
    from ax.service.ax_client import AxClient, ObjectiveProperties
    from ax.plot.contour import plot_contour
    from ax.plot.feature_importances import plot_feature_importance_by_feature
    AX_AVAILABLE = True
except ImportError:
    AX_AVAILABLE = False
    print("Warning: Ax not available. Install with: pip install ax-platform")

# Plotting imports
import matplotlib
matplotlib.use('TkAgg')  # Backend for embedding plots in tkinter GUI
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from scipy import stats as scipy_stats
import seaborn as sns

# Set plotting style
sns.set_style("whitegrid")
plt.rcParams['font.size'] = 9

# Suppress verbose Ax logging
import logging
logging.getLogger('ax').setLevel(logging.WARNING)


# Data Handler

class DataHandler:
    """Data loading and preprocessing"""
    
    # Metadata columns to exclude from analysis
    METADATA_COLUMNS = ['ID', 'Plate_96', 'Well_96', 'Well_384', 'Source', 'Batch']
    
    def __init__(self):
        self.data = None
        self.clean_data = None
        self.factor_columns = []
        self.categorical_factors = []
        self.numeric_factors = []
        self.response_column = None
        self.stock_concentrations = {}  # Store stock concentrations from metadata
        
    def load_excel(self, filepath: str):
        """Load data from Excel file"""
        self.data = pd.read_excel(filepath)
        
        # Try to load stock concentrations from metadata sheet
        self._load_stock_concentrations(filepath)
    
    def _load_stock_concentrations(self, filepath: str):
        """Load stock concentrations from Stock_Concentrations sheet if it exists"""
        try:
            # Try to read the Stock_Concentrations sheet
            stock_df = pd.read_excel(filepath, sheet_name="Stock_Concentrations")
            
            # Parse the stock concentrations with intelligent matching
            self.stock_concentrations = {}
            for _, row in stock_df.iterrows():
                factor_name = str(row['Factor Name']).strip()
                stock_value_raw = row['Stock Value']

                # Skip rows with None or NaN values
                if pd.isna(stock_value_raw):
                    continue

                stock_value = float(stock_value_raw)

                # Smart matching algorithm - convert display name to internal key
                internal_name = self._smart_factor_match(factor_name)

                if internal_name:
                    self.stock_concentrations[internal_name] = stock_value
            
            print(f"✓ Loaded stock concentrations from metadata: {self.stock_concentrations}")
            
        except Exception as e:
            # Sheet doesn't exist or error reading - that's okay, will use dialog
            print(f"Note: Stock concentrations sheet not found or error reading ({e})")
            self.stock_concentrations = {}
    
    def _smart_factor_match(self, display_name: str) -> str:
        """
        Intelligently match display name to internal factor name.
        Handles any factor format by stripping units and normalizing.
        
        Examples:
        "NaCl (mM)" → "nacl"
        "Detergent (%)" → "detergent_concentration"
        "Buffer Conc (mM)" → "buffer_concentration"
        "Reducing Agent (mM)" → "reducing_agent_concentration"
        """
        # Remove leading/trailing whitespace
        name = display_name.strip()
        
        # Special case: Buffer pH (no units, but needs to stay as is)
        if "Buffer pH" in name or "buffer pH" in name:
            return "buffer pH"
        
        # Extract base name by removing units in parentheses
        # "NaCl (mM)" → "NaCl"
        if '(' in name:
            base_name = name.split('(')[0].strip()
        else:
            base_name = name
        
        # Normalize: lowercase, replace spaces with underscores
        normalized = base_name.lower().replace(' ', '_').replace('-', '_')
        
        # Handle special cases for concentration factors
        # These need "_concentration" suffix based on the original name
        if "buffer conc" in base_name.lower() or "buffer_conc" in normalized:
            return "buffer_concentration"
        elif "detergent" in base_name.lower() and ("%" in name or "conc" in base_name.lower()):
            return "detergent_concentration"
        elif "reducing agent" in base_name.lower() or "reducing_agent" in normalized:
            return "reducing_agent_concentration"
        
        # For everything else, return the normalized name
        # Examples: "NaCl" → "nacl", "Zinc" → "zinc", "Glycerol" → "glycerol"
        return normalized
    
    def get_stock_concentrations(self) -> dict:
        """Get stock concentrations (either from metadata or empty dict)"""
        return self.stock_concentrations.copy()
    
    def detect_columns(self, response_column: str):
        """Detect factor types, exclude metadata"""
        if self.data is None:
            raise ValueError("No data loaded")
        
        self.response_column = response_column
        
        # All columns except response and metadata are factors
        self.factor_columns = [
            col for col in self.data.columns 
            if col != response_column and col not in self.METADATA_COLUMNS
        ]
        
        # Detect categorical vs numeric factors
        self.categorical_factors = []
        self.numeric_factors = []
        
        for col in self.factor_columns:
            if self.data[col].dtype == 'object' or pd.api.types.is_categorical_dtype(self.data[col]):
                self.categorical_factors.append(col)
            else:
                self.numeric_factors.append(col)
    
    def preprocess_data(self):
        """Clean and prepare data for analysis"""
        if self.data is None:
            raise ValueError("No data loaded")
        
        # Create copy
        self.clean_data = self.data.copy()
        
        # Drop metadata columns
        columns_to_drop = [col for col in self.METADATA_COLUMNS if col in self.clean_data.columns]
        if columns_to_drop:
            self.clean_data = self.clean_data.drop(columns=columns_to_drop)
        
        # Drop rows with missing response
        self.clean_data = self.clean_data.dropna(subset=[self.response_column])
        
        # Handle categorical variables - fill NaN with 'None'
        for col in self.categorical_factors:
            self.clean_data[col] = self.clean_data[col].fillna('None')
            self.clean_data[col] = self.clean_data[col].astype(str)
        
        # Handle numeric variables - drop rows with NaN in factors
        for col in self.numeric_factors:
            self.clean_data = self.clean_data.dropna(subset=[col])
        
        return self.clean_data


# DoE Analyzer

class DoEAnalyzer:
    """Statistical analysis via regression"""
    
    MODEL_TYPES = {
        'linear': 'Linear (main effects only)',
        'interactions': 'Linear with 2-way interactions',
        'quadratic': 'Quadratic (interactions + squared terms)',
        'purequadratic': 'Pure quadratic (squared terms only)'
    }
    
    def __init__(self):
        self.data = None
        self.model = None
        self.model_type = 'linear'
        self.factor_columns = []
        self.categorical_factors = []
        self.numeric_factors = []
        self.response_column = None
        self.results = None
        
    def set_data(self, data, factor_columns, categorical_factors, numeric_factors, response_column):
        """Set data and factor information"""
        self.data = data.copy()
        self.factor_columns = factor_columns
        self.categorical_factors = categorical_factors
        self.numeric_factors = numeric_factors
        self.response_column = response_column
    
    def _build_interaction_terms(self, factor_terms):
        """Build interaction terms for all factor combinations"""
        interactions = []
        for i in range(len(factor_terms)):
            for j in range(i+1, len(factor_terms)):
                interactions.append(f"{factor_terms[i]}:{factor_terms[j]}")
        return interactions
        
    def build_formula(self, model_type='linear'):
        """Build regression formula based on model type"""
        self.model_type = model_type
        
        # Prepare factor terms
        factor_terms = []
        for factor in self.factor_columns:
            if factor in self.categorical_factors:
                # C() treats as categorical, Q() quotes column names with spaces
                factor_terms.append(f"C(Q('{factor}'))")
            else:
                factor_terms.append(f"Q('{factor}')")
        
        # Build formula
        if model_type == 'linear':
            formula = f"Q('{self.response_column}') ~ " + " + ".join(factor_terms)
            
        elif model_type == 'interactions':
            main_effects = " + ".join(factor_terms)
            interactions = self._build_interaction_terms(factor_terms)
            formula = f"Q('{self.response_column}') ~ {main_effects}"
            if interactions:
                formula += " + " + " + ".join(interactions)
                
        elif model_type == 'quadratic':
            main_effects = " + ".join(factor_terms)
            interactions = self._build_interaction_terms(factor_terms)
            squared_terms = []
            for factor in self.numeric_factors:
                # I() treats expression as-is (prevents ** being interpreted as interaction)
                squared_terms.append(f"I(Q('{factor}')**2)")
            formula = f"Q('{self.response_column}') ~ {main_effects}"
            if interactions:
                formula += " + " + " + ".join(interactions)
            if squared_terms:
                formula += " + " + " + ".join(squared_terms)
                
        elif model_type == 'purequadratic':
            main_effects = " + ".join(factor_terms)
            squared_terms = []
            for factor in self.numeric_factors:
                # I() treats expression as-is
                squared_terms.append(f"I(Q('{factor}')**2)")
            formula = f"Q('{self.response_column}') ~ {main_effects}"
            if squared_terms:
                formula += " + " + " + ".join(squared_terms)
        else:
            raise ValueError(f"Unknown model type: {model_type}")
        
        return formula
    
    def fit_model(self, model_type='linear'):
        """Fit regression model"""
        if self.data is None:
            raise ValueError("No data set")
        
        formula = self.build_formula(model_type)
        self.model = smf.ols(formula=formula, data=self.data).fit()
        self.results = self._extract_results()
        return self.results
    
    def _extract_results(self):
        """Extract and organize model results"""
        summary_df = pd.DataFrame({
            'Coefficient': self.model.params,
            'Std Error': self.model.bse,
            't-statistic': self.model.tvalues,
            'p-value': self.model.pvalues,
            'Significant': self.model.pvalues < 0.05
        })
        
        model_stats = {
            'R-squared': self.model.rsquared,
            'Adjusted R-squared': self.model.rsquared_adj,
            'RMSE': np.sqrt(self.model.mse_resid),
            'F-statistic': self.model.fvalue,
            'F p-value': self.model.f_pvalue,
            'AIC': self.model.aic,
            'BIC': self.model.bic,
            'Observations': int(self.model.nobs),
            'DF Residuals': int(self.model.df_resid),
            'DF Model': int(self.model.df_model)
        }
        
        return {
            'coefficients': summary_df,
            'model_stats': model_stats,
            'model_type': self.model_type,
            'formula': self.model.model.formula,
            'predictions': self.model.fittedvalues,
            'residuals': self.model.resid
        }
    
    def get_significant_factors(self, alpha=0.05):
        """Get list of significant factors"""
        if self.results is None:
            raise ValueError("No results available")
        
        coef_df = self.results['coefficients']
        significant = coef_df[coef_df['p-value'] < alpha]
        return [idx for idx in significant.index if idx != 'Intercept']
    
    def calculate_main_effects(self):
        """Calculate main effects for each factor"""
        if self.data is None:
            raise ValueError("No data available")
        
        main_effects = {}
        for factor in self.factor_columns:
            # Group by factor levels and calculate mean/std/count for each
            effects = self.data.groupby(factor)[self.response_column].agg(['mean', 'std', 'count'])
            effects.columns = ['Mean Response', 'Std Dev', 'Count']
            main_effects[factor] = effects
        
        return main_effects


# Plotter class for visualizations

class DoEPlotter:
    """Plotting functions for DoE"""

    # Professional colorblind-safe palette (Okabe-Ito)
    COLORS = {
        'primary': '#0173B2',      # Blue
        'secondary': '#DE8F05',    # Orange
        'accent': '#CC78BC',       # Reddish Purple
        'warning': '#D55E00',      # Vermillion
        'success': '#029E73',      # Bluish Green
        'palette': ['#0173B2', '#DE8F05', '#029E73', '#D55E00',
                   '#56B4E9', '#CC78BC', '#ECE133', '#000000']
    }

    def __init__(self):
        self.data = None
        self.factor_columns = []
        self.response_column = None
        
    def set_data(self, data, factor_columns, response_column):
        """Set data for plotting"""
        self.data = data
        self.factor_columns = factor_columns
        self.response_column = response_column
    
    def plot_main_effects(self, save_path=None):
        """Create main effects plot"""
        num_factors = len(self.factor_columns)
        ncols = min(3, num_factors)
        nrows = (num_factors + ncols - 1) // ncols
        
        fig, axes = plt.subplots(nrows, ncols, figsize=(3*ncols, 2*nrows))
        if num_factors == 1:
            axes = [axes]
        else:
            axes = axes.flatten()
        
        for idx, factor in enumerate(self.factor_columns):
            ax = axes[idx]
            grouped = self.data.groupby(factor)[self.response_column].agg(['mean', 'std'])
            levels = sorted(self.data[factor].unique())
            means = [grouped.loc[level, 'mean'] for level in levels]
            stds = [grouped.loc[level, 'std'] for level in levels]
            
            ax.plot(range(len(levels)), means, 'o-', linewidth=2, markersize=8, color=self.COLORS['primary'])
            # Shaded region shows ± 1 std dev
            ax.fill_between(range(len(levels)),
                           [m - s for m, s in zip(means, stds)],
                           [m + s for m, s in zip(means, stds)],
                           alpha=0.2, color=self.COLORS['primary'])
            
            ax.set_xlabel(factor, fontsize=11, fontweight='bold')
            ax.set_ylabel('Mean Response', fontsize=11, fontweight='bold')
            ax.set_title(f'Main Effect: {factor}', fontsize=12, fontweight='bold')
            ax.set_xticks(range(len(levels)))
            ax.set_xticklabels(levels, rotation=45, ha='right')
            ax.grid(True, alpha=0.3)
        
        for idx in range(num_factors, len(axes)):
            fig.delaxes(axes[idx])
        
        plt.tight_layout(pad=0.5)
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
        
        return fig
    
    def plot_interaction_effects(self, save_path=None, max_factors=6):
        """Create interaction plot matrix"""
        factors_to_plot = self.factor_columns[:max_factors]
        num_factors = len(factors_to_plot)
        
        if num_factors < 2:
            return None
        
        fig, axes = plt.subplots(num_factors, num_factors, 
                                figsize=(1.8*num_factors, 1.8*num_factors))
        
        for i, factor1 in enumerate(factors_to_plot):
            for j, factor2 in enumerate(factors_to_plot):
                ax = axes[i, j]
                
                # Diagonal: main effects
                if i == j:
                    grouped = self.data.groupby(factor1)[self.response_column].mean()
                    levels = sorted(self.data[factor1].unique())
                    means = [grouped.loc[level] for level in levels]

                    ax.plot(range(len(levels)), means, 'o-', linewidth=2, color=self.COLORS['primary'])
                    ax.set_xticks(range(len(levels)))
                    ax.set_xticklabels(levels, rotation=45, ha='right', fontsize=8)
                    
                    if j == 0:
                        ax.set_ylabel('Mean\nResponse', fontsize=9)
                    
                # Lower triangle: interaction plots
                elif i > j:
                    levels1 = sorted(self.data[factor1].unique())
                    levels2 = sorted(self.data[factor2].unique())

                    for idx, level2 in enumerate(levels2):
                        subset = self.data[self.data[factor2] == level2]
                        grouped = subset.groupby(factor1)[self.response_column].mean()
                        means = [grouped.loc[level1] if level1 in grouped.index else np.nan
                                for level1 in levels1]
                        color = self.COLORS['palette'][idx % len(self.COLORS['palette'])]
                        ax.plot(range(len(levels1)), means, 'o-', linewidth=1.5,
                               label=f'{factor2}={level2}', alpha=0.85, color=color)
                    
                    ax.set_xticks(range(len(levels1)))
                    ax.set_xticklabels(levels1, rotation=45, ha='right', fontsize=8)
                    
                    if j == 0:
                        ax.set_ylabel('Mean\nResponse', fontsize=9)
                    
                    if i == 1 and j == 0 and len(levels2) <= 5:
                        ax.legend(fontsize=7, loc='best')
                    
                # Upper triangle: just label the comparison
                else:
                    ax.text(0.5, 0.5, f'{factor1}\nvs\n{factor2}', 
                           ha='center', va='center', fontsize=9,
                           transform=ax.transAxes, style='italic')
                    ax.set_xticks([])
                    ax.set_yticks([])
                
                ax.grid(True, alpha=0.3)
                
                if i == 0:
                    ax.set_title(factor2, fontsize=10, fontweight='bold')
                
                if j == 0 and i > 0:
                    ax.set_ylabel(factor1, fontsize=10, fontweight='bold', rotation=0, 
                                 ha='right', va='center', labelpad=20)
        
        plt.tight_layout(pad=0.5)
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
        
        return fig
    
    def plot_residuals(self, predictions, residuals, save_path=None):
        """Create residual diagnostic plots"""
        fig, axes = plt.subplots(2, 2, figsize=(9, 7))
        
        # Residuals vs Fitted
        axes[0, 0].scatter(predictions, residuals, alpha=0.6, color=self.COLORS['primary'], edgecolors='white', linewidth=0.5)
        axes[0, 0].axhline(y=0, color=self.COLORS['warning'], linestyle='--', linewidth=2)
        axes[0, 0].set_xlabel('Fitted Values', fontsize=11)
        axes[0, 0].set_ylabel('Residuals', fontsize=11)
        axes[0, 0].set_title('Residuals vs Fitted', fontsize=12, fontweight='bold')
        axes[0, 0].grid(True, alpha=0.3)
        
        # Q-Q plot (This was just made to have inverted axes to match PRISM style)
        qq_data = scipy_stats.probplot(residuals, dist="norm")
        theoretical_quantiles = qq_data[0][0]  # Y-axis
        ordered_residuals = qq_data[0][1]      # X-axis
        slope = qq_data[1][0]
        intercept = qq_data[1][1]
        
        # Plot data points with swapped axes
        axes[0, 1].plot(ordered_residuals, theoretical_quantiles, 'o', markersize=6,
                       color=self.COLORS['primary'], alpha=0.7, markeredgecolor='white', markeredgewidth=0.5)

        # Reference line: Since normal relationship is y = slope*x + intercept
        # When we swap axes, the line becomes: theoretical = (actual - intercept) / slope
        # Or simplified: theoretical = (1/slope) * actual - (intercept/slope)
        x_line = np.array([ordered_residuals.min(), ordered_residuals.max()])
        y_line = (x_line - intercept) / slope
        axes[0, 1].plot(x_line, y_line, '-', linewidth=2, color=self.COLORS['warning'])
        
        axes[0, 1].set_xlabel('Actual residual', fontsize=11)
        axes[0, 1].set_ylabel('Predicted residual', fontsize=11)
        axes[0, 1].set_title('Normal Q-Q Plot', fontsize=12, fontweight='bold')
        axes[0, 1].grid(True, alpha=0.3)
        
        # Scale-Location
        standardized_resid = residuals / residuals.std()
        axes[1, 0].scatter(predictions, np.sqrt(np.abs(standardized_resid)), alpha=0.6,
                          color=self.COLORS['primary'], edgecolors='white', linewidth=0.5)
        axes[1, 0].set_xlabel('Fitted Values', fontsize=11)
        axes[1, 0].set_ylabel('√|Standardized Residuals|', fontsize=11)
        axes[1, 0].set_title('Scale-Location', fontsize=12, fontweight='bold')
        axes[1, 0].grid(True, alpha=0.3)
        
        # Histogram
        axes[1, 1].hist(residuals, bins=30, color=self.COLORS['primary'],
                       edgecolor='white', alpha=0.8, linewidth=1.2)
        axes[1, 1].set_xlabel('Residuals', fontsize=11)
        axes[1, 1].set_ylabel('Frequency', fontsize=11)
        axes[1, 1].set_title('Histogram of Residuals', fontsize=12, fontweight='bold')
        axes[1, 1].grid(True, alpha=0.3)
        
        plt.tight_layout(pad=0.5)
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
        
        return fig


# Results Exporter

class ResultsExporter:
    """Exports results to various formats"""
    
    def __init__(self):
        self.results = None
        self.main_effects = None
        
    def set_results(self, results, main_effects):
        """Set results to export"""
        self.results = results
        self.main_effects = main_effects
    
    def export_statistics_excel(self, filepath):
        """Export statistics to Excel"""
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Model Statistics
            model_stats_df = pd.DataFrame([self.results['model_stats']]).T
            model_stats_df.columns = ['Value']
            model_stats_df.index.name = 'Statistic'
            model_stats_df.to_excel(writer, sheet_name='Model Statistics')
            
            # Coefficients
            coef_df = self.results['coefficients'].copy()
            coef_df['p-value'] = coef_df['p-value'].apply(lambda x: f"{x:.6e}")
            coef_df.to_excel(writer, sheet_name='Coefficients')
            
            # Main Effects
            main_effects_combined = []
            for factor, effects_df in self.main_effects.items():
                effects_df_copy = effects_df.copy()
                effects_df_copy.insert(0, 'Factor', factor)
                effects_df_copy.reset_index(inplace=True)
                effects_df_copy.rename(columns={effects_df_copy.columns[1]: 'Level'}, inplace=True)
                main_effects_combined.append(effects_df_copy)
            
            combined_df = pd.concat(main_effects_combined, ignore_index=True)
            combined_df.to_excel(writer, sheet_name='Main Effects', index=False)
            
            # Significant Factors
            sig_factors = self.results['coefficients'][
                self.results['coefficients']['p-value'].astype(float) < 0.05
            ].copy()
            sig_factors = sig_factors[sig_factors.index != 'Intercept']
            sig_factors['p-value'] = sig_factors['p-value'].apply(lambda x: f"{x:.6e}")
            sig_factors.to_excel(writer, sheet_name='Significant Factors')


# Bayesian Optimizer

class BayesianOptimizer:
    """Bayesian Optimization for intelligent experiment suggestions"""

    # Professional colorblind-safe palette (Okabe-Ito)
    COLORS = {
        'primary': '#0173B2',      # Blue
        'accent': '#CC78BC',       # Reddish Purple
        'warning': '#D55E00',      # Vermillion
    }

    def __init__(self):
        self.ax_client = None
        self.data = None
        self.factor_columns = []
        self.numeric_factors = []
        self.categorical_factors = []
        self.response_column = None
        self.factor_bounds = {}
        self.is_initialized = False
        self.name_mapping = {}  # Maps sanitized names back to original names
        self.reverse_mapping = {}  # Maps original names to sanitized names
    
    def _sanitize_name(self, name):
        """Replace spaces and special characters with underscores for Ax compatibility"""
        sanitized = name.replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '')
        return sanitized
    
    def set_data(self, data, factor_columns, categorical_factors, numeric_factors, response_column):
        """Set data and factor information"""
        self.data = data.copy()
        self.factor_columns = factor_columns
        self.categorical_factors = categorical_factors
        self.numeric_factors = numeric_factors
        self.response_column = response_column
        
        # Create name mappings
        self.reverse_mapping = {name: self._sanitize_name(name) for name in factor_columns}
        self.name_mapping = {self._sanitize_name(name): name for name in factor_columns}
        
        self._calculate_bounds()
    
    def _calculate_bounds(self):
        """Calculate bounds for each factor from the data"""
        self.factor_bounds = {}
        
        for factor in self.numeric_factors:
            min_val = float(self.data[factor].min())
            max_val = float(self.data[factor].max())
            self.factor_bounds[factor] = (min_val, max_val)
        
        for factor in self.categorical_factors:
            unique_vals = self.data[factor].unique().tolist()
            self.factor_bounds[factor] = unique_vals
    
    def initialize_optimizer(self, minimize=False):
        """Initialize Ax client with parameters"""
        if not AX_AVAILABLE:
            raise ImportError("Ax platform not available. Install with: pip install ax-platform")
        
        # Build parameters list with sanitized names
        parameters = []
        for factor in self.factor_columns:
            sanitized_name = self.reverse_mapping[factor]
            
            if factor in self.numeric_factors:
                # Special handling for pH - treat as ordered categorical
                if 'ph' in factor.lower() and 'buffer' in factor.lower():
                    # Get unique pH values that were ACTUALLY TESTED
                    tested_ph_values = sorted(self.data[factor].unique().tolist())
                    
                    print(f"ℹ️  Treating '{factor}' as ordered categorical parameter")
                    print(f"   Tested pH values: {tested_ph_values}")
                    print(f"   BO will only suggest from these tested values (no extrapolation)")
                    
                    parameters.append({
                        "name": sanitized_name,
                        "type": "choice",
                        "values": tested_ph_values,
                        "is_ordered": True,  # Tells Ax that pH values have natural ordering
                        "value_type": "float"
                    })
                else:
                    # Normal continuous numeric factors
                    min_val, max_val = self.factor_bounds[factor]
                    parameters.append({
                        "name": sanitized_name,
                        "type": "range",
                        "bounds": [min_val, max_val],
                        "value_type": "float"
                    })
            elif factor in self.categorical_factors:
                parameters.append({
                    "name": sanitized_name,
                    "type": "choice",
                    "values": self.factor_bounds[factor],
                    "value_type": "str"
                })
        
        # Create Ax client
        self.ax_client = AxClient()
        self.ax_client.create_experiment(
            name="doe_optimization",
            parameters=parameters,
            objectives={self.response_column: ObjectiveProperties(minimize=minimize)},
            choose_generation_strategy_kwargs={"num_initialization_trials": 0}  # Skip Sobol, go straight to BO
        )
        
        # Add existing data as completed trials using sanitized names
        for idx, row in self.data.iterrows():
            params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                val = row[factor]
                # Ensure proper types
                if factor in self.numeric_factors:
                    # pH is now categorical, but still pass as float
                    if 'ph' in factor.lower() and 'buffer' in factor.lower():
                        params[sanitized_name] = float(val)
                    else:
                        params[sanitized_name] = float(val)
                else:
                    params[sanitized_name] = str(val)
            
            # Add trial
            _, trial_index = self.ax_client.attach_trial(parameters=params)
            self.ax_client.complete_trial(
                trial_index=trial_index,
                raw_data=float(row[self.response_column])
            )
        
        self.is_initialized = True
    
    def get_next_suggestions(self, n=5):
        """Get next experiment suggestions with original factor names and proper rounding"""
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized. Call initialize_optimizer() first.")
        
        suggestions = []
        for _ in range(n):
            params, trial_index = self.ax_client.get_next_trial()
            
            # Convert sanitized names back to original names and apply rounding
            original_params = {}
            for sanitized_name, value in params.items():
                original_name = self.name_mapping[sanitized_name]
                
                # Apply rounding based on factor type
                if isinstance(value, (int, float)):
                    # Check if this is a pH factor (now categorical, no rounding needed)
                    if 'ph' in original_name.lower() and 'buffer' in original_name.lower():
                        # pH is categorical - value comes directly from tested values
                        original_params[original_name] = float(value)
                    else:
                        # Round to 2 decimals for other numeric factors
                        rounded_value = round(value, 2)
                        original_params[original_name] = rounded_value
                else:
                    # Categorical factors - keep as is
                    original_params[original_name] = value
            
            suggestions.append(original_params)
            # Abandon trial so we can generate more suggestions
            self.ax_client.abandon_trial(trial_index)
        
        return suggestions
    
    def _create_suggestion_heatmap(self, factor_x_orig, factor_y_orig,
                                   factor_x_san, factor_y_san, X, Y):
        """Create a heatmap showing where BO suggests exploring - ENHANCED VERSION"""
        try:
            # Larger for export quality
            fig, ax = plt.subplots(1, 1, figsize=(9, 7))

            # Generate many suggestions and extract their x,y positions
            n_suggestions = 50
            suggestions_x = []
            suggestions_y = []

            for _ in range(n_suggestions):
                try:
                    params, trial_idx = self.ax_client.get_next_trial()
                    suggestions_x.append(params[factor_x_san])
                    suggestions_y.append(params[factor_y_san])
                    self.ax_client.abandon_trial(trial_idx)
                except:
                    break

            if len(suggestions_x) < 5:
                print("Could not generate enough suggestions for heatmap")
                return None

            # Create 2D histogram / density plot
            from scipy.stats import gaussian_kde
            xy = np.vstack([suggestions_x, suggestions_y])
            z = gaussian_kde(xy)(xy)

            # Plot scatter with density colors - modern colormap
            scatter = ax.scatter(suggestions_x, suggestions_y, c=z, s=150,
                               cmap='plasma', alpha=0.7,
                               edgecolors='white', linewidth=2)

            # Enhanced colorbar
            cbar = plt.colorbar(scatter, ax=ax, label='Suggestion Density')
            cbar.ax.tick_params(labelsize=11)
            cbar.set_label('Suggestion Density\n(Higher = More Recommended)',
                          fontsize=12, fontweight='bold')

            # Plot existing experiments with modern styling
            existing_x = self.data[factor_x_orig].values
            existing_y = self.data[factor_y_orig].values
            ax.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                      edgecolors='white', linewidth=3,
                      label='Existing Experiments', zorder=5, marker='s', alpha=0.9)

            # Enhanced labels and title
            ax.set_xlabel(factor_x_orig, fontsize=13, fontweight='bold', color='#333333')
            ax.set_ylabel(factor_y_orig, fontsize=13, fontweight='bold', color='#333333')
            ax.set_title('Bayesian Optimization: Recommended Exploration Regions',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')

            # Modern legend
            ax.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')

            # Modern grid
            ax.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax.set_axisbelow(True)

            # Set background
            ax.set_facecolor('#FAFAFA')

            # Borders
            for spine in ax.spines.values():
                spine.set_edgecolor('#CCCCCC')
                spine.set_linewidth(1.5)

            plt.tight_layout()
            return fig

        except Exception as e:
            print(f"Error creating suggestion heatmap: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _select_most_important_factors(self):
        """Select the 2 most important numeric factors using feature importances"""
        if len(self.numeric_factors) <= 2:
            return self.numeric_factors[:2] if len(self.numeric_factors) == 2 else None

        try:
            # Use BoTorch's feature_importances method
            adapter = self.ax_client.generation_strategy.adapter

            if hasattr(adapter, 'generator') and hasattr(adapter.generator, 'feature_importances'):
                # Get feature importances from BoTorchGenerator
                importances = adapter.generator.feature_importances()

                # importances is a numpy array - flatten it
                import numpy as np
                importance_values = np.array(importances).flatten()

                # Get parameter names in the order they appear in the model
                param_names = list(adapter.parameters)

                # Create dict mapping original factor names to importance values
                factor_importances = {}
                for idx, param_name in enumerate(param_names):
                    if idx < len(importance_values):
                        # Find the original factor name from the sanitized parameter name
                        for orig_factor in self.numeric_factors:
                            sanitized = self.reverse_mapping.get(orig_factor, orig_factor)
                            if sanitized == param_name:
                                factor_importances[orig_factor] = float(importance_values[idx])
                                break

                if len(factor_importances) >= 2:
                    print(f"✓ Using feature importances for factor selection")
                    print(f"  Feature importances: {factor_importances}")

                    # Sort by importance and take top 2
                    sorted_factors = sorted(factor_importances.items(), key=lambda x: x[1], reverse=True)
                    selected = [f[0] for f in sorted_factors[:2]]

                    print(f"  Selected most important factors: {selected}")
                    return selected

            # If feature importances not available, fall back
            print("⚠ Could not extract feature importances, falling back to range-based selection")

        except NotImplementedError:
            # Some models (e.g., mixed continuous/categorical) don't support feature importances
            print("⚠ Feature importances not supported for this model type")
            print("  Falling back to range-based selection")
        except Exception as e:
            print(f"⚠ Error extracting feature importances: {e}")
            print("  Falling back to range-based selection")

        # Fallback: Sobol sensitivity analysis - works for all model types
        # Uses Ax's built-in global sensitivity analysis with Sobol indices
        try:
            print("  Using Sobol sensitivity analysis for factor selection...")

            # Try to use Ax's built-in Sobol sensitivity analysis
            from ax.analysis.plotly.sensitivity import SensitivityAnalysisPlot

            analysis = SensitivityAnalysisPlot()
            card = analysis.compute(
                experiment=self.ax_client._experiment,
                generation_strategy=self.ax_client._generation_strategy
            )

            # Extract Sobol indices from the analysis card
            # The blob is a JSON string containing the plotly figure
            if card is not None:
                import json
                import plotly.graph_objects as go

                # Parse the JSON blob to get the plotly figure data
                blob_json = json.loads(card.blob)
                fig = go.Figure(blob_json)

                # Parse Sobol indices from the plotly figure data
                # The data structure contains parameter names and their total Sobol indices
                sensitivities = {}

                if hasattr(fig, 'data') and len(fig.data) > 0:
                    # Extract from plotly bar chart data
                    # Note: In plotly's encoding, parameter names are in y and values are in x (as binary data)
                    for idx, trace in enumerate(fig.data):
                        if hasattr(trace, 'x') and hasattr(trace, 'y'):
                            # Extract parameter names (in trace.y) and Sobol values (in trace.x as binary)
                            param_names = trace.y  # Parameter names are in y
                            sobol_data = trace.x   # Sobol values are in x (binary encoded)

                            # Decode binary-encoded Sobol values
                            if isinstance(param_names, (list, tuple)) and isinstance(sobol_data, dict):
                                import base64
                                import numpy as np

                                # Decode base64 binary data
                                binary_data = base64.b64decode(sobol_data['bdata'])
                                dtype = sobol_data.get('dtype', 'f8')
                                sobol_values = np.frombuffer(binary_data, dtype=dtype)

                                print(f"  Decoded {len(sobol_values)} Sobol indices from analysis")

                                # Map parameter names to Sobol values
                                for param_name, sobol_val in zip(param_names, sobol_values):
                                    # Map sanitized parameter names back to original factor names
                                    for orig_factor in self.numeric_factors:
                                        sanitized = self.reverse_mapping.get(orig_factor, orig_factor)
                                        if sanitized == param_name:
                                            sensitivities[orig_factor] = float(sobol_val)
                                            break
                else:
                    print(f"  No Sobol data found in analysis")

                if len(sensitivities) >= 2:
                    print(f"✓ Using Sobol sensitivity indices for factor selection")
                    print(f"  Sobol indices: {sensitivities}")

                    sorted_factors = sorted(sensitivities.items(), key=lambda x: x[1], reverse=True)
                    selected = [f[0] for f in sorted_factors[:2]]

                    print(f"  Selected most important factors: {selected}")
                    return selected
                else:
                    print(f"  Could not extract enough Sobol indices (got {len(sensitivities)})")

        except ImportError:
            print(f"  Sobol analysis not available (Ax version may not support it)")
        except Exception as e:
            print(f"  Sobol analysis failed: {e}")
            import traceback
            traceback.print_exc()

        # Ultimate fallback: largest parameter ranges
        factor_ranges = {}
        for factor in self.numeric_factors:
            min_val, max_val = self.factor_bounds[factor]
            data_range = max_val - min_val
            if data_range > 0:
                factor_ranges[factor] = data_range

        sorted_factors = sorted(factor_ranges.items(), key=lambda x: x[1], reverse=True)
        selected = [f[0] for f in sorted_factors[:2]]

        print(f"  Selected factors based on range: {selected}")
        return selected
    
    def get_acquisition_plot(self):
        """Generate comprehensive BO visualization with 4 panels (GUI preview)"""
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized")

        # Need at least 2 numeric factors
        if len(self.numeric_factors) < 2:
            return None

        try:
            # Select the 2 most important numeric factors
            selected_factors = self._select_most_important_factors()
            if selected_factors is None or len(selected_factors) < 2:
                return None

            factor_x_original = selected_factors[0]
            factor_y_original = selected_factors[1]

            factor_x_sanitized = self.reverse_mapping[factor_x_original]
            factor_y_sanitized = self.reverse_mapping[factor_y_original]

            print(f"Creating comprehensive BO plots for: {factor_x_original} vs {factor_y_original}")

            # Create grid
            x_min, x_max = self.factor_bounds[factor_x_original]
            y_min, y_max = self.factor_bounds[factor_y_original]

            # Check if bounds are valid (not constant)
            if x_min == x_max or y_min == y_max:
                print(f"Factor has no variation: {factor_x_original}={x_min} or {factor_y_original}={y_min}")
                return None

            # Add small padding to bounds
            x_range = x_max - x_min
            y_range = y_max - y_min
            x_min -= 0.05 * x_range
            x_max += 0.05 * x_range
            y_min -= 0.05 * y_range
            y_max += 0.05 * y_range

            # Use coarser grid for faster GUI preview
            x = np.linspace(x_min, x_max, 15)
            y = np.linspace(y_min, y_max, 15)
            X, Y = np.meshgrid(x, y)

            # Get a template with all other factors at their median values
            template_params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                if factor == factor_x_original or factor == factor_y_original:
                    continue
                if factor in self.numeric_factors:
                    template_params[sanitized_name] = float(self.data[factor].median())
                else:
                    template_params[sanitized_name] = str(self.data[factor].mode()[0])

            # Build list of parameterizations for batch prediction
            parameterizations = []
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    params = template_params.copy()
                    params[factor_x_sanitized] = float(X[i, j])
                    params[factor_y_sanitized] = float(Y[i, j])
                    parameterizations.append(params)

            print(f"Predicting {len(parameterizations)} points using BO model...")

            # Get predictions with uncertainty
            try:
                predictions_list = self.ax_client.get_model_predictions_for_parameterizations(
                    parameterizations=parameterizations,
                    metric_names=[self.response_column]
                )

                # Extract predictions and uncertainty into arrays
                Z_mean = np.zeros_like(X)
                Z_sem = np.zeros_like(X)
                idx = 0
                for i in range(X.shape[0]):
                    for j in range(X.shape[1]):
                        pred_mean, pred_sem = predictions_list[idx][self.response_column]
                        Z_mean[i, j] = pred_mean
                        Z_sem[i, j] = pred_sem
                        idx += 1

                print(f"Successfully predicted all {len(parameterizations)} points")

            except Exception as e:
                print(f"Batch prediction failed: {e}")
                print("Falling back to suggestion density heatmap...")
                return self._create_suggestion_heatmap(factor_x_original, factor_y_original,
                                                       factor_x_sanitized, factor_y_sanitized,
                                                       X, Y)

            # CREATE 2x2 MULTI-PANEL FIGURE (GUI PREVIEW)
            fig = plt.figure(figsize=(12, 9))
            gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.3)

            # PANEL 1: Response Surface (Top-Left)
            ax1 = fig.add_subplot(gs[0, 0])
            contour1 = ax1.contourf(X, Y, Z_mean, levels=15, cmap='RdYlGn', alpha=0.9)
            cbar1 = plt.colorbar(contour1, ax=ax1)
            cbar1.ax.tick_params(labelsize=8)
            cbar1.set_label(f'{self.response_column}', fontsize=9, fontweight='bold')

            contour_lines1 = ax1.contour(X, Y, Z_mean, levels=8, colors='black',
                                        alpha=0.3, linewidths=1)
            ax1.clabel(contour_lines1, inline=True, fontsize=7, fmt='%.1f')

            existing_x = self.data[factor_x_original].values
            existing_y = self.data[factor_y_original].values
            ax1.scatter(existing_x, existing_y, c='#2E86AB', s=80,
                       edgecolors='white', linewidth=2, label='Tested', zorder=5, alpha=0.9)

            ax1.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax1.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax1.set_title('Response Surface (GP Mean)', fontsize=11, fontweight='bold', pad=10)
            ax1.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax1.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax1.set_facecolor('#FAFAFA')

            # PANEL 2: Acquisition Function (Top-Right)
            ax2 = fig.add_subplot(gs[0, 1])
            # Calculate Expected Improvement (EI) - Proper formula
            current_best = self.data[self.response_column].max()  # Assuming maximize

            # Avoid division by zero
            Z_sem_safe = np.where(Z_sem > 1e-6, Z_sem, 1e-6)

            # Standardized improvement
            Z_score = (Z_mean - current_best) / Z_sem_safe

            # Proper EI formula: EI = (mu - best) * Phi(Z) + sigma * phi(Z)
            # where Phi is CDF and phi is PDF of standard normal
            Z_ei = (Z_mean - current_best) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)
            Z_ei = np.maximum(Z_ei, 0)  # EI is always non-negative

            contour2 = ax2.contourf(X, Y, Z_ei, levels=15, cmap='plasma', alpha=0.9)
            cbar2 = plt.colorbar(contour2, ax=ax2)
            cbar2.ax.tick_params(labelsize=8)
            cbar2.set_label('EI Score', fontsize=9, fontweight='bold')

            ax2.scatter(existing_x, existing_y, c='#2E86AB', s=80,
                       edgecolors='white', linewidth=2, label='Tested', zorder=5, alpha=0.9)

            # Mark best point so far
            best_idx = self.data[self.response_column].idxmax()
            best_x = self.data.loc[best_idx, factor_x_original]
            best_y = self.data.loc[best_idx, factor_y_original]
            ax2.scatter([best_x], [best_y], c='gold', s=200, marker='*',
                       edgecolors='black', linewidth=2, label='Best', zorder=6)

            ax2.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax2.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax2.set_title('Acquisition Function (Where to Sample)', fontsize=11, fontweight='bold', pad=10)
            ax2.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax2.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax2.set_facecolor('#FAFAFA')

            # PANEL 3: GP Uncertainty Map (Bottom-Left)
            ax3 = fig.add_subplot(gs[1, 0])

            # Plot uncertainty (standard error) as contour map
            # Z_sem was computed earlier - it shows where the model is uncertain
            contour3 = ax3.contourf(X, Y, Z_sem, levels=15, cmap='YlOrRd', alpha=0.9)

            # Add contour lines for clarity
            contour_lines3 = ax3.contour(X, Y, Z_sem, levels=8, colors='black',
                                         linewidths=0.5, alpha=0.3)
            ax3.clabel(contour_lines3, inline=True, fontsize=7, fmt='%.2f')

            # Mark observed points
            ax3.scatter(existing_x, existing_y, c='black', s=80,
                       marker='o', edgecolors='white', linewidth=2,
                       label='Observed Points', zorder=5)

            # Mark the best point (already calculated above)
            ax3.scatter([best_x], [best_y], c='lime', s=200, marker='*',
                       edgecolors='black', linewidth=2, label='Best Point', zorder=6)

            # Add colorbar
            cbar3 = plt.colorbar(contour3, ax=ax3, orientation='vertical', pad=0.02, shrink=0.9)
            cbar3.set_label('GP Std. Error', fontsize=9, fontweight='bold')
            cbar3.ax.tick_params(labelsize=8)

            # Calculate uncertainty statistics
            mean_uncertainty = np.mean(Z_sem)
            max_uncertainty = np.max(Z_sem)
            ax3.text(0.05, 0.95, f'Mean: {mean_uncertainty:.3f}\nMax: {max_uncertainty:.3f}',
                    transform=ax3.transAxes, fontsize=9, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.8, edgecolor='#CCCCCC'))

            ax3.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax3.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax3.set_title('Model Uncertainty (GP Std. Error)', fontsize=11, fontweight='bold', pad=10)
            ax3.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC', loc='lower right')
            ax3.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax3.set_facecolor('#FAFAFA')

            # PANEL 4: Optimization Progress (Bottom-Right)
            ax4 = fig.add_subplot(gs[1, 1])

            # Get cumulative best values
            response_values = self.data[self.response_column].values
            cumulative_best = np.maximum.accumulate(response_values)  # Assuming maximize
            iterations = np.arange(1, len(cumulative_best) + 1)

            # Plot progress
            ax4.plot(iterations, cumulative_best, 'o-', color='#029E73', linewidth=2.5,
                    markersize=6, markerfacecolor='#029E73', markeredgecolor='white',
                    markeredgewidth=1.5, label='Best So Far', alpha=0.9)

            # Fill area under curve
            ax4.fill_between(iterations, cumulative_best, alpha=0.2, color='#029E73')

            # Mark current best
            ax4.scatter([len(iterations)], [cumulative_best[-1]], c='gold', s=200,
                       marker='*', edgecolors='black', linewidth=2, zorder=5,
                       label='Current Best')

            # Set Y-axis limits based on data range (with 10% margin)
            y_min = min(response_values.min(), cumulative_best.min())
            y_max = cumulative_best.max()
            y_range = y_max - y_min
            ax4.set_ylim(y_min - 0.1 * y_range, y_max + 0.05 * y_range)

            ax4.set_xlabel('Experiment Number', fontsize=10, fontweight='bold')
            ax4.set_ylabel(f'Best {self.response_column}', fontsize=10, fontweight='bold')
            ax4.set_title('Optimization Progress', fontsize=11, fontweight='bold', pad=10)
            ax4.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax4.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax4.set_facecolor('#FAFAFA')

            # Annotate improvement
            total_improvement = cumulative_best[-1] - cumulative_best[0]
            ax4.text(0.05, 0.05, f'Total Δ: {total_improvement:+.2f}',
                    transform=ax4.transAxes, fontsize=9, fontweight='bold',
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.8,
                    edgecolor='#CCCCCC'))

            plt.tight_layout()
            return fig
            
        except Exception as e:
            print(f"Error generating acquisition plot: {e}")
            import traceback
            traceback.print_exc()
            return None

    def export_bo_plots(self, directory, base_name="Experiment", date_str=None, file_format="png", dpi=300):
        """Export individual high-resolution BO plots for publication

        Args:
            directory: Output directory
            base_name: Base filename
            date_str: Date string (YYYYMMDD)
            file_format: File format (png, tiff, pdf, eps)
            dpi: Resolution for raster formats (PNG, TIFF)
        """
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized")

        if date_str is None:
            date_str = datetime.now().strftime('%Y%m%d')

        exported_files = []

        try:
            # Get the necessary data (same as preview)
            selected_factors = self._select_most_important_factors()
            if selected_factors is None or len(selected_factors) < 2:
                return []

            factor_x_original = selected_factors[0]
            factor_y_original = selected_factors[1]
            factor_x_sanitized = self.reverse_mapping[factor_x_original]
            factor_y_sanitized = self.reverse_mapping[factor_y_original]

            # Create higher resolution grid for export
            x_min, x_max = self.factor_bounds[factor_x_original]
            y_min, y_max = self.factor_bounds[factor_y_original]
            x_range = x_max - x_min
            y_range = y_max - y_min
            x_min -= 0.05 * x_range
            x_max += 0.05 * x_range
            y_min -= 0.05 * y_range
            y_max += 0.05 * y_range

            x = np.linspace(x_min, x_max, 30)  # Higher resolution for export
            y = np.linspace(y_min, y_max, 30)
            X, Y = np.meshgrid(x, y)

            # Template params
            template_params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                if factor == factor_x_original or factor == factor_y_original:
                    continue
                if factor in self.numeric_factors:
                    template_params[sanitized_name] = float(self.data[factor].median())
                else:
                    template_params[sanitized_name] = str(self.data[factor].mode()[0])

            # Build parameterizations
            parameterizations = []
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    params = template_params.copy()
                    params[factor_x_sanitized] = float(X[i, j])
                    params[factor_y_sanitized] = float(Y[i, j])
                    parameterizations.append(params)

            # Get predictions
            predictions_list = self.ax_client.get_model_predictions_for_parameterizations(
                parameterizations=parameterizations,
                metric_names=[self.response_column]
            )

            Z_mean = np.zeros_like(X)
            Z_sem = np.zeros_like(X)
            idx = 0
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    pred_mean, pred_sem = predictions_list[idx][self.response_column]
                    Z_mean[i, j] = pred_mean
                    Z_sem[i, j] = pred_sem
                    idx += 1

            existing_x = self.data[factor_x_original].values
            existing_y = self.data[factor_y_original].values

            # EXPORT 1: Response Surface
            fig1, ax1 = plt.subplots(1, 1, figsize=(9, 7))
            contour1 = ax1.contourf(X, Y, Z_mean, levels=20, cmap='RdYlGn', alpha=0.9)
            cbar1 = plt.colorbar(contour1, ax=ax1)
            cbar1.ax.tick_params(labelsize=11)
            cbar1.set_label(f'Predicted {self.response_column}', fontsize=12, fontweight='bold')
            contour_lines1 = ax1.contour(X, Y, Z_mean, levels=10, colors='black',
                                        alpha=0.4, linewidths=1.5)
            ax1.clabel(contour_lines1, inline=True, fontsize=9, fmt='%.1f')
            ax1.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                       edgecolors='white', linewidth=3, label='Existing Experiments',
                       zorder=5, marker='o', alpha=0.9)
            ax1.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax1.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax1.set_title('Bayesian Optimization: Predicted Response Surface',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax1.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax1.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax1.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath1 = os.path.join(directory, f'{base_name}_BO_ResponseSurface_{date_str}.{file_format}')
            fig1.savefig(filepath1, dpi=dpi, bbox_inches='tight')
            plt.close(fig1)
            exported_files.append(filepath1)

            # EXPORT 2: Acquisition Function
            fig2, ax2 = plt.subplots(1, 1, figsize=(9, 7))
            current_best = self.data[self.response_column].max()

            # Proper EI formula
            Z_sem_safe = np.where(Z_sem > 1e-6, Z_sem, 1e-6)
            Z_score = (Z_mean - current_best) / Z_sem_safe
            Z_ei = (Z_mean - current_best) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)
            Z_ei = np.maximum(Z_ei, 0)
            contour2 = ax2.contourf(X, Y, Z_ei, levels=20, cmap='plasma', alpha=0.9)
            cbar2 = plt.colorbar(contour2, ax=ax2)
            cbar2.ax.tick_params(labelsize=11)
            cbar2.set_label('Expected Improvement Score', fontsize=12, fontweight='bold')
            ax2.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                       edgecolors='white', linewidth=3, label='Tested', zorder=5, alpha=0.9)
            best_idx = self.data[self.response_column].idxmax()
            best_x = self.data.loc[best_idx, factor_x_original]
            best_y = self.data.loc[best_idx, factor_y_original]
            ax2.scatter([best_x], [best_y], c='gold', s=300, marker='*',
                       edgecolors='black', linewidth=3, label='Current Best', zorder=6)
            ax2.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax2.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax2.set_title('Bayesian Optimization: Acquisition Function',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax2.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax2.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax2.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath2 = os.path.join(directory, f'{base_name}_BO_Acquisition_{date_str}.{file_format}')
            fig2.savefig(filepath2, dpi=dpi, bbox_inches='tight')
            plt.close(fig2)
            exported_files.append(filepath2)

            # EXPORT 3: GP Uncertainty Map
            fig3, ax3 = plt.subplots(1, 1, figsize=(9, 7))

            # Plot uncertainty (standard error) as contour map
            contour3 = ax3.contourf(X, Y, Z_sem, levels=15, cmap='YlOrRd', alpha=0.9)

            # Add contour lines for clarity
            contour_lines3 = ax3.contour(X, Y, Z_sem, levels=10, colors='black',
                                         linewidths=0.8, alpha=0.4)
            ax3.clabel(contour_lines3, inline=True, fontsize=9, fmt='%.2f')

            # Mark observed points
            ax3.scatter(existing_x, existing_y, c='black', s=120,
                       marker='o', edgecolors='white', linewidth=2.5,
                       label='Observed Points', zorder=5)

            # Mark the best point (already calculated above)
            ax3.scatter([best_x], [best_y], c='lime', s=350, marker='*',
                       edgecolors='black', linewidth=3, label='Best Point', zorder=6)

            # Add colorbar
            cbar3 = plt.colorbar(contour3, ax=ax3, orientation='vertical', pad=0.02, shrink=0.9)
            cbar3.set_label('GP Standard Error', fontsize=13, fontweight='bold')
            cbar3.ax.tick_params(labelsize=11)

            # Calculate uncertainty statistics
            mean_uncertainty = np.mean(Z_sem)
            max_uncertainty = np.max(Z_sem)
            min_uncertainty = np.min(Z_sem)
            ax3.text(0.05, 0.95, f'Mean: {mean_uncertainty:.3f}\\nMax: {max_uncertainty:.3f}\\nMin: {min_uncertainty:.3f}',
                    transform=ax3.transAxes, fontsize=11, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.9, edgecolor='#CCCCCC', linewidth=2))

            ax3.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax3.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax3.set_title('Model Uncertainty (GP Standard Error)',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax3.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='lower right')
            ax3.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax3.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath3 = os.path.join(directory, f'{base_name}_BO_Uncertainty_{date_str}.{file_format}')
            fig3.savefig(filepath3, dpi=dpi, bbox_inches='tight')
            plt.close(fig3)
            exported_files.append(filepath3)

            # EXPORT 4: Progress
            fig4, ax4 = plt.subplots(1, 1, figsize=(9, 7))
            response_values = self.data[self.response_column].values
            cumulative_best = np.maximum.accumulate(response_values)
            iterations = np.arange(1, len(cumulative_best) + 1)
            ax4.plot(iterations, cumulative_best, 'o-', color='#029E73', linewidth=3,
                    markersize=8, markerfacecolor='#029E73', markeredgecolor='white',
                    markeredgewidth=2, label='Best Value Found', alpha=0.9)
            ax4.fill_between(iterations, cumulative_best, alpha=0.2, color='#029E73')
            ax4.scatter([len(iterations)], [cumulative_best[-1]], c='gold', s=400,
                       marker='*', edgecolors='black', linewidth=3, zorder=5,
                       label='Current Best')
            # Set Y-axis limits based on data range (with 10% margin)
            y_min = min(response_values.min(), cumulative_best.min())
            y_max = cumulative_best.max()
            y_range = y_max - y_min
            ax4.set_ylim(y_min - 0.1 * y_range, y_max + 0.05 * y_range)

            total_improvement = cumulative_best[-1] - cumulative_best[0]
            ax4.text(0.05, 0.95, f'Total Improvement: {total_improvement:+.2f}\\nIterations: {len(iterations)}',
                    transform=ax4.transAxes, fontsize=11, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.9, edgecolor='#CCCCCC', linewidth=2))
            ax4.set_xlabel('Experiment Number', fontsize=13, fontweight='bold', color='#333333')
            ax4.set_ylabel(f'Best {self.response_column}', fontsize=13, fontweight='bold', color='#333333')
            ax4.set_title('Optimization Progress',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax4.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax4.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax4.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath4 = os.path.join(directory, f'{base_name}_BO_Progress_{date_str}.{file_format}')
            fig4.savefig(filepath4, dpi=dpi, bbox_inches='tight')
            plt.close(fig4)
            exported_files.append(filepath4)

            return exported_files

        except Exception as e:
            print(f"Error exporting BO plots: {e}")
            import traceback
            traceback.print_exc()
            return exported_files

    def _smart_column_match(self, column_name: str) -> str:
        """
        Intelligently match Excel column name to internal factor name for BO.
        Handles any factor format from ExpModel Suite.
        
        Examples:
        "NaCl (mM)" → "nacl"
        "Detergent" → "detergent" (categorical)
        "Detergent (%)" → "detergent_concentration"
        "Buffer pH" → "buffer pH"
        "Reducing Agent" → "reducing_agent" (categorical)
        "Reducing Agent (mM)" → "reducing_agent_concentration"
        """
        # Handle None or empty column names
        if column_name is None or (isinstance(column_name, str) and not column_name.strip()):
            return None

        name = column_name.strip()
        
        # Special case: Buffer pH (keep as-is for BO)
        if "Buffer pH" in name or "buffer pH" in name:
            return "buffer pH"
        
        # Extract base name by removing units
        if '(' in name:
            base_name = name.split('(')[0].strip()
        else:
            base_name = name
        
        # Normalize
        normalized = base_name.lower().replace(' ', '_').replace('-', '_')
        
        # Handle concentration suffixes
        if "buffer_conc" in normalized or "buffer conc" in base_name.lower():
            return "buffer_concentration"
        elif "detergent" in base_name.lower():
            # Check if it has units (concentration) or is categorical (name)
            if "%" in column_name or "conc" in base_name.lower():
                return "detergent_concentration"
            else:
                return "detergent"
        elif "reducing_agent" in normalized or "reducing agent" in base_name.lower():
            # Check if it has units (concentration) or is categorical (name)
            if "mM" in column_name or "conc" in base_name.lower():
                return "reducing_agent_concentration"
            else:
                return "reducing_agent"
        
        # Default: return normalized name
        return normalized
    
    def export_bo_batch_to_files(self, n_suggestions, batch_number, excel_path, 
                                 stock_concs, final_volume, buffer_ph_values):
        """Export BO suggestions to Excel and Opentrons CSV
        
        Args:
            n_suggestions: Number of BO suggestions to generate
            batch_number: Batch number for this BO iteration
            excel_path: Path to existing Excel file to append to
            stock_concs: Dict of stock concentrations {factor: concentration}
            final_volume: Final volume in µL
            buffer_ph_values: List of buffer pH values used
        
        Returns:
            Tuple of (xlsx_path, csv_path) or None if failed
        """
        try:
            import openpyxl
            import csv as csv_module
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Get BO suggestions (already rounded to 0.5 increments)
            suggestions = self.get_next_suggestions(n=n_suggestions)
            
            if not suggestions:
                print("ERROR: No suggestions returned from BO!")
                return None
            
            print(f"\nGenerating {len(suggestions)} BO suggestions...")
            print(f"First suggestion: {suggestions[0]}")
            
            # Collect pH values from BO suggestions (for dynamic CSV columns)
            bo_ph_values = []
            for suggestion in suggestions:
                # Check both possible pH column names
                for ph_key in ['Buffer pH', 'buffer pH']:
                    if ph_key in suggestion:
                        ph_value = float(suggestion[ph_key])  # Ensure float
                        if ph_value not in bo_ph_values:
                            bo_ph_values.append(ph_value)
                        break
            
            # Convert buffer_ph_values to floats and merge
            original_ph_floats = [float(ph) for ph in buffer_ph_values]
            all_ph_values = sorted(set(original_ph_floats + bo_ph_values))
            
            print(f"Original pH values: {original_ph_floats}")
            print(f"BO suggested pH values: {bo_ph_values}")
            print(f"Combined pH values for CSV: {all_ph_values}")
            
            # Read existing Excel to get last ID and structure
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Find the ACTUAL last row with data (not just max_row which includes empty rows)
            last_row_with_data = 1
            for row_num in range(1, ws.max_row + 1):
                # Check if ID column has data
                if ws.cell(row=row_num, column=1).value is not None:
                    last_row_with_data = row_num
            
            last_id = ws.cell(row=last_row_with_data, column=1).value
            next_id = int(last_id) + 1 if last_id else 1
            
            print(f"\nExcel structure:")
            print(f"  Headers: {headers}")
            print(f"  Excel max_row: {ws.max_row}")
            print(f"  Actual last row with data: {last_row_with_data}")
            print(f"  Last ID: {last_id}, Next ID: {next_id}")
            
            # Prepare new rows for Excel
            new_excel_rows = []
            volume_rows = []
            
            # Find factor column indices in Excel
            factor_col_indices = {}
            display_to_internal = {}
            response_col_idx = None
            
            for idx, header in enumerate(headers):
                # Skip None or empty headers
                if header is None or (isinstance(header, str) and not header.strip()):
                    continue
                if header == 'Response':
                    response_col_idx = idx
                    continue
                if header in ['ID', 'Plate_96', 'Well_96', 'Well_384', 'Source', 'Batch']:
                    continue

                # Use smart matching to map Excel column names to internal names
                # This will work for any factor exported by ExpModel Suite
                internal_name = self._smart_column_match(header)
                if internal_name:
                    factor_col_indices[internal_name] = idx
                    display_to_internal[header] = internal_name
            
            print(f"\nColumn mapping:")
            print(f"  Factor columns: {factor_col_indices}")
            print(f"  Response column index: {response_col_idx}")
            print(f"  Display to internal: {display_to_internal}")
            
            # Generate well positions (restart from A1 for new Opentrons run)
            
            # Collect unique values for categorical factors from all suggestions BEFORE processing
            detergent_values = set()
            reducing_agent_values = set()
            
            for sug in suggestions:
                # Check for detergent (try multiple possible keys)
                for det_key in ['detergent', 'Detergent']:
                    if det_key in sug:
                        det_val = str(sug[det_key]).strip()
                        if det_val:
                            detergent_values.add(det_val)
                        break
                
                # Check for reducing agent (try multiple possible keys)
                for agent_key in ['reducing_agent', 'Reducing Agent']:
                    if agent_key in sug:
                        agent_val = str(sug[agent_key]).strip()
                        if agent_val:
                            reducing_agent_values.add(agent_val)
                        break
            
            print(f"Detected categorical values:")
            print(f"  Detergents: {detergent_values}")
            print(f"  Reducing Agents: {reducing_agent_values}")
            
            for idx, suggestion in enumerate(suggestions):
                # Well position calculation
                plate_num = (idx // 96) + 1
                well_idx = idx % 96
                row_letter = chr(65 + (well_idx // 12))  # A-H
                col_number = (well_idx % 12) + 1
                well_pos = f"{row_letter}{col_number}"
                
                # 384-well conversion
                row_384 = chr(65 + (well_idx // 12) * 2 + (plate_num - 1) % 2)
                col_384 = (well_idx % 12) * 2 + 1 + (plate_num - 1) // 2
                well_384 = f"{row_384}{col_384}"
                
                # Build Excel row matching existing structure
                excel_row = [None] * len(headers)
                excel_row[0] = next_id + idx  # ID
                excel_row[1] = plate_num  # Plate_96
                excel_row[2] = well_pos  # Well_96
                excel_row[3] = well_384  # Well_384
                excel_row[4] = "BO"  # Source
                excel_row[5] = batch_number  # Batch
                
                # Fill in factor values from suggestions
                # BO suggestions come with original display names
                for factor_name, value in suggestion.items():
                    # Try direct match with display name
                    if factor_name in headers:
                        col_idx = headers.index(factor_name)
                        excel_row[col_idx] = value
                    # Try matching through internal name conversion
                    elif factor_name in display_to_internal:
                        internal_name = display_to_internal[factor_name]
                        if internal_name in factor_col_indices:
                            excel_row[factor_col_indices[internal_name]] = value
                
                # Response column (empty) - use the correct index
                if response_col_idx is not None:
                    excel_row[response_col_idx] = ""
                
                new_excel_rows.append(excel_row)
                
                # Calculate volumes for Opentrons CSV
                volumes = {}
                total_volume_used = 0
                
                # Handle buffer pH (categorical - one column per pH value)
                if 'buffer pH' in factor_col_indices:
                    buffer_ph_col = factor_col_indices['buffer pH']
                    buffer_ph = str(excel_row[buffer_ph_col])
                    
                    # Initialize ALL buffer pH columns to 0 (including new BO pHs)
                    for ph in all_ph_values:
                        volumes[f"buffer_{ph}"] = 0
                    
                    # Calculate volume for the specific pH used
                    if 'buffer_concentration' in factor_col_indices:
                        buffer_conc_col = factor_col_indices['buffer_concentration']
                        buffer_conc_value = excel_row[buffer_conc_col]
                        
                        if buffer_conc_value is not None:
                            desired_conc = float(buffer_conc_value)
                            buffer_stock = stock_concs.get('buffer_concentration', 0)
                            
                            if buffer_stock > 0:
                                volume = (desired_conc * final_volume) / buffer_stock
                                volumes[f"buffer_{buffer_ph}"] = round(volume, 2)
                                total_volume_used += volumes[f"buffer_{buffer_ph}"]
                
                # Handle detergent (categorical - one column per detergent type)
                if 'detergent' in factor_col_indices:
                    detergent_col = factor_col_indices['detergent']
                    detergent_type = str(excel_row[detergent_col]).strip()
                    
                    # Initialize all detergent columns to 0
                    for det in detergent_values:
                        det_clean = det.lower().replace(' ', '_').replace('-', '_')
                        volumes[det_clean] = 0
                    
                    # Calculate volume for the specific detergent used
                    if detergent_type and 'detergent_concentration' in factor_col_indices:
                        detergent_conc_col = factor_col_indices['detergent_concentration']
                        detergent_conc_value = excel_row[detergent_conc_col]
                        
                        if detergent_conc_value is not None:
                            desired_conc = float(detergent_conc_value)
                            detergent_stock = stock_concs.get('detergent_concentration', 0)
                            
                            if detergent_stock > 0:
                                volume = (desired_conc * final_volume) / detergent_stock
                                det_clean = detergent_type.lower().replace(' ', '_').replace('-', '_')
                                volumes[det_clean] = round(volume, 2)
                                total_volume_used += volumes[det_clean]
                
                # Handle reducing_agent (categorical - one column per reducing agent type)
                if 'reducing_agent' in factor_col_indices:
                    agent_col = factor_col_indices['reducing_agent']
                    agent_type = str(excel_row[agent_col]).strip()
                    
                    # Initialize all reducing agent columns to 0
                    for agent in reducing_agent_values:
                        agent_clean = agent.lower().replace(' ', '_').replace('-', '_')
                        volumes[agent_clean] = 0
                    
                    # Calculate volume for the specific reducing agent used
                    if agent_type and 'reducing_agent_concentration' in factor_col_indices:
                        agent_conc_col = factor_col_indices['reducing_agent_concentration']
                        agent_conc_value = excel_row[agent_conc_col]
                        
                        if agent_conc_value is not None:
                            desired_conc = float(agent_conc_value)
                            agent_stock = stock_concs.get('reducing_agent_concentration', 0)
                            
                            if agent_stock > 0:
                                volume = (desired_conc * final_volume) / agent_stock
                                agent_clean = agent_type.lower().replace(' ', '_').replace('-', '_')
                                volumes[agent_clean] = round(volume, 2)
                                total_volume_used += volumes[agent_clean]
                
                # Calculate volumes for other simple factors (NaCl, Zinc, Glycerol, etc.)
                for internal_name, col_idx in factor_col_indices.items():
                    if internal_name in ['buffer pH', 'buffer_concentration', 'detergent', 'detergent_concentration',
                                        'reducing_agent', 'reducing_agent_concentration']:
                        continue
                    
                    if internal_name in stock_concs:
                        factor_value = excel_row[col_idx]
                        
                        if factor_value is not None:
                            desired_conc = float(factor_value)
                            stock_conc = stock_concs[internal_name]
                            
                            if stock_conc > 0:
                                volume = (desired_conc * final_volume) / stock_conc
                                volumes[internal_name] = round(volume, 2)
                                total_volume_used += volumes[internal_name]
                
                # Calculate water
                water_volume = round(final_volume - total_volume_used, 2)
                volumes["water"] = water_volume
                
                volume_rows.append(volumes)
            
            print(f"\nGenerated {len(new_excel_rows)} new rows")
            print(f"First Excel row: {new_excel_rows[0]}")
            print(f"First volume row: {volume_rows[0]}")
            
            # Write to Excel at specific row numbers (not append!)
            start_row = last_row_with_data + 1
            print(f"\nWriting to Excel starting at row {start_row}")
            
            for idx, excel_row in enumerate(new_excel_rows):
                row_num = start_row + idx
                for col_idx, value in enumerate(excel_row, start=1):
                    ws.cell(row=row_num, column=col_idx, value=value)
                print(f"  Wrote row {row_num}: ID={excel_row[0]}")
            
            # Save Excel
            wb.save(excel_path)

            # Generate CSV path with standardized naming: [BaseName]_BO_Batch[N]_[Date]_Opentron.csv
            base_path = os.path.splitext(excel_path)[0]
            date_str = datetime.now().strftime('%Y%m%d')

            # Remove existing _Design_YYYYMMDD suffix from original file if present
            import re
            base_path = re.sub(r'_Design_\d{8}$', '', base_path)

            csv_path = f"{base_path}_BO_Batch{batch_number}_{date_str}_Opentron.csv"
            
            # Find which pH values are actually used in this batch
            used_ph_values = set()
            for volumes in volume_rows:
                for ph in all_ph_values:
                    buffer_key = f"buffer_{ph}"
                    if volumes.get(buffer_key, 0) > 0:
                        used_ph_values.add(ph)
            
            # Build CSV headers with categorical columns
            csv_headers = []
            
            # Add buffer pH columns (only used ones)
            if used_ph_values:
                for ph in sorted(used_ph_values):
                    csv_headers.append(f"buffer_{ph}")
            
            # Add detergent columns (all unique detergent types)
            if detergent_values:
                for det in sorted(detergent_values):
                    det_clean = det.lower().replace(' ', '_').replace('-', '_')
                    csv_headers.append(det_clean)
            
            # Add reducing agent columns (all unique reducing agent types)
            if reducing_agent_values:
                for agent in sorted(reducing_agent_values):
                    agent_clean = agent.lower().replace(' ', '_').replace('-', '_')
                    csv_headers.append(agent_clean)
            
            # Add other simple factors (skip categorical factors and their concentrations)
            for internal_name in factor_col_indices.keys():
                if internal_name not in ['buffer pH', 'buffer_concentration', 'detergent', 'detergent_concentration',
                                        'reducing_agent', 'reducing_agent_concentration']:
                    csv_headers.append(internal_name)
            
            csv_headers.append("water")
            
            # Write CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv_module.writer(f)
                writer.writerow(csv_headers)
                
                for volumes in volume_rows:
                    row = [volumes.get(h, 0) for h in csv_headers]
                    writer.writerow(row)
            
            print(f"\n✓ Successfully exported:")
            print(f"  Excel: {excel_path}")
            print(f"  CSV: {csv_path}")
            print(f"  Rows added: {len(new_excel_rows)}")
            
            return excel_path, csv_path
            
        except Exception as e:
            print(f"Error exporting BO batch: {e}")
            import traceback
            traceback.print_exc()
            return None

# GUI Application

class AnalysisTab(ttk.Frame):
    """Main GUI Application - Analysis Tab"""

    def __init__(self, parent, project, main_window):
        super().__init__(parent)
        self.project = project
        self.main_window = main_window

        # Initialize components
        self.handler = DataHandler()
        self.analyzer = DoEAnalyzer()
        self.plotter = DoEPlotter()
        self.exporter = ResultsExporter()
        self.optimizer = BayesianOptimizer() if AX_AVAILABLE else None

        # Data storage
        self.filepath = None
        self.results = None
        self.main_effects = None

        # Setup GUI
        self.setup_gui()
        
    def setup_gui(self):
        """Setup GUI layout"""
        
        # Title
        title_frame = ttk.Frame(self)
        title_frame.pack(fill='x', padx=10, pady=5)

        # File selection
        file_frame = ttk.LabelFrame(self, text="1. Select Data File", padding=10)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side='left', padx=5)
        
        browse_btn = ttk.Button(file_frame, text="Browse...", command=self.browse_file)
        browse_btn.pack(side='right', padx=5)
        
        # Configuration
        config_frame = ttk.LabelFrame(self, text="2. Configure Analysis", padding=10)
        config_frame.pack(fill='x', padx=10, pady=5)
        
        # Model Type row with info button
        ttk.Label(config_frame, text="Model Type:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        
        self.model_type_var = tk.StringVar(value='linear')
        model_combo = ttk.Combobox(config_frame, textvariable=self.model_type_var,
                                   values=['linear', 'interactions', 'purequadratic', 'quadratic'],
                                   state='readonly', width=25)
        model_combo.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # info button
        info_btn = ttk.Button(config_frame, text="?", width=2, command=self.show_model_guide)
        info_btn.grid(row=0, column=2, sticky='w', padx=(2, 0), pady=5)
        
        self.analyze_btn = ttk.Button(config_frame, text="Analyze Data", 
                                      command=self.analyze_data, state='disabled')
        self.analyze_btn.grid(row=0, column=3, padx=20, pady=5)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self, textvariable=self.status_var, 
                              relief=tk.SUNKEN, anchor='w')
        status_bar.pack(fill='x', side='bottom')
        
        # Export buttons
        export_frame = ttk.LabelFrame(self, text="4. Export Results", padding=10)
        export_frame.pack(fill='x', side='bottom', padx=10, pady=5)
        
        self.export_stats_btn = ttk.Button(export_frame, text="Export Statistics (.xlxs)",
                                          command=self.export_statistics, state='disabled')
        self.export_stats_btn.pack(side='left', padx=5)
        
        self.export_plots_btn = ttk.Button(export_frame, text="Export Plots (.png)",
                                          command=self.export_plots, state='disabled')
        self.export_plots_btn.pack(side='left', padx=5)
        
        # Results notebook
        results_frame = ttk.LabelFrame(self, text="3. Results", padding=5)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # It was needed to create a canvas with scrollbar for results to prevent expanding too much
        results_container = ttk.Frame(results_frame)
        results_container.pack(fill='both', expand=True)
        
        self.notebook = ttk.Notebook(results_container)
        self.notebook.pack(fill='both', expand=True)
        
        # Tab 1: Statistics
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistics")
        
        self.stats_text = scrolledtext.ScrolledText(self.stats_frame, wrap=tk.WORD, font=('Courier', 14))
        self.stats_text.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Tab 2: Main Effects
        main_effects_container = ttk.Frame(self.notebook)
        self.notebook.add(main_effects_container, text="Main Effects")
        
        # Tooltip button for main effects
        me_header = ttk.Frame(main_effects_container)
        me_header.pack(fill='x', padx=5, pady=2)
        ttk.Label(me_header, text="Main Effects Plot", font=('TkDefaultFont', 10, 'bold')).pack(side='left')
        ttk.Button(me_header, text="ℹ️ How to Read", width=15,
                  command=lambda: self.show_tooltip("main_effects")).pack(side='right', padx=5)
        
        # Scrollable frame for main effects plot
        self.main_effects_frame = self.create_scrollable_frame(main_effects_container)
        
        # Tab 3: Interactions
        interactions_container = ttk.Frame(self.notebook)
        self.notebook.add(interactions_container, text="Interactions")
        
        # Tooltip button for interactions
        int_header = ttk.Frame(interactions_container)
        int_header.pack(fill='x', padx=5, pady=2)
        ttk.Label(int_header, text="Interaction Effects Plot", font=('TkDefaultFont', 10, 'bold')).pack(side='left')
        ttk.Button(int_header, text="ℹ️ How to Read", width=15,
                  command=lambda: self.show_tooltip("interactions")).pack(side='right', padx=5)
        
        # Scrollable frame for interactions plot
        self.interactions_frame = self.create_scrollable_frame(interactions_container)
        
        # Tab 4: Residuals
        residuals_container = ttk.Frame(self.notebook)
        self.notebook.add(residuals_container, text="Residuals")
        
        # Tooltip button for residuals
        res_header = ttk.Frame(residuals_container)
        res_header.pack(fill='x', padx=5, pady=2)
        ttk.Label(res_header, text="Residuals Diagnostic Plot", font=('TkDefaultFont', 10, 'bold')).pack(side='left')
        ttk.Button(res_header, text="ℹ️ How to Read", width=15,
                  command=lambda: self.show_tooltip("residuals")).pack(side='right', padx=5)
        
        # Scrollable frame for residuals plot
        self.residuals_frame = self.create_scrollable_frame(residuals_container)
        
        # Tab 5: Recommendations
        self.recommendations_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.recommendations_frame, text="Recommendations")

        # Add button frame at top for BO batch export
        if AX_AVAILABLE:
            button_frame = ttk.Frame(self.recommendations_frame)
            button_frame.pack(fill='x', padx=5, pady=5)

            self.export_bo_button = ttk.Button(button_frame, text="📤 Export BO Batch to Files",
                                              command=self.export_bo_batch, state='disabled')
            self.export_bo_button.pack(side='left', padx=5)

            ttk.Label(button_frame, text="(Available after analysis with BO suggestions)",
                     font=('TkDefaultFont', 9)).pack(side='left', padx=5)

        self.recommendations_text = scrolledtext.ScrolledText(self.recommendations_frame,
                                                             wrap=tk.WORD, font=('Courier', 14))
        self.recommendations_text.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Tab 6: Optimization Details (Bayesian Optimization plot)
        if AX_AVAILABLE:
            optimization_container = ttk.Frame(self.notebook)
            self.notebook.add(optimization_container, text="Optimization Details")

            # Tooltip button for optimization and export button
            opt_header = ttk.Frame(optimization_container)
            opt_header.pack(fill='x', padx=5, pady=2)
            ttk.Label(opt_header, text="Bayesian Optimization Analysis",
                     font=('TkDefaultFont', 10, 'bold')).pack(side='left')

            # Export BO Plots button
            self.export_bo_plots_button = ttk.Button(opt_header, text="📊 Export BO Plots",
                                                     command=self.export_bo_plots_gui, state='disabled')
            self.export_bo_plots_button.pack(side='right', padx=5)

            ttk.Button(opt_header, text="ℹ️ How to Read", width=15,
                      command=lambda: self.show_tooltip("optimization")).pack(side='right', padx=5)

            # Scrollable frame for optimization plot
            self.optimization_frame = self.create_scrollable_frame(optimization_container)
    
    def create_scrollable_frame(self, parent):
        """Create a scrollable frame for plot display with comprehensive mousewheel support"""
        import platform
        
        # Canvas with scrollbars and white background
        canvas = tk.Canvas(parent, bg='white', highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(parent, orient='vertical', command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(parent, orient='horizontal', command=canvas.xview)
        
        # Create frame inside canvas
        scrollable_frame = ttk.Frame(canvas)
        
        # Store canvas reference in frame
        scrollable_frame._scroll_canvas = canvas
        
        # Canvas config
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scrollable_frame.bind("<Configure>", on_frame_configure)

        canvas.create_window((0, 0), window=scrollable_frame, anchor="center")
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Configure smooth scrolling
        canvas.configure(yscrollincrement='10')  # Scroll with 10 pixels at a time for smooth but faster scrolling
        
        # Pack scrollbars and canvas
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')
        canvas.pack(side='left', fill='both', expand=True)
        
        # MOUSEWHEEL SOLUTION (Mousewheel wasn't being detected, this is the patch)
        # Detect platform
        system = platform.system()
        
        # Mousewheel handler
        def on_mousewheel(event):
            # Use smaller scroll increments for smooth scrolling
            if system == 'Windows':
                # Windows: delta is typically ±120, divide further for smoothness
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            elif system == 'Darwin':  # macOS
                # macOS: smaller delta values, use directly but scale down
                canvas.yview_scroll(int(-1*event.delta), "units")
            else:  # Linux/Unix
                if event.num == 4:
                    canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    canvas.yview_scroll(1, "units")
        
        # Recursive binding function to bind all widgets
        def bind_tree(widget):
            """Recursively bind mousewheel to widget and all its children"""
            if system == 'Windows' or system == 'Darwin':
                widget.bind('<MouseWheel>', on_mousewheel, add='+')
            else:  # Linux
                widget.bind('<Button-4>', on_mousewheel, add='+')
                widget.bind('<Button-5>', on_mousewheel, add='+')
            
            # Recursively bind to all children
            for child in widget.winfo_children():
                bind_tree(child)
        
        # Bind to canvas
        if system == 'Windows' or system == 'Darwin':
            canvas.bind('<MouseWheel>', on_mousewheel)
        else:  # Linux
            canvas.bind('<Button-4>', on_mousewheel)
            canvas.bind('<Button-5>', on_mousewheel)
        
        # Bind to scrollable frame
        bind_tree(scrollable_frame)
        
        # Store binding function for later use when new widgets are added
        scrollable_frame._bind_mousewheel = bind_tree

        def on_canvas_configure(event):
            # Center the window
            canvas_width = event.width
            frame_width = scrollable_frame.winfo_reqwidth()
            x_position = max(0, (canvas_width - frame_width) // 2)
            if canvas.find_withtag("all"):
                canvas.coords(canvas.find_withtag("all")[0], x_position, 0)
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        return scrollable_frame
    
    def show_tooltip(self, plot_type):
        """Show interpretation guidance for different plot types"""
        tooltips = {
            "main_effects": (
                "How to Read Main Effects Plot\n\n"
                "• Steeper slopes = Stronger factor influence on response\n"
                "• Upward slope = Increasing factor increases response\n"
                "• Downward slope = Increasing factor decreases response\n"
                "• Flat line = Factor has little/no effect\n"
                "• Compare slopes to identify most important factors\n\n"
                "Example: If Salt has a steep upward slope and Buffer pH is flat,\n"
                "Salt is more important for your response."
            ),
            "interactions": (
                "How to Read Interaction Plot\n\n"
                "• Parallel lines = NO interaction (factors work independently)\n"
                "• Non-parallel lines = Interaction present (factors affect each other)\n"
                "• Crossing lines = Strong interaction (optimal level depends on other factor)\n\n"
                "Example: If glycerol lines cross for different salt levels,\n"
                "the best glycerol amount depends on which salt level you use."
            ),
            "residuals": (
                "How to Read Residuals Plot\n\n"
                "• Random scatter around zero = Good model fit ✓\n"
                "• Pattern/curve = Model may be inappropriate ✗\n"
                "• Funnel shape = Variance increases with response ✗\n"
                "• Outliers far from others = Check those data points\n\n"
                "If residuals look bad, your statistical conclusions may be unreliable.\n"
                "Consider trying a different model type or checking for data errors."
            ),
            "optimization": (
                "How to Read Optimization Details\n\n"
                "This plot shows where Bayesian Optimization predicts you should explore:\n\n"
                "• Hot colors (yellow/green) = Predicted high response regions\n"
                "• Cool colors (blue/purple) = Predicted low response regions\n"
                "• Red dots = Your existing experiments\n"
                "• Gaps between dots = Unexplored regions to test\n\n"
                "The algorithm learns from your data to suggest intelligent next experiments\n"
                "that balance exploring new areas with exploiting promising regions."
            )
        }
        
        messagebox.showinfo("Plot Interpretation Guide", tooltips.get(plot_type, "No guide available"))
    
    def show_model_guide(self):
        """Show model selection guide in a popup window"""
        guide_text = (
            "Linear: Simplest model, main effects only (A + B + C)\n"
            "  • Use when: You want to understand which factors matter\n"
            "  • Assumes: Each factor works independently\n"
            "  • Example: Salt effect + Buffer effect\n\n"
            
            "Interactions: Use if factors work together (A + B + A×B)\n"
            "  • Use when: Factors may influence each other\n"
            "  • Detects: Combined effects of multiple factors\n"
            "  • Example: Salt effect depends on Buffer level\n\n"
            
            "Pure Quadratic: For curved responses (A + B + A² + B²)\n"
            "  • Use when: Response has a peak or valley\n"
            "  • Assumes: No interactions, but curved relationships\n"
            "  • Example: Optimum temperature between two extremes\n\n"
            
            "Quadratic: Full model (A + B + A×B + A² + B²)\n"
            "  • Use when: You expect both curvature and interactions\n"
            "  • Most complex: Requires more data points\n"
            "  • Example: Complex biological systems with multiple effects"
        )
        
        messagebox.showinfo("Model Selection Guide", guide_text)
    
    def browse_file(self):
        """Open file dialog"""
        filepath = filedialog.askopenfilename(
            title="Select DoE Data File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filepath:
            self.filepath = filepath
            # Keep label white/colored when file selected
            self.file_label.config(text=os.path.basename(filepath))
            self.status_var.set(f"Loaded: {os.path.basename(filepath)}")
            
            try:
                self.handler.load_excel(filepath)
                
                # Check if "Response" column exists
                if 'Response' not in self.handler.data.columns:
                    messagebox.showerror("Error", 
                                       "Excel file must have a column named 'Response'\n\n"
                                       "Please rename your response column to 'Response' and try again.")
                    self.status_var.set("Error: No 'Response' column found")
                    return
                
                # Enable analyze button
                self.analyze_btn.config(state='normal')
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load file:\n{str(e)}")
                self.status_var.set("Error loading file")
    
    def analyze_data(self):
        """Perform DoE analysis"""
        if not self.filepath:
            messagebox.showwarning("Warning", "Please select a data file first.")
            return
        
        self.status_var.set("Analyzing...")
        self.update()
        
        try:
            # Always use "Response" column in the xlxs file
            response_col = "Response"
            model_type = self.model_type_var.get()
            # Detect which columns are factors vs response
            self.handler.detect_columns(response_col)
            clean_data = self.handler.preprocess_data()
            
            # Pass cleaned data to analyzer
            self.analyzer.set_data(
                data=clean_data,
                factor_columns=self.handler.factor_columns,
                categorical_factors=self.handler.categorical_factors,
                numeric_factors=self.handler.numeric_factors,
                response_column=self.handler.response_column
            )
            
            # Fit regression model then calculations are made
            self.results = self.analyzer.fit_model(model_type)
            self.main_effects = self.analyzer.calculate_main_effects()
            
            self.display_statistics()
            self.display_plots()
            self.display_recommendations()
            
            # Display optimization plot if available
            if AX_AVAILABLE and self.optimizer:
                self.display_optimization_plot()
            
            self.export_stats_btn.config(state='normal')
            self.export_plots_btn.config(state='normal')
            
            self.status_var.set(f"Analysis complete! R² = {self.results['model_stats']['R-squared']:.4f}")
            
            messagebox.showinfo("Success", 
                              f"Analysis completed successfully!\n\n"
                              f"Model: {model_type}\n"
                              f"Observations: {self.results['model_stats']['Observations']}\n"
                              f"R-squared: {self.results['model_stats']['R-squared']:.4f}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Analysis failed:\n{str(e)}")
            self.status_var.set("Analysis failed")
            import traceback
            traceback.print_exc()
    
    def display_statistics(self):
        """Display statistical results with recommendations and warnings"""
        self.stats_text.delete('1.0', tk.END)
        
        self.stats_text.insert(tk.END, "="*80 + "\n")
        self.stats_text.insert(tk.END, "DOE ANALYSIS RESULTS\n")
        self.stats_text.insert(tk.END, "="*80 + "\n\n")
        
        # RED FLAGS SECTION - Show warnings first if model quality is poor
        r_squared = self.results['model_stats']['R-squared']
        n_obs = self.results['model_stats']['Observations']
        sig_factors = self.analyzer.get_significant_factors()
        
        warnings = []
        if r_squared < 0.5:
            warnings.append(f"⚠️  LOW R² ({r_squared:.3f}): Model explains only {r_squared*100:.1f}% of variance; inspect residuals for unexplained structure.")
        elif r_squared < 0.7:
            warnings.append(f"⚠️  MODERATE R² ({r_squared:.3f}): Model is acceptable but could be improved.")
        
        if len(sig_factors) == 0:
            warnings.append("⚠️  NO SIGNIFICANT FACTORS: No factors with p < 0.05 found. Consider reviewing experimental design.")
        
        if n_obs < 20:
            warnings.append(f"⚠️  SMALL SAMPLE SIZE: Only {n_obs} observations. Consider adding replicates for more robust results.")
        
        if warnings:
            self.stats_text.insert(tk.END, "🚩 RED FLAGS / WARNINGS\n")
            self.stats_text.insert(tk.END, "-"*80 + "\n")
            for warning in warnings:
                self.stats_text.insert(tk.END, f"{warning}\n")
            self.stats_text.insert(tk.END, "\n" + "="*80 + "\n\n")
        
        # MODEL STATISTICS
        self.stats_text.insert(tk.END, "MODEL STATISTICS\n")
        self.stats_text.insert(tk.END, "-"*80 + "\n")
        for key, value in self.results['model_stats'].items():
            if isinstance(value, float):
                self.stats_text.insert(tk.END, f"  {key:<25}: {value:>15.6f}\n")
            else:
                self.stats_text.insert(tk.END, f"  {key:<25}: {value:>15}\n")
        
        self.stats_text.insert(tk.END, "\n" + "="*80 + "\n")
        self.stats_text.insert(tk.END, "COEFFICIENTS\n")
        self.stats_text.insert(tk.END, "-"*80 + "\n\n")
        
        coef_str = self.results['coefficients'].to_string()
        self.stats_text.insert(tk.END, coef_str + "\n")
        
        # SIGNIFICANT FACTORS
        self.stats_text.insert(tk.END, "\n" + "="*80 + "\n")
        self.stats_text.insert(tk.END, f"SIGNIFICANT FACTORS (p < 0.05): {len(sig_factors)} found\n")
        self.stats_text.insert(tk.END, "-"*80 + "\n")
        
        if sig_factors:
            # Sort by absolute coefficient value to rank importance
            factor_importance = []
            for factor in sig_factors:
                coef = self.results['coefficients'].loc[factor, 'Coefficient']
                pval = self.results['coefficients'].loc[factor, 'p-value']
                factor_importance.append((factor, abs(coef), coef, pval))
            
            # Rank by effect magnitude
            factor_importance.sort(key=lambda x: x[1], reverse=True)
            
            self.stats_text.insert(tk.END, "Ranked by effect size (most important first):\n\n")
            for rank, (factor, abs_coef, coef, pval) in enumerate(factor_importance, 1):
                self.stats_text.insert(tk.END, f"  {rank}. {factor:<38} coef={coef:>10.4f}  p={pval:.2e}\n")
        else:
            self.stats_text.insert(tk.END, "  None found - no factors are statistically significant.\n")
        
        # INTERACTION DETECTION
        interaction_factors = [f for f in sig_factors if ':' in f]
        if interaction_factors:
            self.stats_text.insert(tk.END, "\n" + "="*80 + "\n")
            self.stats_text.insert(tk.END, f"⚡ SIGNIFICANT INTERACTIONS DETECTED: {len(interaction_factors)}\n")
            self.stats_text.insert(tk.END, "-"*80 + "\n")
            for interaction in interaction_factors:
                coef = self.results['coefficients'].loc[interaction, 'Coefficient']
                pval = self.results['coefficients'].loc[interaction, 'p-value']
                self.stats_text.insert(tk.END, f"  {interaction:<40} coef={coef:>10.4f}  p={pval:.2e}\n")
            self.stats_text.insert(tk.END, "\n⚠️  Interactions mean optimal settings depend on factor combinations!\n")
        
        self.stats_text.insert(tk.END, "\n" + "="*80 + "\n")
    
    def display_plots(self):
        """Display all plots"""
        self.plotter.set_data(
            self.handler.clean_data,
            self.handler.factor_columns,
            self.handler.response_column
        )
        
        self.display_main_effects_plot()
        self.display_interaction_plot()
        self.display_residuals_plot()
    
    def display_main_effects_plot(self):
        """Display main effects plot"""
        for widget in self.main_effects_frame.winfo_children():
            widget.destroy()
        
        fig = self.plotter.plot_main_effects()
        canvas = FigureCanvasTkAgg(fig, master=self.main_effects_frame)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill='both', expand=True)
        
        # Bind mousewheel to the matplotlib canvas widget
        if hasattr(self.main_effects_frame, '_bind_mousewheel'):
            self.main_effects_frame._bind_mousewheel(canvas_widget)
        
        # Reset scroll position to top
        if hasattr(self.main_effects_frame, '_scroll_canvas'):
            self.main_effects_frame._scroll_canvas.update_idletasks()
            self.main_effects_frame._scroll_canvas.yview_moveto(0)
        
        plt.close(fig)
    
    def display_interaction_plot(self):
        """Display interaction plot"""
        for widget in self.interactions_frame.winfo_children():
            widget.destroy()
        
        fig = self.plotter.plot_interaction_effects()
        if fig:
            canvas = FigureCanvasTkAgg(fig, master=self.interactions_frame)
            canvas.draw()
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.pack(fill='both', expand=True)
            
            # Bind mousewheel to the matplotlib canvas widget
            if hasattr(self.interactions_frame, '_bind_mousewheel'):
                self.interactions_frame._bind_mousewheel(canvas_widget)
            
            # Reset scroll position to top
            if hasattr(self.interactions_frame, '_scroll_canvas'):
                self.interactions_frame._scroll_canvas.update_idletasks()
                self.interactions_frame._scroll_canvas.yview_moveto(0)
            
            plt.close(fig)
    
    def display_residuals_plot(self):
        """Display residuals plot"""
        for widget in self.residuals_frame.winfo_children():
            widget.destroy()
        
        fig = self.plotter.plot_residuals(
            self.results['predictions'],
            self.results['residuals']
        )
        canvas = FigureCanvasTkAgg(fig, master=self.residuals_frame)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill='both', expand=True)
        
        # Bind mousewheel to the matplotlib canvas widget
        if hasattr(self.residuals_frame, '_bind_mousewheel'):
            self.residuals_frame._bind_mousewheel(canvas_widget)
        
        # Reset scroll position to top
        if hasattr(self.residuals_frame, '_scroll_canvas'):
            self.residuals_frame._scroll_canvas.update_idletasks()
            self.residuals_frame._scroll_canvas.yview_moveto(0)
        
        plt.close(fig)
    
    def display_recommendations(self):
        """Display recommendations and optimal conditions"""
        self.recommendations_text.delete('1.0', tk.END)
        
        self.recommendations_text.insert(tk.END, "="*80 + "\n")
        self.recommendations_text.insert(tk.END, "RECOMMENDATIONS & OPTIMAL CONDITIONS\n")
        self.recommendations_text.insert(tk.END, "="*80 + "\n\n")
        
        # Determine confidence level
        r_squared = self.results['model_stats']['R-squared']
        n_obs = self.results['model_stats']['Observations']
        sig_factors = self.analyzer.get_significant_factors()
        
        # Determine confidence level based on model quality metrics
        if r_squared >= 0.8 and len(sig_factors) > 0 and n_obs >= 20:
            confidence = "HIGH"
        elif r_squared >= 0.6 and len(sig_factors) > 0:
            confidence = "MEDIUM"
        else:
            confidence = "LOW"
        
        self.recommendations_text.insert(tk.END, f"CONFIDENCE LEVEL: {confidence}\n")
        self.recommendations_text.insert(tk.END, "-"*80 + "\n")
        self.recommendations_text.insert(tk.END, f"Based on R² = {r_squared:.3f}, {n_obs} observations, {len(sig_factors)} significant factors\n\n")
        
        # Find optimal condition from data (row with highest response)
        clean_data = self.handler.clean_data
        max_idx = clean_data[self.handler.response_column].idxmax()
        optimal_response = clean_data.loc[max_idx, self.handler.response_column]
        
        self.recommendations_text.insert(tk.END, "="*80 + "\n")
        self.recommendations_text.insert(tk.END, "BEST OBSERVED CONDITION (from your data)\n")
        self.recommendations_text.insert(tk.END, "-"*80 + "\n")
        self.recommendations_text.insert(tk.END, f"Response Value: {optimal_response:.2f}\n\n")
        
        for factor in self.handler.factor_columns:
            value = clean_data.loc[max_idx, factor]
            self.recommendations_text.insert(tk.END, f"  • {factor:<30}: {value}\n")
        
        # Predicted optimal based on model
        self.recommendations_text.insert(tk.END, "\n" + "="*80 + "\n")
        self.recommendations_text.insert(tk.END, "MODEL-PREDICTED OPTIMAL DIRECTION\n")
        self.recommendations_text.insert(tk.END, "-"*80 + "\n")
        
        if len(sig_factors) > 0:
            # Get only main effect factors (not interactions)
            main_sig_factors = [f for f in sig_factors if ':' not in f]
            
            if main_sig_factors:
                self.recommendations_text.insert(tk.END, "To INCREASE response, adjust these significant factors:\n\n")
                for factor in main_sig_factors:
                    coef = self.results['coefficients'].loc[factor, 'Coefficient']
                    # Positive coef = increase factor to increase response
                    if coef > 0:
                        direction = "INCREASE"
                    else:
                        direction = "DECREASE"
                    
                    # Clean factor name (remove C() and Q() wrappers)
                    clean_factor = factor.replace("C(Q('", "").replace("'))", "").replace("Q('", "").replace("')", "")
                    self.recommendations_text.insert(tk.END, f"  • {clean_factor:<30}: {direction}  (effect: {coef:+.4f})\n")
            
            interaction_factors = [f for f in sig_factors if ':' in f]
            if interaction_factors:
                self.recommendations_text.insert(tk.END, f"\nWARNING: {len(interaction_factors)} interaction(s) detected!\n")
                self.recommendations_text.insert(tk.END, "Optimal levels depend on factor combinations - see Interactions plot.\n")
        else:
            self.recommendations_text.insert(tk.END, "No significant factors found. Consider:\n")
            self.recommendations_text.insert(tk.END, "  • Testing wider factor ranges\n")
            self.recommendations_text.insert(tk.END, "  • Adding more factors to the experiment\n")
            self.recommendations_text.insert(tk.END, "  • Checking measurement accuracy\n")
        
        # Next steps recommendations
        self.recommendations_text.insert(tk.END, "\n" + "="*80 + "\n")
        self.recommendations_text.insert(tk.END, "NEXT STEPS\n")
        self.recommendations_text.insert(tk.END, "-"*80 + "\n")
        
        if confidence == "HIGH":
            self.recommendations_text.insert(tk.END, 
                "1. Run 3-5 confirmation experiments at the predicted optimal condition\n"
                "2. Compare results to model prediction to validate\n"
                "3. If confirmed, implement optimized condition in production\n"
            )
        elif confidence == "MEDIUM":
            self.recommendations_text.insert(tk.END, 
                "1. Run confirmation experiments at predicted optimal condition\n"
                "2. Consider additional replicates to improve model confidence\n"
                "3. May need to refine factor ranges or add more data\n"
            )
        else:
            self.recommendations_text.insert(tk.END, 
                "1. WARNING: Results may not be reliable enough for immediate use\n"
                "2. Consider running more experiments\n"
                "3. Check for experimental errors or measurement issues\n"
                "4. May need to reconsider factors or expand factor ranges\n"
            )
        
        self.recommendations_text.insert(tk.END, "\n" + "="*80 + "\n")
        
        # Add Bayesian Optimization suggestions if available
        if AX_AVAILABLE and self.optimizer:
            self.recommendations_text.insert(tk.END, "\n" + "="*80 + "\n")
            self.recommendations_text.insert(tk.END, "BAYESIAN OPTIMIZATION SUGGESTIONS\n")
            self.recommendations_text.insert(tk.END, "="*80 + "\n")
            self.recommendations_text.insert(tk.END, "Intelligently suggested next experiments (balancing exploration & exploitation):\n\n")
            
            try:
                # Initialize optimizer with current data
                self.optimizer.set_data(
                    data=self.handler.clean_data,
                    factor_columns=self.handler.factor_columns,
                    categorical_factors=self.handler.categorical_factors,
                    numeric_factors=self.handler.numeric_factors,
                    response_column=self.handler.response_column
                )
                self.optimizer.initialize_optimizer(minimize=False)  # Assume maximize response
                
                # Get suggestions
                suggestions = self.optimizer.get_next_suggestions(n=5)
                
                for i, suggestion in enumerate(suggestions, 1):
                    self.recommendations_text.insert(tk.END, f"Suggested Experiment #{i}:\n")
                    for factor, value in suggestion.items():
                        if isinstance(value, float):
                            self.recommendations_text.insert(tk.END, f"  • {factor:<30}: {value:.4f}\n")
                        else:
                            self.recommendations_text.insert(tk.END, f"  • {factor:<30}: {value}\n")
                    self.recommendations_text.insert(tk.END, "\n")
                
                self.recommendations_text.insert(tk.END, 
                    "💡 TIP: These suggestions use machine learning to predict where to test next.\n"
                    "   View 'Optimization Details' tab for visualization of predicted response surface.\n"
                )
                
                # Enable export button after successful BO initialization
                if hasattr(self, 'export_bo_button'):
                    self.export_bo_button.config(state='normal')
                if hasattr(self, 'export_bo_plots_button'):
                    self.export_bo_plots_button.config(state='normal')
                
            except Exception as e:
                self.recommendations_text.insert(tk.END, 
                    f"Could not generate BO suggestions: {str(e)}\n"
                    "This may require more data points or only numeric factors.\n"
                )
                # Disable export button if BO failed
                if hasattr(self, 'export_bo_button'):
                    self.export_bo_button.config(state='disabled')
                if hasattr(self, 'export_bo_plots_button'):
                    self.export_bo_plots_button.config(state='disabled')
            
            self.recommendations_text.insert(tk.END, "\n" + "="*80 + "\n")
    
    def display_optimization_plot(self):
        """Display Bayesian Optimization predicted response surface"""
        if not AX_AVAILABLE or not self.optimizer:
            return
        
        for widget in self.optimization_frame.winfo_children():
            widget.destroy()
        
        try:
            # Check if we have enough numeric factors
            num_numeric = len(self.handler.numeric_factors)
            
            if num_numeric < 2:
                # Show message if plot can't be generated
                message_label = ttk.Label(
                    self.optimization_frame,
                    text=f"Optimization plot requires at least 2 numeric factors.\n"
                         f"Your data has {num_numeric} numeric factor(s).\n\n"
                         f"Check the Recommendations tab for suggested experiments.",
                    font=('TkDefaultFont', 12),
                    justify='center'
                )
                message_label.pack(expand=True)
                return
            
            fig = self.optimizer.get_acquisition_plot()
            
            if fig is None:
                # Show message if plot generation failed
                message_label = ttk.Label(
                    self.optimization_frame,
                    text="Could not generate optimization plot.\n"
                         "Check the Recommendations tab for suggested experiments.",
                    font=('TkDefaultFont', 12),
                    justify='center'
                )
                message_label.pack(expand=True)
                return
            
            canvas = FigureCanvasTkAgg(fig, master=self.optimization_frame)
            canvas.draw()
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.pack(fill='both', expand=True)
            
            # Bind mousewheel to the matplotlib canvas widget
            if hasattr(self.optimization_frame, '_bind_mousewheel'):
                self.optimization_frame._bind_mousewheel(canvas_widget)
            
            # Reset scroll position to top
            if hasattr(self.optimization_frame, '_scroll_canvas'):
                self.optimization_frame._scroll_canvas.update_idletasks()
                self.optimization_frame._scroll_canvas.yview_moveto(0)
            
            plt.close(fig)
            
        except Exception as e:
            error_label = ttk.Label(
                self.optimization_frame,
                text=f"Could not generate optimization plot:\n{str(e)}\n\n"
                     "This feature works best with 2+ numeric factors and sufficient data.\n"
                     "Check the Recommendations tab for suggested experiments.",
                font=('TkDefaultFont', 10),
                justify='center',
                wraplength=600
            )
            error_label.pack(expand=True, pady=20)
    
    def export_statistics(self):
        """Export statistics to Excel"""
        date_str = datetime.now().strftime('%Y%m%d')

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Analysis1.xlsx"
        )

        if filepath:
            try:
                # Generate path with naming convention: [UserName]_Statistics_[Date]
                base_path = os.path.splitext(filepath)[0]

                # Add standardized suffix
                final_path = f"{base_path}_Statistics_{date_str}.xlsx"

                self.exporter.set_results(self.results, self.main_effects)
                self.exporter.export_statistics_excel(final_path)
                messagebox.showinfo("Success", f"Statistics exported to:\n{final_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed:\n{str(e)}")
    
    def _prompt_for_stock_concentrations(self):
        """Prompt user for stock concentrations when metadata is not available"""
        # Create dialog
        stock_dialog = tk.Toplevel(self)
        stock_dialog.title("Enter Stock Concentrations")
        stock_dialog.geometry("500x400")
        stock_dialog.transient(self)
        stock_dialog.grab_set()
        
        ttk.Label(stock_dialog, text="Stock Concentrations Not Found in File",
                 font=('TkDefaultFont', 12, 'bold')).pack(pady=10)
        ttk.Label(stock_dialog, text="Please enter stock concentrations for each factor:",
                 font=('TkDefaultFont', 10)).pack(pady=5)
        
        # Frame for entries
        entries_frame = ttk.Frame(stock_dialog, padding=20)
        entries_frame.pack(fill='both', expand=True)
        
        # Create entry for each numeric factor
        stock_vars = {}
        row = 0
        for factor in self.handler.numeric_factors:
            # Display name and unit
            if 'buffer' in factor.lower() and 'conc' in factor.lower():
                display_name = "Buffer Concentration"
                unit = "mM"
                default = "1000"
            elif 'salt' in factor.lower():
                display_name = factor.replace('_', ' ').title()
                unit = "mM"
                default = "5000"
            elif 'glycerol' in factor.lower():
                display_name = "Glycerol"
                unit = "%"
                default = "100"
            elif 'dmso' in factor.lower():
                display_name = "DMSO"
                unit = "%"
                default = "100"
            elif 'detergent' in factor.lower():
                display_name = "Detergent"
                unit = "%"
                default = "10"
            else:
                display_name = factor.replace('_', ' ').title()
                unit = ""
                default = ""
            
            ttk.Label(entries_frame, text=f"{display_name}:").grid(row=row, column=0, sticky='e', padx=5, pady=5)
            var = tk.StringVar(value=default)
            entry = ttk.Entry(entries_frame, textvariable=var, width=15)
            entry.grid(row=row, column=1, padx=5, pady=5)
            ttk.Label(entries_frame, text=unit).grid(row=row, column=2, sticky='w', padx=5, pady=5)
            
            stock_vars[factor] = var
            row += 1
        
        result = {'confirmed': False, 'stocks': {}}
        
        def confirm():
            try:
                stocks = {}
                for factor, var in stock_vars.items():
                    value_str = var.get().strip()
                    if value_str:
                        stocks[factor] = float(value_str)
                
                result['confirmed'] = True
                result['stocks'] = stocks
                stock_dialog.destroy()
            except ValueError:
                messagebox.showerror("Invalid Input", "Please enter valid numeric values for all fields.")
        
        def cancel():
            stock_dialog.destroy()
        
        # Buttons
        button_frame = ttk.Frame(stock_dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Confirm", command=confirm).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=cancel).pack(side='left', padx=5)
        
        # Wait for dialog to close
        self.wait_window(stock_dialog)
        
        if result['confirmed']:
            return result['stocks']
        else:
            return None
    
    def _ask_plot_format(self):
        """Ask user for plot export format"""
        dialog = tk.Toplevel(self)
        dialog.title("Select Plot Format")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()

        result = {'format': None, 'dpi': 300}

        ttk.Label(dialog, text="Select Export Format", font=('TkDefaultFont', 12, 'bold')).pack(pady=10)

        # Format selection
        format_frame = ttk.LabelFrame(dialog, text="File Format", padding=10)
        format_frame.pack(fill='x', padx=20, pady=10)

        format_var = tk.StringVar(value="png")

        formats = [
            ("PNG - Good for presentations, web (300 DPI)", "png"),
            ("TIFF - Publication quality, lossless (300 DPI)", "tiff"),
            ("PDF - Vector format, scalable", "pdf"),
            ("EPS - Vector format, publication standard", "eps")
        ]

        for text, value in formats:
            ttk.Radiobutton(format_frame, text=text, variable=format_var, value=value).pack(anchor='w', pady=2)

        # DPI selection (only for raster formats)
        dpi_frame = ttk.LabelFrame(dialog, text="Resolution (PNG/TIFF only)", padding=10)
        dpi_frame.pack(fill='x', padx=20, pady=10)

        dpi_var = tk.IntVar(value=300)

        dpi_options = [
            ("300 DPI - Standard publication quality", 300),
            ("600 DPI - High quality publication", 600),
            ("150 DPI - Screen/web quality", 150)
        ]

        for text, value in dpi_options:
            ttk.Radiobutton(dpi_frame, text=text, variable=dpi_var, value=value).pack(anchor='w', pady=2)

        def confirm():
            result['format'] = format_var.get()
            result['dpi'] = dpi_var.get()
            dialog.destroy()

        def cancel():
            dialog.destroy()

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="OK", command=confirm).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=cancel).pack(side='left', padx=5)

        self.wait_window(dialog)
        return result

    def export_plots(self):
        """Export plots in user-selected format"""
        # Ask for format first
        format_options = self._ask_plot_format()
        if not format_options['format']:
            return

        file_format = format_options['format']
        dpi = format_options['dpi']
        date_str = datetime.now().strftime('%Y%m%d')

        # File type mapping for dialog
        format_map = {
            'png': ("PNG files", "*.png"),
            'tiff': ("TIFF files", "*.tiff"),
            'pdf': ("PDF files", "*.pdf"),
            'eps': ("EPS files", "*.eps")
        }

        # Ask user for base name
        base_name = filedialog.asksaveasfilename(
            defaultextension=f".{file_format}",
            filetypes=[format_map[file_format], ("All files", "*.*")],
            initialfile=f"Plots1.{file_format}",
            title="Choose base name for plots (will create multiple files)"
        )

        if not base_name:
            return

        # Extract directory and base name without extension
        directory = os.path.dirname(base_name)
        base_path = os.path.splitext(os.path.basename(base_name))[0]

        try:
            # Export main effects plot
            fig1 = self.plotter.plot_main_effects()
            fig1.savefig(os.path.join(directory, f'{base_path}_MainEffects_{date_str}.{file_format}'),
                        dpi=dpi, bbox_inches='tight')
            plt.close(fig1)

            # Export interactions plot
            fig2 = self.plotter.plot_interaction_effects()
            if fig2:
                fig2.savefig(os.path.join(directory, f'{base_path}_Interactions_{date_str}.{file_format}'),
                            dpi=dpi, bbox_inches='tight')
                plt.close(fig2)

            # Export residuals plot
            fig3 = self.plotter.plot_residuals(
                self.results['predictions'],
                self.results['residuals']
            )
            fig3.savefig(os.path.join(directory, f'{base_path}_Residuals_{date_str}.{file_format}'),
                        dpi=dpi, bbox_inches='tight')
            plt.close(fig3)

            messagebox.showinfo("Success", f"Plots exported to:\n{directory}\n\n"
                              f"Format: {file_format.upper()}, DPI: {dpi if file_format in ['png', 'tiff'] else 'Vector'}\n\n"
                              f"Files created:\n"
                              f"- {base_path}_MainEffects_{date_str}.{file_format}\n"
                              f"- {base_path}_Interactions_{date_str}.{file_format}\n"
                              f"- {base_path}_Residuals_{date_str}.{file_format}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
    
    def export_bo_batch(self):
        """Export BO suggestions to Excel and Opentrons CSV"""
        if not AX_AVAILABLE or not self.optimizer or not self.optimizer.is_initialized:
            messagebox.showerror("Error", "Bayesian Optimization not available or not initialized.")
            return
        
        if not self.filepath:
            messagebox.showerror("Error", "No Excel file loaded. Please load data first.")
            return
        
        # Dialog to get parameters
        dialog = tk.Toplevel(self)
        dialog.title("Export BO Batch")
        dialog.geometry("500x400")
        dialog.transient(self)
        dialog.grab_set()
        
        # Number of suggestions
        ttk.Label(dialog, text="Number of BO Suggestions:", font=('TkDefaultFont', 10, 'bold')).pack(pady=(20,5))
        n_var = tk.IntVar(value=5)
        n_spin = ttk.Spinbox(dialog, from_=1, to=20, textvariable=n_var, width=10)
        n_spin.pack()
        
        # Batch number
        ttk.Label(dialog, text="Batch Number:", font=('TkDefaultFont', 10, 'bold')).pack(pady=(15,5))
        
        # Auto-detect next batch number from loaded data
        try:
            if 'Batch' in self.handler.data.columns:
                max_batch = int(self.handler.data['Batch'].max())
                next_batch = max_batch + 1
            else:
                next_batch = 1
        except:
            next_batch = 1
        
        batch_var = tk.IntVar(value=next_batch)
        batch_frame = ttk.Frame(dialog)
        batch_frame.pack()
        ttk.Label(batch_frame, text=f"Suggested: {next_batch} (next after current data)").pack(side='left', padx=5)
        batch_spin = ttk.Spinbox(batch_frame, from_=1, to=100, textvariable=batch_var, width=10)
        batch_spin.pack(side='left')
        
        # Final volume
        ttk.Label(dialog, text="Final Volume (µL):", font=('TkDefaultFont', 10, 'bold')).pack(pady=(15,5))
        vol_var = tk.DoubleVar(value=100.0)
        vol_entry = ttk.Entry(dialog, textvariable=vol_var, width=15)
        vol_entry.pack()
        
        # Stock concentrations note
        stock_concs_from_metadata = self.handler.get_stock_concentrations()
        if stock_concs_from_metadata:
            ttk.Label(dialog, text="✓ Using stock concentrations from file metadata",
                     font=('TkDefaultFont', 9, 'bold'), foreground='green').pack(pady=(15,5))
        else:
            ttk.Label(dialog, text="⚠ No metadata found - will prompt for stock concentrations",
                     font=('TkDefaultFont', 9), foreground='orange').pack(pady=(15,5))
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        
        def do_export():
            try:
                n_suggestions = n_var.get()
                batch_number = batch_var.get()
                final_volume = vol_var.get()
                
                # Get stock concentrations from metadata first
                stock_concs = self.handler.get_stock_concentrations()
                
                # If no metadata, show dialog to ask user
                if not stock_concs:
                    stock_concs = self._prompt_for_stock_concentrations()
                    if stock_concs is None:
                        # User cancelled
                        return
                
                # Get buffer pH values from data
                buffer_ph_values = []
                if 'buffer pH' in self.handler.data.columns or 'Buffer pH' in self.handler.data.columns:
                    ph_col = 'buffer pH' if 'buffer pH' in self.handler.data.columns else 'Buffer pH'
                    unique_phs = sorted(self.handler.data[ph_col].dropna().unique())
                    buffer_ph_values = [str(ph) for ph in unique_phs]
                
                # Export
                result = self.optimizer.export_bo_batch_to_files(
                    n_suggestions=n_suggestions,
                    batch_number=batch_number,
                    excel_path=self.filepath,
                    stock_concs=stock_concs,
                    final_volume=final_volume,
                    buffer_ph_values=buffer_ph_values
                )
                
                if result:
                    xlsx_path, csv_path = result
                    dialog.destroy()
                    messagebox.showinfo("Export Successful!",
                        f"BO Batch exported successfully!\n\n"
                        f"📊 Excel updated:\n{xlsx_path}\n\n"
                        f"🤖 Opentrons CSV:\n{csv_path}\n\n"
                        f"Batch {batch_number}: {n_suggestions} new experiments added")
                else:
                    messagebox.showerror("Error", "Export failed. Check console for details.")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Export failed:\n{str(e)}")
        
        ttk.Button(button_frame, text="Export", command=do_export).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def export_bo_plots_gui(self):
        """GUI wrapper for exporting BO plots"""
        if not AX_AVAILABLE or not self.optimizer or not self.optimizer.is_initialized:
            messagebox.showerror("Error", "Bayesian Optimization not available or not initialized.")
            return

        # Ask for format first
        format_options = self._ask_plot_format()
        if not format_options['format']:
            return

        file_format = format_options['format']
        dpi = format_options['dpi']
        date_str = datetime.now().strftime('%Y%m%d')

        # File type mapping for dialog
        format_map = {
            'png': ("PNG files", "*.png"),
            'tiff': ("TIFF files", "*.tiff"),
            'pdf': ("PDF files", "*.pdf"),
            'eps': ("EPS files", "*.eps")
        }

        # Ask user for base name
        base_name = filedialog.asksaveasfilename(
            defaultextension=f".{file_format}",
            filetypes=[format_map[file_format], ("All files", "*.*")],
            initialfile=f"Plots1.{file_format}",
            title="Choose base name for BO plots (will create multiple files)"
        )

        if not base_name:
            return

        # Extract directory and base name
        directory = os.path.dirname(base_name)
        base_path = os.path.splitext(os.path.basename(base_name))[0]

        try:
            exported_files = self.optimizer.export_bo_plots(directory, base_path, date_str, file_format, dpi)

            if exported_files:
                filenames = "\\n".join([os.path.basename(f) for f in exported_files])
                messagebox.showinfo("Success",
                                  f"Exported {len(exported_files)} BO plots\\n\\n"
                                  f"Format: {file_format.upper()}, DPI: {dpi if file_format in ['png', 'tiff'] else 'Vector'}\\n\\n"
                                  f"{filenames}\\n\\nLocation: {directory}")
            else:
                messagebox.showwarning("Warning", "No plots were exported. Check console for details.")

        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\\n{str(e)}")


    def load_results(self):
        """Load results from Excel (called from main window menu)"""
        self.browse_file()

    def refresh(self):
        """Refresh UI from project data (called when switching tabs)"""
        # Reload display if data exists
        if self.handler.data is not None:
            self._update_display()
