"""Export service - generates Excel and CSV files in memory matching original SCOUT format"""

import io
from typing import Dict, List, Optional, Tuple

import pandas as pd
from utils.constants import AVAILABLE_FACTORS
from config.design_config import CATEGORICAL_FACTORS
from core.volume_calculator import VolumeCalculator
from core.well_mapper import WellMapper

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _extract_unit(display_name: str) -> str:
    """Extract unit from display name, e.g. 'NaCl (mM)' -> 'mM'"""
    if "(" in display_name and ")" in display_name:
        start = display_name.rfind("(") + 1
        end = display_name.rfind(")")
        return display_name[start:end]
    return ""


def build_bo_volume_data(
    suggestions: List[Dict],
    stock_concs: Dict[str, float],
    per_level_concs: Optional[Dict],
    final_volume: float,
    batch_number: int,
    existing_data: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Convert raw BO suggestions into (excel_df, volume_df) for Opentron-compatible export.

    Reuses VolumeCalculator for C1V1=C2V2 and WellMapper for well positions,
    matching the format produced by DoEDesigner.build_factorial_design().
    """
    calculator = VolumeCalculator()
    mapper = WellMapper()

    # Auto-detect next ID from existing data
    start_id = 1
    if existing_data is not None and "ID" in existing_data.columns:
        start_id = int(existing_data["ID"].max()) + 1

    # Collect all unique buffer pH values across suggestions + existing data
    buffer_ph_values = set()
    for s in suggestions:
        if "buffer pH" in s:
            buffer_ph_values.add(str(s["buffer pH"]).strip())
    if existing_data is not None and "buffer pH" in existing_data.columns:
        for v in existing_data["buffer pH"].dropna().unique():
            buffer_ph_values.add(str(v).strip())
    # Also check display name column
    if existing_data is not None and "Buffer pH" in existing_data.columns:
        for v in existing_data["Buffer pH"].dropna().unique():
            buffer_ph_values.add(str(v).strip())
    buffer_ph_values = sorted(buffer_ph_values)

    # Collect all unique categorical values for decomposition headers
    all_detergents = set()
    all_agents = set()
    for s in suggestions:
        if "detergent" in s:
            all_detergents.add(str(s["detergent"]))
        if "reducing_agent" in s:
            all_agents.add(str(s["reducing_agent"]))

    # Build volume headers (same order as factorial design)
    volume_headers = ["ID"]
    for ph in buffer_ph_values:
        volume_headers.append(f"buffer_{ph}")

    # Determine which numeric factors appear in suggestions (excluding categoricals/concentrations)
    skip_factors = set(CATEGORICAL_FACTORS) | {"buffer_concentration", "detergent_concentration", "reducing_agent_concentration"}
    numeric_factor_names = []
    for key in suggestions[0]:
        if key not in skip_factors:
            numeric_factor_names.append(key)

    # Detergent decomposed columns
    for det in sorted(all_detergents):
        norm = VolumeCalculator._normalize_factor_name(det)
        if norm not in volume_headers:
            volume_headers.append(norm)

    # Reducing agent decomposed columns
    for agent in sorted(all_agents):
        norm = VolumeCalculator._normalize_factor_name(agent)
        if norm not in volume_headers:
            volume_headers.append(norm)

    # Regular numeric factors
    for fn in numeric_factor_names:
        if fn not in volume_headers and fn not in ("buffer pH",):
            volume_headers.append(fn)

    volume_headers.append("water")

    # Build excel headers (Sample Tracking format)
    excel_headers = ["ID", "Plate_96", "Well_96", "Well_384", "Source", "Batch"]
    factor_names = list(suggestions[0].keys())
    for fn in factor_names:
        display = AVAILABLE_FACTORS.get(fn, fn)
        if display not in excel_headers:
            excel_headers.append(display)
    excel_headers.append("Response")

    # Calculate volumes for each suggestion
    excel_rows = []
    volume_rows = []
    for idx, s in enumerate(suggestions):
        sample_id = start_id + idx
        plate_96, well_96, well_384 = mapper.generate_well_position_384_order(idx)

        # Build factor_values dict (convert numeric strings)
        factor_values = {}
        for k, v in s.items():
            try:
                factor_values[k] = float(v)
            except (ValueError, TypeError):
                factor_values[k] = v

        volumes = calculator.calculate_volumes(
            factor_values,
            stock_concs,
            final_volume,
            buffer_ph_values if buffer_ph_values else None,
            per_level_concs=per_level_concs,
        )

        # Excel row
        excel_row = {
            "ID": sample_id,
            "Plate_96": plate_96,
            "Well_96": well_96,
            "Well_384": well_384,
            "Source": "BO",
            "Batch": batch_number,
        }
        for fn in factor_names:
            display = AVAILABLE_FACTORS.get(fn, fn)
            excel_row[display] = s[fn]
        excel_row["Response"] = ""
        excel_rows.append(excel_row)

        # Volume row
        vol_row = {"ID": sample_id}
        for h in volume_headers:
            if h == "ID":
                continue
            vol_row[h] = round(volumes.get(h, 0.0), 2)
        volume_rows.append(vol_row)

    excel_df = pd.DataFrame(excel_rows, columns=excel_headers)
    volume_df = pd.DataFrame(volume_rows, columns=volume_headers)

    return excel_df, volume_df


def generate_excel_bytes(
    excel_df: pd.DataFrame,
    volume_df: pd.DataFrame,
    stock_concs: Dict[str, float],
    project_name: str = "Design",
    per_level_concs: Optional[Dict] = None,
    protein_stock: Optional[float] = None,
    protein_final: Optional[float] = None,
    final_volume: float = 200.0,
) -> bytes:
    """Generate Excel file as bytes matching real SCOUT export format.

    Creates:
    - Sheet 1: 'Sample Tracking' - design data with well positions
    - Sheet 2: 'Stock_Concentrations' - 5-column format matching original
    - Sheet 3: 'Reagent Setup Guide' - reservoir positions and total volumes
    """
    output = io.BytesIO()

    wb = openpyxl.Workbook()

    # --- Sheet 1: Sample Tracking ---
    ws = wb.active
    ws.title = "Sample Tracking"

    headers = list(excel_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

    for row_idx, (_, row) in enumerate(excel_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            try:
                numeric = float(value)
                ws.cell(row=row_idx, column=col_idx, value=numeric)
            except (ValueError, TypeError):
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths for Sample Tracking
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # --- Sheet 2: Stock_Concentrations (5-column unified format) ---
    stock_ws = wb.create_sheet(title="Stock_Concentrations")

    stock_headers = ["Factor Name", "Level", "Stock Value", "Final Value", "Unit"]
    for col_idx, header in enumerate(stock_headers, 1):
        cell = stock_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

    row_idx = 2

    # Normal stock concentrations (Level column empty)
    for factor_name, stock_value in stock_concs.items():
        display_name = AVAILABLE_FACTORS.get(factor_name, factor_name)
        unit = _extract_unit(display_name)

        stock_ws.cell(row=row_idx, column=1, value=display_name)
        stock_ws.cell(row=row_idx, column=2, value="")
        stock_ws.cell(row=row_idx, column=3, value=stock_value)
        stock_ws.cell(row=row_idx, column=4, value="")
        stock_ws.cell(row=row_idx, column=5, value=unit)
        row_idx += 1

    # Per-level concentrations (Level column filled)
    if per_level_concs:
        for factor_name, level_concs in per_level_concs.items():
            if not level_concs:
                continue

            if factor_name == "detergent":
                conc_factor = "detergent_concentration"
                unit = "%"
            elif factor_name == "reducing_agent":
                conc_factor = "reducing_agent_concentration"
                unit = "mM"
            else:
                continue

            display_name = AVAILABLE_FACTORS.get(conc_factor, conc_factor)

            for level, conc_data in level_concs.items():
                stock_ws.cell(row=row_idx, column=1, value=display_name)
                stock_ws.cell(row=row_idx, column=2, value=level)
                stock_ws.cell(row=row_idx, column=3, value=conc_data.get("stock", ""))
                stock_ws.cell(row=row_idx, column=4, value=conc_data.get("final", ""))
                stock_ws.cell(row=row_idx, column=5, value=unit)
                row_idx += 1

    # Protein info (if provided)
    if protein_stock and protein_final and protein_stock > 0:
        row_idx += 1  # Separator row
        protein_vol = round((protein_final * final_volume) / protein_stock, 2)

        stock_ws.cell(row=row_idx, column=1, value="Protein (added manually)")
        stock_ws.cell(row=row_idx, column=2, value="")
        stock_ws.cell(row=row_idx, column=3, value=protein_stock)
        stock_ws.cell(row=row_idx, column=4, value=protein_final)
        stock_ws.cell(row=row_idx, column=5, value="mg/mL")
        for c in range(1, 6):
            stock_ws.cell(row=row_idx, column=c).fill = PatternFill(
                start_color="81C784", end_color="81C784", fill_type="solid"
            )
        row_idx += 1

        stock_ws.cell(row=row_idx, column=1, value="→ Volume per well")
        stock_ws.cell(row=row_idx, column=2, value="")
        stock_ws.cell(row=row_idx, column=3, value=protein_vol)
        stock_ws.cell(row=row_idx, column=4, value="")
        stock_ws.cell(row=row_idx, column=5, value="µL")

    # Auto-adjust column widths for Stock_Concentrations
    for col in stock_ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 30)
        stock_ws.column_dimensions[col_letter].width = adjusted_width

    # --- Sheet 3: Reagent Setup Guide ---
    _create_reagent_setup_guide(wb, volume_df, stock_concs, per_level_concs)

    wb.save(output)
    output.seek(0)
    return output.read()


def _create_reagent_setup_guide(
    workbook,
    volume_df: pd.DataFrame,
    stock_concs: Dict[str, float],
    per_level_concs: Optional[Dict] = None,
) -> None:
    """Create Reagent Setup Guide sheet showing reservoir positions and total volumes.

    Matches the original SCOUT Tkinter export: one row per reagent with
    reservoir position, stock concentration, total volume needed, and
    volume with 20% overage for the Cytiva 24-position reservoir.
    """
    guide_sheet = workbook.create_sheet(title="Reagent Setup Guide")

    # Cytiva 24 reservoir positions (column-wise: A1, B1, C1, D1, A2, B2...)
    reservoir_positions = []
    for col in range(1, 7):
        for row in ['A', 'B', 'C', 'D']:
            reservoir_positions.append(f"{row}{col}")

    # Headers
    headers = ["Position", "Reagent", "Stock Concentration", "Volume Needed", "Add to Reservoir (with 20% overage)"]
    for col_idx, header in enumerate(headers, 1):
        cell = guide_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

    row_idx = 2
    position_idx = 0

    for col_name in volume_df.columns:
        # Skip ID column
        if col_name.upper() == "ID":
            continue

        # Calculate total volume for this reagent
        total_volume = volume_df[col_name].sum()

        # Add 20% overage
        volume_with_overage = round(total_volume * 1.20, 1)

        # Determine stock concentration display
        stock_display = "N/A"

        # Buffer pH columns (buffer_7.5, buffer_8.5, etc.)
        if col_name.startswith("buffer_"):
            if "buffer_concentration" in stock_concs:
                stock_display = f"{stock_concs['buffer_concentration']} mM"

        # Check per-level concentrations (detergent types, reducing agents)
        elif per_level_concs:
            col_normalized = col_name.lower().replace(' ', '_').replace('-', '_')
            if "detergent" in per_level_concs:
                for level, conc_data in per_level_concs["detergent"].items():
                    level_normalized = level.lower().replace(' ', '_').replace('-', '_')
                    if level_normalized == col_normalized:
                        stock_display = f"{conc_data.get('stock', 'N/A')}%"
                        break
            if stock_display == "N/A" and "reducing_agent" in per_level_concs:
                for level, conc_data in per_level_concs["reducing_agent"].items():
                    level_normalized = level.lower().replace(' ', '_').replace('-', '_')
                    if level_normalized == col_normalized:
                        stock_display = f"{conc_data.get('stock', 'N/A')} mM"
                        break

        # Regular numeric factor
        if stock_display == "N/A" and col_name in stock_concs:
            stock_val = stock_concs[col_name]
            display_name = AVAILABLE_FACTORS.get(col_name, col_name)
            unit = _extract_unit(display_name)
            stock_display = f"{stock_val} {unit}".strip()

        # Water has no stock concentration
        if col_name == "water":
            stock_display = "—"

        # Build display name
        reagent_display = col_name.replace('_', ' ').title()
        if col_name.startswith("buffer_"):
            ph_value = col_name.replace("buffer_", "")
            reagent_display = f"Buffer pH {ph_value}"

        # Get reservoir position
        position = reservoir_positions[position_idx] if position_idx < len(reservoir_positions) else f"Extra-{position_idx}"

        # Write row
        guide_sheet.cell(row=row_idx, column=1, value=position)
        guide_sheet.cell(row=row_idx, column=2, value=reagent_display)
        guide_sheet.cell(row=row_idx, column=3, value=stock_display)
        guide_sheet.cell(row=row_idx, column=4, value=f"{total_volume:.1f} µL")
        guide_sheet.cell(row=row_idx, column=5, value=f"{volume_with_overage:.1f} µL")

        row_idx += 1
        position_idx += 1

    # Auto-adjust column widths
    for col in guide_sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 40)
        guide_sheet.column_dimensions[col_letter].width = adjusted_width


def generate_csv_bytes(volume_df: pd.DataFrame) -> bytes:
    """Generate Opentrons-compatible CSV as bytes"""
    output = io.StringIO()
    volume_df.to_csv(output, index=False)
    return output.getvalue().encode("utf-8")
