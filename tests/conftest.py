"""Shared test fixtures"""
import pytest
import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
import openpyxl
from openpyxl.styles import Font


@pytest.fixture
def sample_doe_data():
    """Create sample DoE data for testing"""
    return pd.DataFrame({
        'ID': [1, 2, 3, 4, 5],
        'Plate_96': ['Plate1', 'Plate1', 'Plate1', 'Plate1', 'Plate1'],
        'Well_96': ['A1', 'A2', 'A3', 'A4', 'A5'],
        'NaCl': [50, 100, 150, 50, 100],
        'pH': [6.0, 7.0, 8.0, 6.5, 7.5],
        'buffer': ['Tris', 'Tris', 'HEPES', 'Tris', 'HEPES'],
        'Response': [0.5, 0.7, 0.9, 0.6, 0.8]
    })


@pytest.fixture
def sample_doe_data_with_missing():
    """Create sample DoE data with missing values"""
    return pd.DataFrame({
        'ID': [1, 2, 3, 4, 5],
        'NaCl': [50, 100, np.nan, 50, 100],
        'pH': [6.0, 7.0, 8.0, np.nan, 7.5],
        'buffer': ['Tris', 'Tris', None, 'Tris', 'HEPES'],
        'Response': [0.5, 0.7, 0.9, np.nan, 0.8]
    })


@pytest.fixture
def temp_excel_file(sample_doe_data, tmp_path):
    """Create temporary Excel file with sample data"""
    filepath = tmp_path / "test_data.xlsx"
    sample_doe_data.to_excel(filepath, index=False)
    return filepath


@pytest.fixture
def temp_excel_with_stock_concs(tmp_path):
    """Create temporary Excel file matching real SCOUT export format.

    Matches the structure of actual design files:
    - 'Sample Tracking' sheet with ID, Plate_96, Well_96, Well_384, Source, Batch, factors, Response
    - 'Stock_Concentrations' sheet with Factor Name, Level, Stock Value, Final Value, Unit
      including per-level entries for categorical factors (e.g. Detergent)
    """
    filepath = tmp_path / "test_data_with_stocks.xlsx"

    wb = openpyxl.Workbook()

    # --- Sample Tracking sheet (main data) ---
    ws = wb.active
    ws.title = "Sample Tracking"

    headers = [
        "ID", "Plate_96", "Well_96", "Well_384", "Source", "Batch",
        "Buffer pH", "Buffer Conc (mM)", "NaCl (mM)", "Glycerol (%)",
        "Detergent", "Reducing Agent", "Reducing Agent (mM)", "Response",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    rows = [
        [1, 1, "A1", "A1", "FULL_FACTORIAL", 0, 6.5, 10, 100, 0, "C12E8", "TCEP", 0, 36.0],
        [2, 1, "B1", "C1", "FULL_FACTORIAL", 0, 6.5, 10, 100, 0, "C12E8", "TCEP", 0.5, 36.4],
        [3, 1, "C1", "E1", "FULL_FACTORIAL", 0, 6.5, 10, 100, 0, "CHAPS", "TCEP", 0, 32.4],
        [4, 1, "D1", "G1", "FULL_FACTORIAL", 0, 6.5, 10, 100, 0, "CHAPS", "TCEP", 0.5, 32.0],
        [5, 1, "E1", "I1", "FULL_FACTORIAL", 0, 7.5, 10, 500, 15, "DDM", "TCEP", 0, 40.1],
        [6, 1, "F1", "K1", "FULL_FACTORIAL", 0, 7.5, 10, 500, 15, "DDM", "TCEP", 0.5, 41.3],
        [7, 1, "G1", "M1", "FULL_FACTORIAL", 0, 8.5, 50, 100, 0, "C12E8", "TCEP", 0, 38.2],
        [8, 1, "H1", "O1", "FULL_FACTORIAL", 0, 8.5, 50, 500, 15, "CHAPS", "TCEP", 0.5, 44.5],
    ]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # --- Stock_Concentrations sheet ---
    stock_ws = wb.create_sheet(title="Stock_Concentrations")
    stock_headers = ["Factor Name", "Level", "Stock Value", "Final Value", "Unit"]
    for c, h in enumerate(stock_headers, 1):
        cell = stock_ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    # Normal factors (Level is empty)
    # Per-level factors (Level is filled, e.g. detergent types)
    stock_rows = [
        # [Factor Name, Level, Stock Value, Final Value, Unit]
        ["Buffer Conc (mM)", None, 100, None, "mM"],
        ["NaCl (mM)", None, 2000, None, "mM"],
        ["Glycerol (%)", None, 50, None, "%"],
        ["Reducing Agent (mM)", None, 5, None, "mM"],
        # Per-level detergent concentrations
        ["Detergent (%)", "C12E8", 0.05, 0.00336, "%"],
        ["Detergent (%)", "CHAPS", 5, 0.34, "%"],
        ["Detergent (%)", "DDM", 0.1, 0.00609, "%"],
        # Separator / empty row
        [None, None, None, None, None],
        # Protein (informational, skipped by loader)
        ["Protein (added manually)", None, 29.75, 0.5, "mg/mL"],
    ]

    for r_idx, row in enumerate(stock_rows, 2):
        for c_idx, val in enumerate(row, 1):
            if val is not None:
                stock_ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(filepath)
    return filepath


@pytest.fixture
def real_ff_template():
    """Path to real FF design template (if available in examples/)"""
    path = Path(__file__).parent.parent / "examples" / "FF_Design_Template.xlsx"
    if path.exists():
        return path
    pytest.skip("FF_Design_Template.xlsx not found in examples/")


@pytest.fixture
def real_lhs_template():
    """Path to real LHS design template (if available in examples/)"""
    path = Path(__file__).parent.parent / "examples" / "LHS_Design_Template.xlsx"
    if path.exists():
        return path
    pytest.skip("LHS_Design_Template.xlsx not found in examples/")


@pytest.fixture
def sample_factorial_design():
    """Create sample full factorial design"""
    return pd.DataFrame({
        'Factor1': [10, 10, 20, 20],
        'Factor2': [100, 200, 100, 200],
        'Response': [0.5, 0.7, 0.6, 0.9]
    })


@pytest.fixture
def sample_analysis_results():
    """Create sample analysis results"""
    return {
        'model_type': 'main_effects',
        'aic': 45.2,
        'bic': 48.7,
        'r_squared': 0.85,
        'adj_r_squared': 0.82,
        'predictions': np.array([0.51, 0.69, 0.61, 0.88]),
        'residuals': np.array([-0.01, 0.01, -0.01, 0.02]),
        'formula': 'Response ~ Factor1 + Factor2'
    }
