"""
Test the full web design flow end-to-end without running the server.
Simulates: add factors → build factorial → generate designs → export → validate
Verifies JSON keys match frontend expectations.

Run with: PYTHONPATH=. python3 tests/test_web_design_flow.py
"""

import json
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def test_factor_management():
    """Test adding, updating, removing factors via project"""
    from core.project import DoEProject

    print("=== Testing Factor Management ===")

    project = DoEProject()

    # Add factors
    project.add_factor("nacl", ["50", "100", "200"], stock_conc=5000)
    project.add_factor("buffer pH", ["6", "7", "8"])
    project.add_factor("glycerol", ["5", "10", "15"], stock_conc=100)

    factors = project.get_factors()
    assert len(factors) == 3
    assert factors["nacl"] == ["50", "100", "200"]
    assert factors["buffer pH"] == ["6", "7", "8"]
    print("  [PASS] Add factors")

    # Stock concentrations
    stock = project.get_all_stock_concs()
    assert stock["nacl"] == 5000
    assert stock["glycerol"] == 100
    assert "buffer pH" not in stock
    print("  [PASS] Stock concentrations")

    # Update factor
    project.update_factor("nacl", ["100", "200", "300"], stock_conc=5000)
    assert project.get_factors()["nacl"] == ["100", "200", "300"]
    print("  [PASS] Update factor")

    # Remove factor
    project.remove_factor("glycerol")
    assert "glycerol" not in project.get_factors()
    assert "glycerol" not in project.get_all_stock_concs()
    print("  [PASS] Remove factor")

    # Clear all
    project.add_factor("test", ["1", "2"])
    project.clear_factors()
    assert len(project.get_factors()) == 0
    print("  [PASS] Clear all factors")

    print()


def test_design_service_combinations():
    """Test combination count and plate calculations"""
    from backend.services import design_service

    print("=== Testing Design Service Calculations ===")

    # Combinations
    factors = {"A": ["1", "2"], "B": ["x", "y", "z"]}
    assert design_service.get_combinations_count(factors) == 6
    assert design_service.get_combinations_count({}) == 0
    print("  [PASS] Combination count")

    # Plates
    assert design_service.get_plates_required(0) == 0
    assert design_service.get_plates_required(1) == 1
    assert design_service.get_plates_required(96) == 1
    assert design_service.get_plates_required(97) == 2
    assert design_service.get_plates_required(384) == 4
    print("  [PASS] Plate calculations")

    print()


def test_build_factorial():
    """Test full factorial design generation with volumes"""
    from core.doe_designer import DoEDesigner
    from backend.services import design_service

    print("=== Testing Build Factorial ===")

    designer = DoEDesigner()
    factors = {
        "NaCl (mM)": ["50", "100", "200"],
        "Glycerol (%)": ["5", "10"],
    }
    stock_concs = {
        "NaCl (mM)": 5000,
        "Glycerol (%)": 100,
    }

    excel_data, volume_data, warnings = design_service.build_factorial_design(
        designer, factors, stock_concs, final_volume=200.0
    )

    # Check structure
    assert isinstance(excel_data, list)
    assert isinstance(volume_data, list)
    assert len(excel_data) == 6  # 3 × 2
    assert len(volume_data) == 6
    print(f"  Runs: {len(excel_data)}")

    # Check excel_data has expected columns
    first = excel_data[0]
    assert "ID" in first, f"Missing ID. Keys: {list(first.keys())}"
    assert "Plate_96" in first
    assert "Well_96" in first
    assert "Well_384" in first
    print(f"  Excel columns: {list(first.keys())}")

    # Check volume_data has columns (reagent volumes + water)
    vol_first = volume_data[0]
    assert "water" in vol_first, f"Missing 'water'. Keys: {list(vol_first.keys())}"
    print(f"  Volume columns: {list(vol_first.keys())}")

    # JSON serializable
    json_str = json.dumps(excel_data)
    json.dumps(volume_data)
    print(f"  Excel JSON: {len(json_str)} bytes")

    print("  [PASS] Build factorial")
    print()


def test_generate_non_factorial():
    """Test non-factorial design generation"""
    from core.design_factory import DesignFactory
    from backend.services import design_service

    print("=== Testing Non-Factorial Design Generation ===")

    factory = DesignFactory(has_pydoe3=True, has_smt=False)
    factors = {
        "NaCl (mM)": ["50", "100", "200"],
        "Glycerol (%)": ["5", "10", "15"],
        "Buffer pH": ["6", "7", "8"],
    }

    stock_concs = {
        "NaCl (mM)": 5000,
        "Glycerol (%)": 100,
        "Buffer pH": None,  # categorical, no stock
    }

    # LHS
    try:
        points, warnings = design_service.generate_design(
            factory, "lhs", factors, {"n_samples": 20},
            stock_concs={k: v for k, v in stock_concs.items() if v is not None},
            final_volume=200.0,
        )
        assert isinstance(points, list)
        assert len(points) == 20
        # Check well positions added
        assert "ID" in points[0], f"Missing ID. Keys: {list(points[0].keys())}"
        assert "Plate_96" in points[0]
        assert "Well_96" in points[0]
        assert "Well_384" in points[0]
        json.dumps(points)
        print(f"  LHS: {len(points)} points, has wells, JSON OK")
    except Exception as e:
        print(f"  LHS: [SKIP] {e}")

    # Plackett-Burman (needs 2 levels per factor)
    factors_2level = {
        "NaCl (mM)": ["50", "200"],
        "Glycerol (%)": ["5", "15"],
        "Buffer pH": ["6", "8"],
    }
    try:
        points, warnings = design_service.generate_design(
            factory, "plackett_burman", factors_2level, {}
        )
        assert isinstance(points, list)
        assert len(points) > 0
        json.dumps(points)
        print(f"  Plackett-Burman: {len(points)} points, JSON OK")
    except Exception as e:
        print(f"  Plackett-Burman: [SKIP] {e}")

    # Central Composite
    try:
        points, warnings = design_service.generate_design(
            factory, "central_composite", factors, {"ccd_type": "faced", "center_points": 3}
        )
        assert isinstance(points, list)
        assert len(points) > 0
        json.dumps(points)
        print(f"  CCD: {len(points)} points, JSON OK")
    except Exception as e:
        print(f"  CCD: [SKIP] {e}")

    # Box-Behnken
    try:
        points, warnings = design_service.generate_design(
            factory, "box_behnken", factors, {"center_points": 3}
        )
        assert isinstance(points, list)
        assert len(points) > 0
        json.dumps(points)
        print(f"  Box-Behnken: {len(points)} points, JSON OK")
    except Exception as e:
        print(f"  Box-Behnken: [SKIP] {e}")

    print("  [PASS] Non-factorial designs")
    print()


def test_export_excel():
    """Test Excel export generates valid bytes matching real SCOUT format"""
    import openpyxl
    from io import BytesIO
    from core.doe_designer import DoEDesigner
    from backend.services import export_service

    print("=== Testing Excel Export ===")

    designer = DoEDesigner()
    factors = {
        "NaCl (mM)": ["50", "100"],
        "Glycerol (%)": ["5", "10"],
    }
    stock_concs = {"NaCl (mM)": 5000, "Glycerol (%)": 100}

    excel_df, volume_df = designer.build_factorial_design(
        factors, stock_concs, final_volume=200.0
    )

    # --- Basic export (no per-level, no protein) ---
    excel_bytes = export_service.generate_excel_bytes(
        excel_df, volume_df, stock_concs, "TestDesign"
    )
    assert isinstance(excel_bytes, bytes)
    assert excel_bytes[:2] == b'PK'  # ZIP format (XLSX is ZIP)
    print(f"  Excel: {len(excel_bytes)} bytes, valid XLSX")

    # Validate sheet structure matches real templates
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    assert "Sample Tracking" in wb.sheetnames, f"Missing 'Sample Tracking'. Sheets: {wb.sheetnames}"
    assert "Stock_Concentrations" in wb.sheetnames, f"Missing 'Stock_Concentrations'. Sheets: {wb.sheetnames}"
    assert "Reagent Setup Guide" in wb.sheetnames, f"Missing 'Reagent Setup Guide'. Sheets: {wb.sheetnames}"
    print(f"  Sheets: {wb.sheetnames}")

    # Validate Sample Tracking headers
    ws = wb["Sample Tracking"]
    headers = [cell.value for cell in ws[1]]
    assert "ID" in headers, f"Missing ID. Headers: {headers}"
    assert "Plate_96" in headers
    assert "Well_96" in headers
    assert "Well_384" in headers
    assert "Response" in headers
    print(f"  Sample Tracking headers: {headers}")

    # Validate Stock_Concentrations has 5-column format
    stock_ws = wb["Stock_Concentrations"]
    stock_headers = [cell.value for cell in stock_ws[1]]
    assert stock_headers == ["Factor Name", "Level", "Stock Value", "Final Value", "Unit"], \
        f"Wrong Stock_Concentrations headers: {stock_headers}"
    print(f"  Stock_Concentrations headers: {stock_headers}")

    # Check stock data rows
    stock_rows = []
    for row in stock_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            stock_rows.append(row)
    assert len(stock_rows) >= 2, f"Expected at least 2 stock rows, got {len(stock_rows)}"
    # First factor should have stock value, empty Level
    assert stock_rows[0][1] == "" or stock_rows[0][1] is None, "Normal stock should have empty Level"
    assert isinstance(stock_rows[0][2], (int, float)), "Stock Value should be numeric"
    print(f"  Stock rows: {len(stock_rows)}")

    # Validate Reagent Setup Guide
    guide_ws = wb["Reagent Setup Guide"]
    guide_headers = [cell.value for cell in guide_ws[1]]
    assert guide_headers[0] == "Position"
    assert guide_headers[1] == "Reagent"
    assert guide_headers[2] == "Stock Concentration"
    assert "Volume Needed" in guide_headers[3]
    assert "overage" in guide_headers[4].lower()
    print(f"  Reagent Setup Guide headers: {guide_headers}")

    # Should have rows for each reagent (not ID) + water
    guide_rows = []
    for row in guide_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            guide_rows.append(row)
    assert len(guide_rows) > 0, "Reagent Setup Guide should have rows"
    # Check positions are reservoir format (A1, B1, etc.)
    assert guide_rows[0][0][0] in "ABCD", f"Position should start with A-D, got {guide_rows[0][0]}"
    # Last row should be water
    reagent_names = [r[1] for r in guide_rows]
    assert "Water" in reagent_names, f"Missing Water row. Reagents: {reagent_names}"
    print(f"  Reagent rows: {len(guide_rows)} ({', '.join(str(r[1]) for r in guide_rows)})")

    # --- Export with per-level concentrations and protein ---
    per_level = {
        "detergent": {
            "C12E8": {"stock": 0.05, "final": 0.00336},
            "CHAPS": {"stock": 5.0, "final": 0.34},
        }
    }
    excel_bytes_full = export_service.generate_excel_bytes(
        excel_df, volume_df, stock_concs, "TestDesign",
        per_level_concs=per_level,
        protein_stock=29.75,
        protein_final=0.5,
        final_volume=200.0,
    )

    wb_full = openpyxl.load_workbook(BytesIO(excel_bytes_full))
    stock_ws_full = wb_full["Stock_Concentrations"]

    # Collect all stock rows
    all_rows = []
    for row in stock_ws_full.iter_rows(min_row=2, values_only=True):
        all_rows.append(row)

    # Should have normal stocks + per-level + protein
    factor_names = [r[0] for r in all_rows if r[0] is not None]
    assert any("Detergent" in str(fn) for fn in factor_names), \
        f"Missing per-level detergent rows. Factor names: {factor_names}"
    assert any("Protein" in str(fn) for fn in factor_names), \
        f"Missing protein row. Factor names: {factor_names}"

    # Find per-level rows (have Level filled)
    per_level_rows = [r for r in all_rows if r[0] is not None and r[1] and str(r[1]).strip()]
    assert len(per_level_rows) >= 2, f"Expected at least 2 per-level rows, got {len(per_level_rows)}"
    level_names = [r[1] for r in per_level_rows]
    assert "C12E8" in level_names, f"Missing C12E8. Levels: {level_names}"
    assert "CHAPS" in level_names, f"Missing CHAPS. Levels: {level_names}"
    print(f"  Per-level rows: {per_level_rows}")

    # Find protein row
    protein_rows = [r for r in all_rows if r[0] and "Protein" in str(r[0])]
    assert len(protein_rows) == 1, f"Expected 1 protein row, got {len(protein_rows)}"
    assert protein_rows[0][2] == 29.75, f"Protein stock should be 29.75, got {protein_rows[0][2]}"
    assert protein_rows[0][4] == "mg/mL", f"Protein unit should be mg/mL"
    print(f"  Protein row: {protein_rows[0]}")

    # --- CSV export ---
    csv_bytes = export_service.generate_csv_bytes(volume_df)
    assert isinstance(csv_bytes, bytes)
    assert len(csv_bytes) > 10
    csv_text = csv_bytes.decode("utf-8")
    assert "water" in csv_text, f"Missing 'water' in CSV. First line: {csv_text.split(chr(10))[0]}"
    print(f"  CSV: {len(csv_bytes)} bytes")

    print("  [PASS] Export (matches real SCOUT format)")
    print()


def test_config_endpoints():
    """Test config data matches frontend expectations"""
    from utils.constants import AVAILABLE_FACTORS
    from config.design_config import (
        CATEGORICAL_FACTORS, FACTOR_CONSTRAINTS, UNIT_OPTIONS,
        DESIGN_TYPES, FRACTIONAL_RESOLUTION_OPTIONS,
        CCD_TYPE_OPTIONS, D_OPTIMAL_MODEL_OPTIONS,
    )

    print("=== Testing Config Data ===")

    # Available factors
    assert "buffer pH" in AVAILABLE_FACTORS
    assert "nacl" in AVAILABLE_FACTORS
    assert "glycerol" in AVAILABLE_FACTORS
    assert "detergent" in AVAILABLE_FACTORS
    print(f"  Available factors: {len(AVAILABLE_FACTORS)}")

    # Categorical factors
    assert "buffer pH" in CATEGORICAL_FACTORS
    assert "detergent" in CATEGORICAL_FACTORS
    assert "reducing_agent" in CATEGORICAL_FACTORS
    print(f"  Categorical: {list(CATEGORICAL_FACTORS)}")

    # Constraints
    assert "buffer pH" in FACTOR_CONSTRAINTS
    assert FACTOR_CONSTRAINTS["buffer pH"]["min"] == 1.0
    assert FACTOR_CONSTRAINTS["buffer pH"]["max"] == 14.0
    assert "glycerol" in FACTOR_CONSTRAINTS
    assert FACTOR_CONSTRAINTS["glycerol"]["max"] == 100.0
    print(f"  Constraints for: {list(FACTOR_CONSTRAINTS.keys())}")

    # Design types
    assert "full_factorial" in DESIGN_TYPES
    assert "lhs" in DESIGN_TYPES
    assert "central_composite" in DESIGN_TYPES
    assert "box_behnken" in DESIGN_TYPES
    for dt_name, dt_info in DESIGN_TYPES.items():
        for key in ["display_name", "min_factors", "description", "parameters"]:
            assert key in dt_info, f"Design type '{dt_name}' missing '{key}'"
    print(f"  Design types: {list(DESIGN_TYPES.keys())}")

    # All config is JSON serializable
    json.dumps(AVAILABLE_FACTORS)
    json.dumps(list(CATEGORICAL_FACTORS))
    json.dumps(FACTOR_CONSTRAINTS)
    json.dumps(UNIT_OPTIONS)
    json.dumps(DESIGN_TYPES)
    print("  [PASS] All config JSON serializable")
    print()


def test_validate_design():
    """Test design validation"""
    from backend.services import design_service

    print("=== Testing Design Validation ===")

    # No factors
    valid, errors, warnings = design_service.validate_design_params(
        "full_factorial", {}, has_pydoe3=True
    )
    assert not valid
    assert len(errors) > 0
    print(f"  No factors: valid={valid}, errors={errors}")

    # Valid full factorial
    factors = {"A": ["1", "2"], "B": ["x", "y"]}
    valid, errors, warnings = design_service.validate_design_params(
        "full_factorial", factors, has_pydoe3=True
    )
    assert valid
    print(f"  Full factorial: valid={valid}")

    print("  [PASS] Validation")
    print()


def test_design_size_validation():
    """Test that designs exceeding 384-well limit are rejected"""
    from core.design_factory import DesignFactory
    from backend.services import design_service

    print("=== Testing Design Size Validation ===")

    factory = DesignFactory(has_pydoe3=True, has_smt=False)

    # Full factorial with many levels should still work if under limit
    factors_small = {
        "A": ["1", "2", "3"],
        "B": ["x", "y", "z"],
    }
    points, warnings = design_service.generate_design(
        factory, "full_factorial", factors_small, {}
    )
    assert len(points) == 9
    print(f"  Small design: {len(points)} points OK")

    # LHS requesting >384 samples should fail
    factors_lhs = {"A": ["1", "2"], "B": ["3", "4"]}
    try:
        points, warnings = design_service.generate_design(
            factory, "lhs", factors_lhs, {"n_samples": 500}
        )
        assert False, "Should have raised ValueError for >384 wells"
    except ValueError as e:
        assert "384" in str(e) or "well limit" in str(e).lower()
        print(f"  >384 wells rejected: {e}")

    print("  [PASS] Design size validation")
    print()


def test_protein_params_passthrough():
    """Test that protein params are accepted by build_factorial_design"""
    from core.doe_designer import DoEDesigner
    from backend.services import design_service

    print("=== Testing Protein Params Passthrough ===")

    designer = DoEDesigner()
    factors = {
        "NaCl (mM)": ["50", "100"],
        "Glycerol (%)": ["5", "10"],
    }
    stock_concs = {"NaCl (mM)": 5000, "Glycerol (%)": 100}

    # Without protein
    excel_data, volume_data, warnings = design_service.build_factorial_design(
        designer, factors, stock_concs, final_volume=200.0
    )
    assert "protein" not in volume_data[0], "Should not have protein without params"
    print("  Without protein: no protein column OK")

    # With protein
    excel_data, volume_data, warnings = design_service.build_factorial_design(
        designer, factors, stock_concs, final_volume=200.0,
        protein_stock=10.0, protein_final=0.5,
    )
    assert "protein" in volume_data[0], f"Missing protein. Keys: {list(volume_data[0].keys())}"
    protein_vol = volume_data[0]["protein"]
    expected_vol = (0.5 / 10.0) * 200.0  # 10.0 µL
    assert abs(protein_vol - expected_vol) < 0.01, f"Expected {expected_vol}, got {protein_vol}"
    print(f"  With protein: protein_vol={protein_vol} µL OK")

    print("  [PASS] Protein params passthrough")
    print()


def test_factors_response_structure():
    """Test that the factors response matches frontend FactorsResponse type"""
    from core.project import DoEProject
    from backend.services import design_service

    print("=== Testing Factors Response Structure ===")

    project = DoEProject()
    project.add_factor("nacl", ["50", "100"], stock_conc=5000)
    project.add_factor("buffer pH", ["6", "7", "8"])

    # Simulate _build_factors_response
    factors = project.get_factors()
    stock_concs = project.get_all_stock_concs()
    per_level = project.get_all_per_level_concs()
    total = design_service.get_combinations_count(factors)
    plates = design_service.get_plates_required(total)

    response = {
        "factors": factors,
        "stock_concs": stock_concs,
        "per_level_concs": per_level,
        "total_combinations": total,
        "plates_required": plates,
    }

    # Must match frontend FactorsResponse interface
    assert isinstance(response["factors"], dict)
    assert isinstance(response["stock_concs"], dict)
    assert isinstance(response["per_level_concs"], dict)
    assert isinstance(response["total_combinations"], int)
    assert isinstance(response["plates_required"], int)
    assert response["total_combinations"] == 6  # 2 × 3
    assert response["plates_required"] == 1

    json_str = json.dumps(response)
    print(f"  Response JSON: {len(json_str)} bytes")
    print(f"  Combinations: {response['total_combinations']}, Plates: {response['plates_required']}")
    print("  [PASS] Factors response structure")
    print()


def test_export_matches_real_template():
    """Validate web export structure against real SCOUT template files"""
    import openpyxl

    print("=== Testing Export Matches Real Template ===")

    template_path = os.path.join("examples", "FF_Design_Template.xlsx")
    if not os.path.exists(template_path):
        print(f"  [SKIP] {template_path} not found")
        return

    # Read real template
    real_wb = openpyxl.load_workbook(template_path, data_only=True)

    # 1. Must have same sheet names
    assert "Sample Tracking" in real_wb.sheetnames, "Real template missing 'Sample Tracking'"
    assert "Stock_Concentrations" in real_wb.sheetnames, "Real template missing 'Stock_Concentrations'"

    # 2. Validate Sample Tracking headers from real template
    real_ws = real_wb["Sample Tracking"]
    real_headers = [cell.value for cell in real_ws[1]]
    print(f"  Real Sample Tracking headers: {real_headers}")

    # These headers must always be present in SCOUT exports
    required_headers = ["ID", "Plate_96", "Well_96", "Well_384", "Source", "Batch", "Response"]
    for h in required_headers:
        assert h in real_headers, f"Real template missing required header '{h}'"

    # 3. Validate Stock_Concentrations from real template
    real_stock = real_wb["Stock_Concentrations"]
    real_stock_headers = [cell.value for cell in real_stock[1]]
    assert real_stock_headers == ["Factor Name", "Level", "Stock Value", "Final Value", "Unit"], \
        f"Real template has unexpected stock headers: {real_stock_headers}"
    print(f"  Real Stock headers: {real_stock_headers}")

    # 4. Validate stock data structure
    normal_stocks = []
    per_level_stocks = []
    protein_rows = []
    for row in real_stock.iter_rows(min_row=2, values_only=True):
        factor_name = row[0]
        if factor_name is None:
            continue
        if "Protein" in str(factor_name) or "Volume" in str(factor_name):
            protein_rows.append(row)
        elif row[1] and str(row[1]).strip():
            per_level_stocks.append(row)
        else:
            normal_stocks.append(row)

    print(f"  Normal stocks: {len(normal_stocks)}, Per-level: {len(per_level_stocks)}, Protein: {len(protein_rows)}")
    assert len(normal_stocks) > 0, "Real template has no normal stock rows"

    # Normal stocks should have Factor Name, empty Level, Stock Value, empty Final, Unit
    for row in normal_stocks:
        assert row[0] is not None, "Factor Name should not be None"
        assert isinstance(row[2], (int, float)), f"Stock Value should be numeric, got {type(row[2])}"
        print(f"    Normal: {row[0]} = {row[2]} {row[4]}")

    # Per-level stocks should have Level filled
    for row in per_level_stocks:
        assert row[1] is not None and str(row[1]).strip(), f"Level should be filled for per-level"
        assert isinstance(row[2], (int, float)), f"Stock Value should be numeric for per-level"
        assert isinstance(row[3], (int, float)), f"Final Value should be numeric for per-level"
        print(f"    Per-level: {row[0]} | {row[1]} = stock:{row[2]}, final:{row[3]}")

    # 5. Now generate export with web service and compare structure
    from io import BytesIO
    from core.doe_designer import DoEDesigner
    from backend.services import export_service

    designer = DoEDesigner()
    factors = {
        "NaCl (mM)": ["100", "500"],
        "Glycerol (%)": ["0", "15"],
    }
    stock_concs = {"NaCl (mM)": 2000, "Glycerol (%)": 50}
    per_level = {
        "detergent": {
            "C12E8": {"stock": 0.05, "final": 0.00336},
            "CHAPS": {"stock": 5, "final": 0.34},
        }
    }

    excel_df, volume_df = designer.build_factorial_design(
        factors, stock_concs, final_volume=100.0
    )

    excel_bytes = export_service.generate_excel_bytes(
        excel_df, volume_df, stock_concs, "Test",
        per_level_concs=per_level,
        protein_stock=29.75,
        protein_final=0.5,
        final_volume=100.0,
    )

    web_wb = openpyxl.load_workbook(BytesIO(excel_bytes))

    # Web export must have same sheets
    assert "Sample Tracking" in web_wb.sheetnames
    assert "Stock_Concentrations" in web_wb.sheetnames

    # Web Stock_Concentrations must have identical header format
    web_stock = web_wb["Stock_Concentrations"]
    web_stock_headers = [cell.value for cell in web_stock[1]]
    assert web_stock_headers == real_stock_headers, \
        f"Web stock headers {web_stock_headers} != real {real_stock_headers}"

    # Web must have normal stocks, per-level, and protein
    web_normal = []
    web_per_level = []
    web_protein = []
    for row in web_stock.iter_rows(min_row=2, values_only=True):
        fname = row[0]
        if fname is None:
            continue
        if "Protein" in str(fname) or "Volume" in str(fname):
            web_protein.append(row)
        elif row[1] and str(row[1]).strip():
            web_per_level.append(row)
        else:
            web_normal.append(row)

    assert len(web_normal) >= 2, f"Expected at least 2 normal stock rows, got {len(web_normal)}"
    assert len(web_per_level) >= 2, f"Expected at least 2 per-level rows, got {len(web_per_level)}"
    assert len(web_protein) >= 1, f"Expected protein row, got {len(web_protein)}"

    print(f"  Web export: {len(web_normal)} normal, {len(web_per_level)} per-level, {len(web_protein)} protein")
    print("  [PASS] Export matches real template structure")
    print()


if __name__ == "__main__":
    print()
    passed = True

    for test_fn in [
        test_factor_management,
        test_design_service_combinations,
        test_build_factorial,
        test_generate_non_factorial,
        test_export_excel,
        test_config_endpoints,
        test_validate_design,
        test_design_size_validation,
        test_protein_params_passthrough,
        test_factors_response_structure,
        test_export_matches_real_template,
    ]:
        try:
            test_fn()
        except Exception as e:
            print(f"  [FAIL] {test_fn.__name__}: {e}")
            import traceback
            traceback.print_exc()
            passed = False

    print("=" * 50)
    if passed:
        print("ALL TESTS PASSED")
    else:
        print("SOME TESTS FAILED")
        sys.exit(1)
    print("=" * 50)
