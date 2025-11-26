"""
Bayesian Optimization Logic
Extracted from doe_analysis_gui.pyw
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats as scipy_stats
import os
from datetime import datetime

# Check if Ax is available
try:
    from ax.service.ax_client import AxClient, ObjectiveProperties
    # Note: OutcomeConstraint and ComparisonOp imports removed - not currently used
    AX_AVAILABLE = True
except ImportError:
    AX_AVAILABLE = False

# Debug flag for Pareto frontier calculations
DEBUG = True


class BayesianOptimizer:
    """Bayesian Optimization for intelligent experiment suggestions"""

    # Professional colorblind-safe palette (Okabe-Ito)
    COLORS = {
        'primary': '#0173B2',      # Blue
        'accent': '#CC78BC',       # Reddish Purple
        'warning': '#D55E00',      # Vermillion
    }

    # Debug flag - set to True to see detailed console output
    DEBUG = True

    def __init__(self):
        self.ax_client = None
        self.data = None
        self.factor_columns = []
        self.numeric_factors = []
        self.categorical_factors = []
        self.response_column = None  # Backward compatibility (primary response)
        self.response_columns = []  # Multi-response support
        self.response_directions = {}  # {response_name: 'maximize' or 'minimize'}
        self.response_constraints = {}  # {response_name: {'min': val, 'max': val}}
        self.objective_constraints = {}  # Constraints on objectives (for post-filtering)
        self.exploration_mode = False  # Allow suggestions outside constraints
        self.factor_bounds = {}
        self.is_initialized = False
        self.is_multi_objective = False  # Track if using multi-objective optimization
        self.name_mapping = {}  # Maps sanitized names back to original names
        self.reverse_mapping = {}  # Maps original names to sanitized names
        self.trial_metadata = {}  # Maps trial_index to original row data
    
    def _sanitize_name(self, name):
        """Replace spaces and special characters with underscores for Ax compatibility"""
        sanitized = name.replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '')
        return sanitized
    
    def set_data(self, data, factor_columns, categorical_factors, numeric_factors,
                 response_column=None, response_columns=None, response_directions=None,
                 response_constraints=None, exploration_mode=False):
        """Set data and factor information

        Args:
            response_column: Single response (backward compatibility)
            response_columns: List of response columns (multi-objective)
            response_directions: Dict of {response_name: 'maximize' or 'minimize'}
            response_constraints: Dict of {response_name: {'min': val, 'max': val}}
            exploration_mode: Allow suggestions outside constraints for exploration
        """
        self.data = data.copy()
        self.factor_columns = factor_columns
        self.categorical_factors = categorical_factors
        self.numeric_factors = numeric_factors

        # Support both single and multiple responses
        if response_columns is not None:
            self.response_columns = response_columns if isinstance(response_columns, list) else [response_columns]
            self.response_column = self.response_columns[0]  # Primary response for backward compatibility
        elif response_column is not None:
            self.response_column = response_column
            self.response_columns = [response_column]
        else:
            raise ValueError("Must specify either response_column or response_columns")

        # Store optimization directions
        self.response_directions = response_directions or {}
        if self.DEBUG:
            print("\n" + "="*70)
            print("OPTIMIZER DATA SETUP")
            print("="*70)
            print(f"[DEBUG OPTIMIZER] Response columns: {self.response_columns}")
            print(f"[DEBUG OPTIMIZER] Response directions (input): {response_directions}")

        # Default to maximize if not specified
        for resp in self.response_columns:
            if resp not in self.response_directions:
                self.response_directions[resp] = 'maximize'
                if self.DEBUG:
                    print(f"  '{resp}' direction not specified, defaulting to 'maximize'")

        if self.DEBUG:
            print(f"[DEBUG OPTIMIZER] Response directions (final): {self.response_directions}")

        # Store response constraints
        self.response_constraints = response_constraints or {}
        if self.DEBUG:
            print(f"[DEBUG OPTIMIZER] Response constraints: {self.response_constraints}")

        # Store exploration mode
        self.exploration_mode = exploration_mode
        if self.DEBUG:
            print(f"[DEBUG OPTIMIZER] Exploration mode: {self.exploration_mode}")

        # Multi-objective if more than one response
        self.is_multi_objective = len(self.response_columns) > 1
        if self.DEBUG:
            print(f"[DEBUG OPTIMIZER] Multi-objective: {self.is_multi_objective}")

        # Create name mappings
        self.reverse_mapping = {name: self._sanitize_name(name) for name in factor_columns}
        self.name_mapping = {self._sanitize_name(name): name for name in factor_columns}
        
        self._calculate_bounds()
    
    def _calculate_bounds(self):
        """Calculate bounds for each factor from the data"""
        self.factor_bounds = {}
        
        for factor in self.numeric_factors:
            min_val = float(self.data[factor].min())
            max_val = float(self.data[factor].max())
            self.factor_bounds[factor] = (min_val, max_val)
        
        for factor in self.categorical_factors:
            unique_vals = self.data[factor].unique().tolist()
            self.factor_bounds[factor] = unique_vals
    
    def initialize_optimizer(self, minimize=False):
        """Initialize Ax client with parameters

        Args:
            minimize: If single-objective, whether to minimize (backward compatibility)
                     Ignored if multi-objective (uses response_directions instead)
        """
        if not AX_AVAILABLE:
            raise ImportError("Ax platform not available. Install with: pip install ax-platform")

        # Build parameters list with sanitized names
        parameters = []
        for factor in self.factor_columns:
            sanitized_name = self.reverse_mapping[factor]

            if factor in self.numeric_factors:
                # Special handling for pH - treat as ordered categorical
                if 'ph' in factor.lower() and 'buffer' in factor.lower():
                    # Get unique pH values that were ACTUALLY TESTED
                    tested_ph_values = sorted(self.data[factor].unique().tolist())

                    print(f"ℹ️  Treating '{factor}' as ordered categorical parameter")
                    print(f"   Tested pH values: {tested_ph_values}")
                    print(f"   BO will only suggest from these tested values (no extrapolation)")

                    parameters.append({
                        "name": sanitized_name,
                        "type": "choice",
                        "values": tested_ph_values,
                        "is_ordered": True,  # Tells Ax that pH values have natural ordering
                        "value_type": "float"
                    })
                else:
                    # Normal continuous numeric factors
                    min_val, max_val = self.factor_bounds[factor]
                    parameters.append({
                        "name": sanitized_name,
                        "type": "range",
                        "bounds": [min_val, max_val],
                        "value_type": "float"
                    })
            elif factor in self.categorical_factors:
                parameters.append({
                    "name": sanitized_name,
                    "type": "choice",
                    "values": self.factor_bounds[factor],
                    "value_type": "str"
                })

        # Build objectives dict
        objectives = {}
        if self.is_multi_objective:
            # Multi-objective optimization
            if self.DEBUG:
                print("\n" + "="*70)
                print("BAYESIAN OPTIMIZATION INITIALIZATION")
                print("="*70)
                print(f"[DEBUG INIT] Multi-objective optimization")
                print(f"[DEBUG INIT] Optimizing {len(self.response_columns)} objectives:")
            for response in self.response_columns:
                direction = self.response_directions[response]
                minimize_this = (direction == 'minimize')
                objectives[response] = ObjectiveProperties(minimize=minimize_this)
                if self.DEBUG:
                    arrow = '↓' if minimize_this else '↑'
                    print(f"  {arrow} {response}: {direction} (Ax minimize={minimize_this})")
        else:
            # Single-objective (backward compatible)
            direction = self.response_directions.get(self.response_column, 'maximize')
            minimize = (direction == 'minimize')
            objectives[self.response_column] = ObjectiveProperties(minimize=minimize)
            if self.DEBUG:
                print("\n" + "="*70)
                print("BAYESIAN OPTIMIZATION INITIALIZATION")
                print("="*70)
                print(f"[DEBUG INIT] Single-objective optimization")
                arrow = '↓' if minimize else '↑'
                print(f"  {arrow} {self.response_column}: {direction} (Ax minimize={minimize})")

        # Build outcome constraints from response_constraints
        # Ax does NOT allow constraints on objective metrics, so separate them
        outcome_constraints = []
        objective_constraints = {}  # Store constraints on objectives for post-filtering

        if self.DEBUG:
            print(f"\n[DEBUG INIT] Building outcome constraints...")
            print(f"  Response constraints: {self.response_constraints}")
            print(f"  Exploration mode: {self.exploration_mode}")
            print(f"  Objectives: {self.response_columns}")

        if self.response_constraints and not self.exploration_mode:
            if self.DEBUG:
                print(f"[DEBUG INIT] Processing response constraints:")
            for response, constraint in self.response_constraints.items():
                if self.DEBUG:
                    print(f"  Constraint for '{response}': {constraint}")

                # Check if this constraint is on an objective metric
                if response in self.response_columns:
                    if self.DEBUG:
                        print(f"    ⚠️  '{response}' is an OBJECTIVE - Ax won't accept constraint")
                        print(f"    → Will post-filter BO suggestions instead")
                    objective_constraints[response] = constraint
                else:
                    # This is a constraint on a non-objective metric - Ax can handle it
                    if self.DEBUG:
                        print(f"    ✓ '{response}' is NOT an objective - can add to Ax")
                    try:
                        if 'min' in constraint:
                            min_val = constraint['min']
                            constraint_str = f"{response} >= {min_val}"
                            outcome_constraints.append(constraint_str)
                            if self.DEBUG:
                                print(f"      ✓ Added: {constraint_str}")
                        if 'max' in constraint:
                            max_val = constraint['max']
                            constraint_str = f"{response} <= {max_val}"
                            outcome_constraints.append(constraint_str)
                            if self.DEBUG:
                                print(f"      ✓ Added: {constraint_str}")
                    except Exception as e:
                        if self.DEBUG:
                            print(f"      ❌ ERROR: {e}")
                            import traceback
                            traceback.print_exc()

            if self.DEBUG:
                print(f"[DEBUG INIT] Constraints to Ax: {len(outcome_constraints)}")
                print(f"[DEBUG INIT] Constraints for post-filtering: {len(objective_constraints)}")
        elif self.response_constraints and self.exploration_mode:
            if self.DEBUG:
                print(f"[DEBUG INIT] Exploration mode enabled - constraints tracked but NOT enforced:")
                for response, constraint in self.response_constraints.items():
                    parts = []
                    if 'min' in constraint:
                        parts.append(f"{response} ≥ {constraint['min']}")
                    if 'max' in constraint:
                        parts.append(f"{response} ≤ {constraint['max']}")
                    if parts:
                        print(f"  {' and '.join(parts)} (guidance only)")
        else:
            if self.DEBUG:
                print(f"[DEBUG INIT] No constraints to apply")

        # Store objective constraints for post-filtering in get_next_suggestions
        self.objective_constraints = objective_constraints

        # Create Ax client
        if self.DEBUG:
            print(f"\n[DEBUG INIT] Creating Ax experiment...")
            print(f"  Parameters: {len(parameters)} defined")
            print(f"  Objectives: {list(objectives.keys())}")
            print(f"  Outcome constraints: {len(outcome_constraints)} defined")

        self.ax_client = AxClient()
        try:
            self.ax_client.create_experiment(
                name="doe_optimization",
                parameters=parameters,
                objectives=objectives,
                outcome_constraints=outcome_constraints if outcome_constraints else None,
                choose_generation_strategy_kwargs={"num_initialization_trials": 0}
            )
            if self.DEBUG:
                print(f"[DEBUG INIT] ✓ Ax experiment created successfully")
        except Exception as e:
            if self.DEBUG:
                print(f"[DEBUG INIT] ❌ ERROR creating Ax experiment: {e}")
                import traceback
                traceback.print_exc()
            raise

        # Add existing data as completed trials using sanitized names
        if self.DEBUG:
            print(f"\n[DEBUG INIT] Adding {len(self.data)} existing trials to Ax...")
            print(f"  Available columns: {list(self.data.columns)}")
            print(f"  Looking for ID column...")
        for idx, row in self.data.iterrows():
            params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                val = row[factor]
                # Ensure proper types
                if factor in self.numeric_factors:
                    # pH is now categorical, but still pass as float
                    if 'ph' in factor.lower() and 'buffer' in factor.lower():
                        params[sanitized_name] = float(val)
                    else:
                        params[sanitized_name] = float(val)
                else:
                    params[sanitized_name] = str(val)

            # Add trial
            _, trial_index = self.ax_client.attach_trial(parameters=params)

            # Store metadata for this trial
            # Try multiple possible ID column names (case-insensitive)
            exp_id = None
            id_col_found = None
            for id_col in ['ID', 'Id', 'id', 'Exp ID', 'Exp_ID', 'ExpID']:
                if id_col in row.index:
                    exp_id = row.get(id_col)
                    id_col_found = id_col
                    break

            # If no ID column found, use the pandas index + 1 (convert to 1-based)
            if exp_id is None:
                exp_id = idx + 1  # Convert 0-based index to 1-based ID
                if self.DEBUG and idx == 0:
                    print(f"  ⚠️  No ID column found, using pandas index + 1 as ID (1-based)")

            # Convert ID to Python int (handles numpy types from DataFrame)
            if exp_id is not None:
                exp_id = int(exp_id)

            metadata = {
                'row_index': idx,
                'id': exp_id,
                'parameters': {factor: row[factor] for factor in self.factor_columns}
            }
            self.trial_metadata[trial_index] = metadata

            if self.DEBUG and idx < 3:  # Show first 3 for debugging
                if id_col_found:
                    print(f"  Trial {trial_index}: ID={exp_id} (from column '{id_col_found}'), pandas_row={idx}")
                else:
                    print(f"  Trial {trial_index}: ID={exp_id} (pandas_row {idx} + 1), pandas_row={idx}")

            # Build raw_data dict for all responses
            if self.is_multi_objective:
                raw_data = {response: float(row[response]) for response in self.response_columns}
            else:
                raw_data = float(row[self.response_column])

            self.ax_client.complete_trial(
                trial_index=trial_index,
                raw_data=raw_data
            )

        self.is_initialized = True
        if self.DEBUG:
            print(f"[DEBUG INIT] ✓ Initialized BO with {len(self.data)} trials")
            print(f"  Stored metadata for {len(self.trial_metadata)} trials")

    def check_constraint_violations(self, predicted_values):
        """Check if predicted response values violate constraints

        Args:
            predicted_values: Dict of {response_name: value}

        Returns:
            List of constraint violation messages
        """
        violations = []
        if not self.response_constraints:
            return violations

        for response, constraint in self.response_constraints.items():
            if response not in predicted_values:
                continue

            value = predicted_values[response]

            if 'min' in constraint and value < constraint['min']:
                msg = f"{response} = {value:.2f} < {constraint['min']} (min)"
                violations.append(msg)
            if 'max' in constraint and value > constraint['max']:
                msg = f"{response} = {value:.2f} > {constraint['max']} (max)"
                violations.append(msg)

        return violations

    def get_next_suggestions(self, n=5):
        """Get next experiment suggestions with original factor names and proper rounding

        If objective constraints exist, generates extra suggestions and filters them.
        """
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized. Call initialize_optimizer() first.")

        # Check if we have objective constraints
        # NOTE: We don't filter objective constraints because:
        # 1. Model predictions aren't reliably available for pending trials
        # 2. BO naturally respects optimization direction
        # 3. Constraints on objectives are guidance, not hard limits
        has_objective_constraints = hasattr(self, 'objective_constraints') and self.objective_constraints

        # For objective constraints, NEVER filter - just return BO suggestions
        # User can see predicted values and decide which to run
        max_attempts = n
        need_filtering = False

        all_suggestions = []
        filtered_suggestions = []

        if self.DEBUG:
            print("\n" + "="*70)
            print("BAYESIAN OPTIMIZATION SUGGESTIONS")
            print("="*70)
            print(f"[DEBUG SUGGESTIONS] Generating {n} suggestions")
            if has_objective_constraints:
                print(f"  Objective constraints (informational):")
                for constraint, value in self.objective_constraints.items():
                    print(f"    {constraint}: {value}")
                print(f"  Note: BO suggestions not filtered by constraints")

        for i in range(max_attempts):
            params, trial_index = self.ax_client.get_next_trial()

            # Convert sanitized names back to original names and apply rounding
            original_params = {}
            for sanitized_name, value in params.items():
                original_name = self.name_mapping[sanitized_name]

                # Apply rounding based on factor type
                if isinstance(value, (int, float)):
                    # Check if this is a pH factor (now categorical, no rounding needed)
                    if 'ph' in original_name.lower() and 'buffer' in original_name.lower():
                        original_params[original_name] = float(value)
                    else:
                        # Round to 2 decimals for other numeric factors
                        rounded_value = round(value, 2)
                        original_params[original_name] = rounded_value
                else:
                    # Categorical factors - keep as is
                    original_params[original_name] = value

            # Get model prediction for this suggestion
            model_prediction = self.ax_client.get_model_predictions()

            # Extract predicted means from the trial
            # Structure: model_prediction = ({metric_name: {trial_index: (mean, sem)}}, {cov_data})
            predicted_values = {}
            if model_prediction and model_prediction[0]:
                for metric_name in self.response_columns:
                    if metric_name in model_prediction[0] and trial_index in model_prediction[0][metric_name]:
                        # Get (mean, sem) tuple and extract mean
                        predicted_values[metric_name] = model_prediction[0][metric_name][trial_index][0]

            if self.DEBUG and i < 2:  # Debug first 2
                print(f"  Trial {trial_index}: predicted {predicted_values}")

            # Abandon trial so we can generate more suggestions
            self.ax_client.abandon_trial(trial_index)

            # For compatible constraints (Minimize+Min or Maximize+Max), don't filter
            # Just accept all suggestions since BO naturally respects these
            if not need_filtering:
                all_suggestions.append(original_params)
                if len(all_suggestions) >= n:
                    break
                continue  # Skip filtering logic

            # Check if this suggestion meets objective constraints (for incompatible constraints only)
            if has_objective_constraints and predicted_values:
                meets_constraints = True
                violations = []

                for response, constraint in self.objective_constraints.items():
                    if response in predicted_values:
                        value = predicted_values[response]
                        direction = self.response_directions.get(response, 'maximize')

                        # Check constraints, but be lenient for "compatible" constraints
                        # Compatible: Minimize + Min constraint OR Maximize + Max constraint
                        # These define optimization boundaries, not hard filters

                        if 'min' in constraint:
                            is_compatible = (direction == 'minimize')  # Minimizing toward a minimum threshold
                            if value < constraint['min']:
                                if not is_compatible:  # Only enforce strictly if incompatible
                                    meets_constraints = False
                                    violations.append(f"{response}={value:.2f} < {constraint['min']}")
                                # For compatible constraints, allow some tolerance
                                elif value < constraint['min'] * 0.95:  # 5% tolerance
                                    meets_constraints = False
                                    violations.append(f"{response}={value:.2f} < {constraint['min']} (>5% below)")

                        if 'max' in constraint:
                            is_compatible = (direction == 'maximize')  # Maximizing toward a maximum threshold
                            if value > constraint['max']:
                                if not is_compatible:  # Only enforce strictly if incompatible
                                    meets_constraints = False
                                    violations.append(f"{response}={value:.2f} > {constraint['max']}")
                                # For compatible constraints, allow some tolerance
                                elif value > constraint['max'] * 1.05:  # 5% tolerance
                                    meets_constraints = False
                                    violations.append(f"{response}={value:.2f} > {constraint['max']} (>5% above)")

                if meets_constraints:
                    filtered_suggestions.append(original_params)
                    if len(filtered_suggestions) >= n:
                        break  # Got enough valid suggestions
                else:
                    if self.DEBUG and i < 3:  # Show first few rejections
                        print(f"  ⚠️  Suggestion {i+1} rejected: {', '.join(violations)}")
            else:
                # No objective constraints, accept all suggestions
                all_suggestions.append(original_params)
                if len(all_suggestions) >= n:
                    break

        result = filtered_suggestions if need_filtering else all_suggestions

        if self.DEBUG:
            if need_filtering:
                print(f"\n[DEBUG SUGGESTIONS] ✓ Generated {max_attempts} candidates, {len(result)} met constraints")
                if len(result) < n:
                    print(f"[DEBUG SUGGESTIONS] ⚠️  Could only find {len(result)}/{n} valid suggestions")
            else:
                print(f"\n[DEBUG SUGGESTIONS] ✓ Generated {len(result)} suggestions (no filtering needed)")

        return result[:n]  # Return at most n suggestions

    def get_pareto_frontier(self):
        """Get Pareto frontier for multi-objective optimization

        Returns:
            List of dicts with 'parameters', 'objectives', 'id', and 'row_index' for Pareto-optimal points
        """
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized")

        if not self.is_multi_objective:
            return None  # No Pareto frontier for single-objective

        try:
            if self.DEBUG:
                print("\n" + "="*70)
                print("PARETO FRONTIER EXTRACTION")
                print("="*70)
            # Get Pareto frontier from Ax
            pareto_frontier = self.ax_client.get_pareto_optimal_parameters()
            if self.DEBUG:
                print(f"[DEBUG PARETO] Ax returned {len(pareto_frontier)} Pareto-optimal trials")

            # Convert to more usable format
            pareto_points = []
            for i, (arm_name, (params, values)) in enumerate(pareto_frontier.items(), 1):
                # Extract trial index from arm_name
                # Ax may return either int directly or string like "trial_0_arm_0"
                if isinstance(arm_name, int):
                    trial_index = arm_name
                else:
                    trial_index = int(arm_name.split('_')[0])

                # Get metadata for this trial
                metadata = self.trial_metadata.get(trial_index, {})

                # Use stored original parameters from metadata
                original_params = metadata.get('parameters', {})

                # If metadata not found, fall back to converting sanitized names
                if not original_params:
                    if self.DEBUG:
                        print(f"  ⚠️  Pareto point {i}: No metadata found for trial {trial_index}, using fallback")
                    original_params = {}
                    for sanitized_name, value in params.items():
                        if sanitized_name in self.name_mapping:
                            original_name = self.name_mapping[sanitized_name]
                            original_params[original_name] = value

                # Extract mean values from tuple: values = (mean_dict, cov_dict)
                mean_dict = values[0] if isinstance(values, tuple) else values

                exp_id = metadata.get('id')
                row_index = metadata.get('row_index')
                if self.DEBUG and i <= 3:  # Show first 3 for debugging
                    print(f"\n  Pareto point {i}:")
                    print(f"    Trial index (Ax): {trial_index}")
                    print(f"    Row index (pandas): {row_index}")
                    print(f"    ID: {exp_id}")
                    print(f"    Parameters (experimental conditions):")
                    for param_name, param_value in original_params.items():
                        print(f"      {param_name}: {param_value}")
                    print(f"    Objectives (results):")
                    for obj_name, obj_value in mean_dict.items():
                        print(f"      {obj_name}: {obj_value:.2f}")

                pareto_points.append({
                    'parameters': original_params,
                    'objectives': mean_dict,
                    'id': exp_id,
                    'row_index': metadata.get('row_index')
                })

            if self.DEBUG:
                print(f"\n[DEBUG PARETO] ✓ Extracted {len(pareto_points)} Pareto points with metadata")
            return pareto_points

        except Exception as e:
            if self.DEBUG:
                print(f"\n[DEBUG PARETO] ⚠️  Could not extract Pareto frontier: {str(e)}")
                import traceback
                traceback.print_exc()
            return None

    def calculate_hypervolume(self, reference_point=None):
        """Calculate hypervolume indicator for multi-objective optimization

        Args:
            reference_point: Dict of {response_name: reference_value}
                           If None, uses worst observed values

        Returns:
            Hypervolume value (higher is better)
        """
        if not self.is_multi_objective:
            return None

        try:
            from ax.modelbridge.modelbridge_utils import observed_hypervolume

            # Get reference point
            if reference_point is None:
                # Use worst observed values as reference
                reference_point = {}
                for response in self.response_columns:
                    direction = self.response_directions[response]
                    if direction == 'maximize':
                        reference_point[response] = float(self.data[response].min())
                    else:  # minimize
                        reference_point[response] = float(self.data[response].max())

            # Calculate hypervolume
            hv = observed_hypervolume(
                modelbridge=None,
                experiment=self.ax_client.experiment,
                metrics=list(self.response_columns),
                reference_point=reference_point
            )

            return hv

        except Exception as e:
            print(f"⚠️  Could not calculate hypervolume: {str(e)}")
            return None

    def plot_pareto_frontier(self):
        """
        Create Pareto frontier visualization for multi-objective optimization.

        Scalable strategy based on number of objectives:
        - 2 objectives: 2D scatter + parallel coordinates
        - 3 objectives: Parallel coordinates + radar + optional 3D
        - 4-6 objectives: Parallel coordinates + radar
        - 7+ objectives: Parallel coordinates only

        Returns:
            matplotlib Figure with subplots or None
        """
        if not self.is_multi_objective:
            return None

        pareto_points = self.get_pareto_frontier()

        if pareto_points is None or len(pareto_points) == 0:
            print("⚠️  No Pareto frontier points available")
            return None

        n_objectives = len(self.response_columns)

        # Strategy: Always show parallel coordinates (universal),
        # add other plots based on n_objectives
        if n_objectives == 2:
            # 2 objectives: Show 2D scatter + parallel coordinates side by side
            fig = plt.figure(figsize=(16, 7))

            # Left: 2D scatter (original)

        if n_objectives == 2:
            # 2D scatter plot
            fig, ax = plt.subplots(1, 1, figsize=(8, 6))

            resp1, resp2 = self.response_columns[0], self.response_columns[1]

            # Plot all observed points
            all_r1 = self.data[resp1].values
            all_r2 = self.data[resp2].values
            ax.scatter(all_r1, all_r2, c='lightgray', s=100, alpha=0.5,
                      label='Observed', zorder=1, edgecolors='gray', linewidths=0.5)

            # Separate Pareto points by constraint violations
            valid_r1, valid_r2 = [], []
            violated_r1, violated_r2 = [], []

            for p in pareto_points:
                violations = self.check_constraint_violations(p['objectives'])
                if violations:
                    violated_r1.append(p['objectives'][resp1])
                    violated_r2.append(p['objectives'][resp2])
                else:
                    valid_r1.append(p['objectives'][resp1])
                    valid_r2.append(p['objectives'][resp2])

            # Plot valid Pareto points (meets constraints)
            if valid_r1:
                ax.scatter(valid_r1, valid_r2, c=self.COLORS['primary'], s=200,
                          label='Pareto Frontier (meets constraints)', zorder=2, marker='*',
                          edgecolors='black', linewidths=1.5)

            # Plot violated Pareto points (violates constraints)
            if violated_r1:
                ax.scatter(violated_r1, violated_r2, c=self.COLORS['warning'], s=200,
                          label='Pareto Frontier (violates constraints)', zorder=2, marker='*',
                          edgecolors='darkred', linewidths=1.5)

            # Add arrows showing optimization direction
            dir1 = self.response_directions[resp1]
            dir2 = self.response_directions[resp2]
            arrow1 = '→' if dir1 == 'maximize' else '←'
            arrow2 = '↑' if dir2 == 'maximize' else '↓'

            ax.set_xlabel(f"{resp1} ({dir1}) {arrow1}", fontsize=12, fontweight='bold')
            ax.set_ylabel(f"{resp2} ({dir2}) {arrow2}", fontsize=12, fontweight='bold')
            ax.set_title(f'Pareto Frontier: {resp1} vs {resp2}\n({len(pareto_points)} optimal points)',
                        fontsize=14, fontweight='bold', pad=15)
            ax.legend(fontsize=11, loc='best')
            ax.grid(True, alpha=0.3)

            plt.tight_layout()
            return fig

        elif n_objectives == 3:
            # 3D scatter plot
            from mpl_toolkits.mplot3d import Axes3D
            fig = plt.figure(figsize=(9, 7))
            ax = fig.add_subplot(111, projection='3d')

            resp1, resp2, resp3 = self.response_columns[0], self.response_columns[1], self.response_columns[2]

            # Plot all observed points
            all_r1 = self.data[resp1].values
            all_r2 = self.data[resp2].values
            all_r3 = self.data[resp3].values
            ax.scatter(all_r1, all_r2, all_r3, c='lightgray', s=80, alpha=0.4,
                      label='Observed', edgecolors='gray', linewidths=0.5)

            # Separate Pareto points by constraint violations
            valid_r1, valid_r2, valid_r3 = [], [], []
            violated_r1, violated_r2, violated_r3 = [], [], []

            for p in pareto_points:
                violations = self.check_constraint_violations(p['objectives'])
                if violations:
                    violated_r1.append(p['objectives'][resp1])
                    violated_r2.append(p['objectives'][resp2])
                    violated_r3.append(p['objectives'][resp3])
                else:
                    valid_r1.append(p['objectives'][resp1])
                    valid_r2.append(p['objectives'][resp2])
                    valid_r3.append(p['objectives'][resp3])

            # Plot valid Pareto points (meets constraints)
            if valid_r1:
                ax.scatter(valid_r1, valid_r2, valid_r3, c=self.COLORS['primary'], s=250,
                          label='Pareto (meets constraints)', marker='*',
                          edgecolors='black', linewidths=1.5)

            # Plot violated Pareto points (violates constraints)
            if violated_r1:
                ax.scatter(violated_r1, violated_r2, violated_r3, c=self.COLORS['warning'], s=250,
                          label='Pareto (violates constraints)', marker='*',
                          edgecolors='darkred', linewidths=1.5)

            # Labels
            dir1 = self.response_directions[resp1]
            dir2 = self.response_directions[resp2]
            dir3 = self.response_directions[resp3]

            ax.set_xlabel(f"{resp1} ({dir1})", fontsize=11, fontweight='bold')
            ax.set_ylabel(f"{resp2} ({dir2})", fontsize=11, fontweight='bold')
            ax.set_zlabel(f"{resp3} ({dir3})", fontsize=11, fontweight='bold')
            ax.set_title(f'3D Pareto Frontier\n({len(pareto_points)} optimal points)',
                        fontsize=14, fontweight='bold', pad=15)
            ax.legend(fontsize=10)

            plt.tight_layout()
            return fig

        else:
            print(f"⚠️  Pareto frontier visualization supports 2-3 objectives (you have {n_objectives})")
            return None

    def plot_pareto_parallel_coordinates(self):
        """
        Create parallel coordinates plot for Pareto frontier.
        Works for 2+ objectives. Industry standard for multi-objective visualization.

        Returns:
            matplotlib Figure or None
        """
        if not self.is_multi_objective:
            return None

        pareto_points = self.get_pareto_frontier()

        if pareto_points is None or len(pareto_points) == 0:
            print("⚠️  No Pareto frontier points available")
            return None

        n_objectives = len(self.response_columns)

        # Create figure (sized to fit window like other plots)
        fig, ax = plt.subplots(1, 1, figsize=(8, 5))

        # Normalize all objectives to 0-1 scale for visualization
        # Get min/max for each objective from ALL data
        normalized_data = {}
        if DEBUG:
            print("\n" + "="*70)
            print("PARALLEL COORDINATES NORMALIZATION DEBUG")
            print("="*70)

        for resp in self.response_columns:
            data_min = self.data[resp].min()
            data_max = self.data[resp].max()
            normalized_data[resp] = {'min': data_min, 'max': data_max}

            if DEBUG:
                direction = self.response_directions[resp]
                print(f"\n{resp}:")
                print(f"  Direction: {direction}")
                print(f"  Data range: [{data_min:.2f}, {data_max:.2f}]")
                print(f"  After normalization: lower values → 0, higher values → 1")
                if direction == 'minimize':
                    print(f"  After flip: LOWER values (better) → 1, HIGHER values (worse) → 0")

        # Plot all observed points (gray, thin lines)
        for idx, row in self.data.iterrows():
            values = []
            for resp in self.response_columns:
                val = row[resp]
                norm_val = (val - normalized_data[resp]['min']) / (normalized_data[resp]['max'] - normalized_data[resp]['min'] + 1e-10)
                # Flip for minimize objectives (so "better" is always up)
                if self.response_directions[resp] == 'minimize':
                    norm_val = 1 - norm_val
                values.append(norm_val)

            ax.plot(range(n_objectives), values, c='lightgray', alpha=0.3, linewidth=0.8, zorder=1)

        # Plot Pareto points (colored, thick lines)
        colors = plt.cm.viridis(np.linspace(0, 1, len(pareto_points)))

        if DEBUG:
            print(f"\nPareto Points (showing first 3):")

        for i, (p, color) in enumerate(zip(pareto_points, colors)):
            values = []
            if DEBUG and i < 3:
                print(f"\n  Pareto {i+1}:")

            for resp in self.response_columns:
                val = p['objectives'][resp]
                norm_val = (val - normalized_data[resp]['min']) / (normalized_data[resp]['max'] - normalized_data[resp]['min'] + 1e-10)

                if DEBUG and i < 3:
                    print(f"    {resp}: raw={val:.2f}, normalized={norm_val:.3f}", end="")

                # Flip for minimize objectives
                if self.response_directions[resp] == 'minimize':
                    norm_val = 1 - norm_val
                    if DEBUG and i < 3:
                        print(f" → flipped={norm_val:.3f}")
                elif DEBUG and i < 3:
                    print()

                values.append(norm_val)

            # Check constraints
            violations = self.check_constraint_violations(p['objectives'])
            linestyle = '--' if violations else '-'
            linewidth = 2.5 if not violations else 2.0

            label = f"Pareto {i+1}"
            if p.get('id') is not None:
                label += f" (ID: {p['id']})"

            ax.plot(range(n_objectives), values, c=color, linewidth=linewidth,
                   linestyle=linestyle, marker='o', markersize=8, label=label, zorder=2)

        # Customize axes
        ax.set_xticks(range(n_objectives))

        # Create labels with direction arrows
        labels = []
        for resp in self.response_columns:
            direction = self.response_directions[resp]
            arrow = '↑' if direction == 'maximize' else '↓'
            labels.append(f"{resp}\n{arrow} {direction}")

        ax.set_xticklabels(labels, fontsize=10, fontweight='bold')
        ax.set_ylabel('Normalized Performance\n(higher is better after normalization)',
                     fontsize=11, fontweight='bold')
        ax.set_ylim(-0.05, 1.05)
        ax.set_title(f'Pareto Frontier - Parallel Coordinates\n({len(pareto_points)} optimal trade-offs)',
                    fontsize=13, fontweight='bold', pad=15)

        # Add grid
        ax.grid(True, alpha=0.3, axis='y')
        ax.set_axisbelow(True)

        # Legend (show only first 10 to avoid clutter)
        if len(pareto_points) <= 10:
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=9)
        else:
            ax.text(1.02, 0.5, f'{len(pareto_points)} Pareto points\n(colors show different solutions)',
                   transform=ax.transAxes, fontsize=10, verticalalignment='center')

        plt.tight_layout()
        return fig

    def plot_pareto_radar(self):
        """
        Create radar/spider chart for Pareto frontier.
        Works well for 3-6 objectives. Shows trade-off shapes.

        Returns:
            matplotlib Figure or None
        """
        if not self.is_multi_objective:
            return None

        pareto_points = self.get_pareto_frontier()

        if pareto_points is None or len(pareto_points) == 0:
            print("⚠️  No Pareto frontier points available")
            return None

        n_objectives = len(self.response_columns)

        if n_objectives < 3:
            print("⚠️  Radar chart requires at least 3 objectives (you have {n_objectives})")
            return None

        if n_objectives > 6:
            print("⚠️  Radar chart not recommended for >6 objectives (too cluttered)")
            return None

        # Setup radar chart
        angles = np.linspace(0, 2 * np.pi, n_objectives, endpoint=False).tolist()
        angles += angles[:1]  # Complete the circle

        # Create figure (sized to fit window like other plots)
        fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(projection='polar'))

        # Normalize objectives to 0-1
        normalized_data = {}
        for resp in self.response_columns:
            data_min = self.data[resp].min()
            data_max = self.data[resp].max()
            normalized_data[resp] = {'min': data_min, 'max': data_max}

        # Plot each Pareto point
        colors = plt.cm.viridis(np.linspace(0, 1, len(pareto_points)))

        for i, (p, color) in enumerate(zip(pareto_points, colors)):
            values = []
            for resp in self.response_columns:
                val = p['objectives'][resp]
                norm_val = (val - normalized_data[resp]['min']) / (normalized_data[resp]['max'] - normalized_data[resp]['min'] + 1e-10)
                # Flip for minimize objectives
                if self.response_directions[resp] == 'minimize':
                    norm_val = 1 - norm_val
                values.append(norm_val)

            values += values[:1]  # Complete the circle

            violations = self.check_constraint_violations(p['objectives'])
            linestyle = '--' if violations else '-'
            linewidth = 2.0 if not violations else 1.5

            label = f"Pareto {i+1}"
            if p.get('id') is not None:
                label += f" (ID: {p['id']})"

            ax.plot(angles, values, 'o-', linewidth=linewidth, linestyle=linestyle,
                   color=color, label=label, markersize=6)
            ax.fill(angles, values, alpha=0.15, color=color)

        # Customize
        ax.set_xticks(angles[:-1])

        labels = []
        for resp in self.response_columns:
            direction = self.response_directions[resp]
            arrow = '↑' if direction == 'maximize' else '↓'
            labels.append(f"{resp} {arrow}")

        ax.set_xticklabels(labels, fontsize=10, fontweight='bold')
        ax.set_ylim(0, 1)
        ax.set_yticks([0.25, 0.5, 0.75, 1.0])
        ax.set_yticklabels(['25%', '50%', '75%', '100%'], fontsize=9)
        ax.set_title(f'Pareto Frontier - Radar Chart\n({len(pareto_points)} optimal trade-offs)',
                    fontsize=13, fontweight='bold', pad=20, y=1.08)

        # Legend
        if len(pareto_points) <= 8:
            ax.legend(bbox_to_anchor=(1.3, 1.0), loc='upper left', fontsize=9)

        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        return fig

    def _create_suggestion_heatmap(self, factor_x_orig, factor_y_orig,
                                   factor_x_san, factor_y_san, X, Y):
        """Create a heatmap showing where BO suggests exploring - ENHANCED VERSION"""
        try:
            # Larger for export quality
            fig, ax = plt.subplots(1, 1, figsize=(9, 7))

            # Generate many suggestions and extract their x,y positions
            n_suggestions = 50
            suggestions_x = []
            suggestions_y = []

            for _ in range(n_suggestions):
                try:
                    params, trial_idx = self.ax_client.get_next_trial()
                    suggestions_x.append(params[factor_x_san])
                    suggestions_y.append(params[factor_y_san])
                    self.ax_client.abandon_trial(trial_idx)
                except:
                    break

            if len(suggestions_x) < 5:
                print("⚠ Could not generate enough suggestions for heatmap")
                return None

            # Create 2D histogram / density plot
            from scipy.stats import gaussian_kde
            xy = np.vstack([suggestions_x, suggestions_y])
            z = gaussian_kde(xy)(xy)

            # Plot scatter with density colors - colorblind-safe colormap
            scatter = ax.scatter(suggestions_x, suggestions_y, c=z, s=150,
                               cmap='viridis', alpha=0.7,
                               edgecolors='white', linewidth=2)

            # Enhanced colorbar
            cbar = plt.colorbar(scatter, ax=ax, label='Suggestion Density')
            cbar.ax.tick_params(labelsize=11)
            cbar.set_label('Suggestion Density\n(Higher = More Recommended)',
                          fontsize=12, fontweight='bold')

            # Plot existing experiments with modern styling
            existing_x = self.data[factor_x_orig].values
            existing_y = self.data[factor_y_orig].values
            ax.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                      edgecolors='white', linewidth=3,
                      label='Existing Experiments', zorder=5, marker='s', alpha=0.9)

            # Enhanced labels and title
            ax.set_xlabel(factor_x_orig, fontsize=13, fontweight='bold', color='#333333')
            ax.set_ylabel(factor_y_orig, fontsize=13, fontweight='bold', color='#333333')
            ax.set_title('Bayesian Optimization: Recommended Exploration Regions',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')

            # Modern legend
            ax.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')

            # Modern grid
            ax.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax.set_axisbelow(True)

            # Set background
            ax.set_facecolor('#FAFAFA')

            # Borders
            for spine in ax.spines.values():
                spine.set_edgecolor('#CCCCCC')
                spine.set_linewidth(1.5)

            plt.tight_layout()
            return fig

        except Exception as e:
            print(f"❌ Error creating suggestion heatmap: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _select_most_important_factors(self):
        """Select the 2 most important numeric factors using feature importances"""
        if len(self.numeric_factors) <= 2:
            return self.numeric_factors[:2] if len(self.numeric_factors) == 2 else None

        try:
            # Use BoTorch's feature_importances method
            adapter = self.ax_client.generation_strategy.adapter

            if hasattr(adapter, 'generator') and hasattr(adapter.generator, 'feature_importances'):
                # Get feature importances from BoTorchGenerator
                importances = adapter.generator.feature_importances()

                # importances is a numpy array - flatten it
                import numpy as np
                importance_values = np.array(importances).flatten()

                # Get parameter names in the order they appear in the model
                param_names = list(adapter.parameters)

                # Create dict mapping original factor names to importance values
                factor_importances = {}
                for idx, param_name in enumerate(param_names):
                    if idx < len(importance_values):
                        # Find the original factor name from the sanitized parameter name
                        for orig_factor in self.numeric_factors:
                            sanitized = self.reverse_mapping.get(orig_factor, orig_factor)
                            if sanitized == param_name:
                                factor_importances[orig_factor] = float(importance_values[idx])
                                break

                if len(factor_importances) >= 2:
                    print(f"✓ Using feature importances for factor selection")
                    print(f"  Feature importances: {factor_importances}")

                    # Sort by importance and take top 2
                    sorted_factors = sorted(factor_importances.items(), key=lambda x: x[1], reverse=True)
                    selected = [f[0] for f in sorted_factors[:2]]

                    print(f"  Selected most important factors: {selected}")
                    return selected

            # If feature importances not available, fall back
            print("⚠ Could not extract feature importances, falling back to range-based selection")

        except NotImplementedError:
            # Some models (e.g., mixed continuous/categorical) don't support feature importances
            print("⚠ Feature importances not supported for this model type")
            print("  Falling back to range-based selection")
        except Exception as e:
            print(f"⚠ Error extracting feature importances: {e}")
            print("  Falling back to range-based selection")

        # Fallback: Sobol sensitivity analysis - works for all model types
        # Uses Ax's built-in global sensitivity analysis with Sobol indices
        try:
            print("ℹ️  Using Sobol sensitivity analysis for factor selection...")

            # Try to use Ax's built-in Sobol sensitivity analysis
            from ax.analysis.plotly.sensitivity import SensitivityAnalysisPlot

            analysis = SensitivityAnalysisPlot()
            card = analysis.compute(
                experiment=self.ax_client._experiment,
                generation_strategy=self.ax_client._generation_strategy
            )

            # Extract Sobol indices from the analysis card
            # The blob is a JSON string containing the plotly figure
            if card is not None:
                import json
                import plotly.graph_objects as go

                # Parse the JSON blob to get the plotly figure data
                blob_json = json.loads(card.blob)
                fig = go.Figure(blob_json)

                # Parse Sobol indices from the plotly figure data
                # The data structure contains parameter names and their total Sobol indices
                sensitivities = {}

                if hasattr(fig, 'data') and len(fig.data) > 0:
                    # Extract from plotly bar chart data
                    # Note: In plotly's encoding, parameter names are in y and values are in x (as binary data)
                    for idx, trace in enumerate(fig.data):
                        if hasattr(trace, 'x') and hasattr(trace, 'y'):
                            # Extract parameter names (in trace.y) and Sobol values (in trace.x as binary)
                            param_names = trace.y  # Parameter names are in y
                            sobol_data = trace.x   # Sobol values are in x (binary encoded)

                            # Decode binary-encoded Sobol values
                            if isinstance(param_names, (list, tuple)) and isinstance(sobol_data, dict):
                                import base64
                                import numpy as np

                                # Decode base64 binary data
                                binary_data = base64.b64decode(sobol_data['bdata'])
                                dtype = sobol_data.get('dtype', 'f8')
                                sobol_values = np.frombuffer(binary_data, dtype=dtype)

                                print(f"ℹ️  Decoded {len(sobol_values)} Sobol indices from analysis")

                                # Map parameter names to Sobol values
                                for param_name, sobol_val in zip(param_names, sobol_values):
                                    # Map sanitized parameter names back to original factor names
                                    for orig_factor in self.numeric_factors:
                                        sanitized = self.reverse_mapping.get(orig_factor, orig_factor)
                                        if sanitized == param_name:
                                            sensitivities[orig_factor] = float(sobol_val)
                                            break
                else:
                    print(f"ℹ️  No Sobol data found in analysis")

                if len(sensitivities) >= 2:
                    print(f"✓ Using Sobol sensitivity indices for factor selection")
                    print(f"  Sobol indices: {sensitivities}")

                    sorted_factors = sorted(sensitivities.items(), key=lambda x: x[1], reverse=True)
                    selected = [f[0] for f in sorted_factors[:2]]

                    print(f"  Selected most important factors: {selected}")
                    return selected
                else:
                    print(f"⚠ Could not extract enough Sobol indices (got {len(sensitivities)})")

        except ImportError:
            print(f"ℹ️  Sobol analysis not available (Ax version may not support it)")
        except Exception as e:
            print(f"⚠ Sobol analysis failed: {e}")
            # Continue to range-based fallback

        # Ultimate fallback: largest parameter ranges
        factor_ranges = {}
        for factor in self.numeric_factors:
            min_val, max_val = self.factor_bounds[factor]
            data_range = max_val - min_val
            if data_range > 0:
                factor_ranges[factor] = data_range

        sorted_factors = sorted(factor_ranges.items(), key=lambda x: x[1], reverse=True)
        selected = [f[0] for f in sorted_factors[:2]]

        print(f"ℹ️  Selected factors based on range: {selected}")
        return selected
    
    def get_acquisition_plot(self):
        """Generate comprehensive BO visualization with 4 panels (GUI preview)

        Note: Only for single-objective optimization.
        For multi-objective, use plot_pareto_frontier() instead.
        """
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized")

        # This function is only for single-objective
        if self.is_multi_objective:
            print("⚠️  get_acquisition_plot() called for multi-objective optimization")
            print("   Use plot_pareto_frontier() instead for multi-objective")
            return None

        # Verify we have the response column
        if self.response_column is None:
            print("⚠️  No response column set for acquisition plot")
            return None

        # Need at least 2 numeric factors
        if len(self.numeric_factors) < 2:
            print(f"⚠️  Need at least 2 numeric factors for acquisition plot (have {len(self.numeric_factors)})")
            return None

        try:
            # Select the 2 most important numeric factors
            selected_factors = self._select_most_important_factors()
            if selected_factors is None or len(selected_factors) < 2:
                return None

            factor_x_original = selected_factors[0]
            factor_y_original = selected_factors[1]

            factor_x_sanitized = self.reverse_mapping[factor_x_original]
            factor_y_sanitized = self.reverse_mapping[factor_y_original]

            print(f"ℹ️  Creating comprehensive BO plots for: {factor_x_original} vs {factor_y_original}")

            # Create grid - handle categorical factors specially
            # For pH (ordered categorical), use only tested values, not continuous range
            is_x_categorical = ('ph' in factor_x_original.lower() and 'buffer' in factor_x_original.lower())
            is_y_categorical = ('ph' in factor_y_original.lower() and 'buffer' in factor_y_original.lower())

            if is_x_categorical:
                # Use only tested pH values for X axis
                x = np.array(sorted(self.data[factor_x_original].unique()))
            else:
                # Continuous factor - use linspace
                x_min, x_max = self.factor_bounds[factor_x_original]
                if x_min == x_max:
                    print(f"⚠ Factor has no variation: {factor_x_original}={x_min}")
                    return None
                # Add small padding to bounds
                x_range = x_max - x_min
                x_min -= 0.05 * x_range
                x_max += 0.05 * x_range
                x = np.linspace(x_min, x_max, 15)

            if is_y_categorical:
                # Use only tested pH values for Y axis
                y = np.array(sorted(self.data[factor_y_original].unique()))
            else:
                # Continuous factor - use linspace
                y_min, y_max = self.factor_bounds[factor_y_original]
                if y_min == y_max:
                    print(f"⚠ Factor has no variation: {factor_y_original}={y_min}")
                    return None
                # Add small padding to bounds
                y_range = y_max - y_min
                y_min -= 0.05 * y_range
                y_max += 0.05 * y_range
                y = np.linspace(y_min, y_max, 15)

            X, Y = np.meshgrid(x, y)

            # Get a template with all other factors at their median values
            template_params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                if factor == factor_x_original or factor == factor_y_original:
                    continue
                if factor in self.numeric_factors:
                    template_params[sanitized_name] = float(self.data[factor].median())
                else:
                    template_params[sanitized_name] = str(self.data[factor].mode()[0])

            # Build list of parameterizations for batch prediction
            parameterizations = []
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    params = template_params.copy()
                    params[factor_x_sanitized] = float(X[i, j])
                    params[factor_y_sanitized] = float(Y[i, j])
                    parameterizations.append(params)

            print(f"ℹ️  Predicting {len(parameterizations)} points using BO model...")
            print(f"   Response metric: {self.response_column}")
            print(f"   Grid shape: {X.shape}")

            # Get predictions with uncertainty
            try:
                predictions_list = self.ax_client.get_model_predictions_for_parameterizations(
                    parameterizations=parameterizations,
                    metric_names=[self.response_column]
                )

                # Extract predictions and uncertainty into arrays
                # IMPORTANT: Use dtype=float to prevent truncation when X is integer (categorical factors)
                Z_mean = np.zeros_like(X, dtype=float)
                Z_sem = np.zeros_like(X, dtype=float)

                idx = 0
                for i in range(X.shape[0]):
                    for j in range(X.shape[1]):
                        pred = predictions_list[idx]

                        # Ax returns dict format: {response_name: (mean, variance)}
                        if isinstance(pred, dict):
                            mean_variance_tuple = pred[self.response_column]
                            pred_mean = mean_variance_tuple[0]
                            variance = mean_variance_tuple[1]
                            pred_sem = np.sqrt(variance)
                        else:
                            raise ValueError(f"Unexpected format: {type(pred)}")

                        Z_mean[i, j] = pred_mean
                        Z_sem[i, j] = pred_sem
                        idx += 1

                print(f"✓ Predicted {len(parameterizations)} points")

            except Exception as e:
                print(f"\n{'='*60}")
                print(f"❌ BATCH PREDICTION FAILED")
                print(f"{'='*60}")
                print(f"Error type: {type(e).__name__}")
                print(f"Error message: {e}")
                print(f"{'='*60}")
                import traceback
                traceback.print_exc()
                print(f"{'='*60}")
                print(f"Falling back to suggestion density heatmap...")
                print(f"{'='*60}\n")
                return self._create_suggestion_heatmap(factor_x_original, factor_y_original,
                                                       factor_x_sanitized, factor_y_sanitized,
                                                       X, Y)

            # CREATE 2x2 MULTI-PANEL FIGURE (GUI PREVIEW)
            fig = plt.figure(figsize=(12, 9))
            gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.3)

            # PANEL 1: Response Surface (Top-Left)
            ax1 = fig.add_subplot(gs[0, 0])
            contour1 = ax1.contourf(X, Y, Z_mean, levels=15, cmap='RdYlGn', alpha=0.9)
            cbar1 = plt.colorbar(contour1, ax=ax1)
            cbar1.ax.tick_params(labelsize=8)
            cbar1.set_label(f'{self.response_column}', fontsize=9, fontweight='bold')

            contour_lines1 = ax1.contour(X, Y, Z_mean, levels=8, colors='black',
                                        alpha=0.3, linewidths=1)
            ax1.clabel(contour_lines1, inline=True, fontsize=7, fmt='%.1f')

            existing_x = self.data[factor_x_original].values
            existing_y = self.data[factor_y_original].values
            ax1.scatter(existing_x, existing_y, c='#2E86AB', s=80,
                       edgecolors='white', linewidth=2, label='Tested', zorder=5, alpha=0.9)

            ax1.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax1.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax1.set_title('Response Surface (GP Mean)', fontsize=11, fontweight='bold', pad=10)
            ax1.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax1.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax1.set_facecolor('#FAFAFA')

            # PANEL 2: Acquisition Function (Top-Right)
            ax2 = fig.add_subplot(gs[0, 1])
            # Calculate Expected Improvement (EI) - Proper formula
            # Get current best based on optimization direction
            direction = self.response_directions.get(self.response_column, 'maximize')
            if direction == 'maximize':
                current_best = self.data[self.response_column].max()
            else:  # minimize
                current_best = self.data[self.response_column].min()

            # Avoid division by zero
            Z_sem_safe = np.where(Z_sem > 1e-6, Z_sem, 1e-6)

            # Standardized improvement (for minimize, we negate to find improvement)
            if direction == 'maximize':
                Z_score = (Z_mean - current_best) / Z_sem_safe
                # Proper EI formula: EI = (mu - best) * Phi(Z) + sigma * phi(Z)
                Z_ei = (Z_mean - current_best) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)
            else:  # minimize
                Z_score = (current_best - Z_mean) / Z_sem_safe
                # For minimize: improvement when prediction is lower than current best
                Z_ei = (current_best - Z_mean) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)

            Z_ei = np.maximum(Z_ei, 0)  # EI is always non-negative

            contour2 = ax2.contourf(X, Y, Z_ei, levels=15, cmap='viridis', alpha=0.9)
            cbar2 = plt.colorbar(contour2, ax=ax2)
            cbar2.ax.tick_params(labelsize=8)
            cbar2.set_label('EI Score', fontsize=9, fontweight='bold')

            ax2.scatter(existing_x, existing_y, c='#2E86AB', s=80,
                       edgecolors='white', linewidth=2, label='Tested', zorder=5, alpha=0.9)

            # Mark best point so far (based on direction)
            if direction == 'maximize':
                best_idx = self.data[self.response_column].idxmax()
            else:  # minimize
                best_idx = self.data[self.response_column].idxmin()
            best_x = self.data.loc[best_idx, factor_x_original]
            best_y = self.data.loc[best_idx, factor_y_original]
            ax2.scatter([best_x], [best_y], c='gold', s=200, marker='*',
                       edgecolors='black', linewidth=2, label='Best', zorder=6)

            ax2.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax2.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax2.set_title('Acquisition Function (Where to Sample)', fontsize=11, fontweight='bold', pad=10)
            ax2.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax2.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax2.set_facecolor('#FAFAFA')

            # PANEL 3: GP Uncertainty Map (Bottom-Left)
            ax3 = fig.add_subplot(gs[1, 0])

            # Plot uncertainty (standard error) as contour map
            # Z_sem was computed earlier - it shows where the model is uncertain
            contour3 = ax3.contourf(X, Y, Z_sem, levels=15, cmap='YlOrRd', alpha=0.9)

            # Add contour lines for clarity
            contour_lines3 = ax3.contour(X, Y, Z_sem, levels=8, colors='black',
                                         linewidths=0.5, alpha=0.3)
            ax3.clabel(contour_lines3, inline=True, fontsize=7, fmt='%.2f')

            # Mark observed points
            ax3.scatter(existing_x, existing_y, c='black', s=80,
                       marker='o', edgecolors='white', linewidth=2,
                       label='Observed Points', zorder=5)

            # Mark the best point (already calculated above)
            ax3.scatter([best_x], [best_y], c='lime', s=200, marker='*',
                       edgecolors='black', linewidth=2, label='Best Point', zorder=6)

            # Add colorbar
            cbar3 = plt.colorbar(contour3, ax=ax3, orientation='vertical', pad=0.02, shrink=0.9)
            cbar3.set_label('GP Std. Error', fontsize=9, fontweight='bold')
            cbar3.ax.tick_params(labelsize=8)

            # Calculate uncertainty statistics
            mean_uncertainty = np.mean(Z_sem)
            max_uncertainty = np.max(Z_sem)
            ax3.text(0.05, 0.95, f'Mean: {mean_uncertainty:.3f}\nMax: {max_uncertainty:.3f}',
                    transform=ax3.transAxes, fontsize=9, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.8, edgecolor='#CCCCCC'))

            ax3.set_xlabel(factor_x_original, fontsize=10, fontweight='bold')
            ax3.set_ylabel(factor_y_original, fontsize=10, fontweight='bold')
            ax3.set_title('Model Uncertainty (GP Std. Error)', fontsize=11, fontweight='bold', pad=10)
            ax3.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC', loc='lower right')
            ax3.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax3.set_facecolor('#FAFAFA')

            # PANEL 4: Optimization Progress (Bottom-Right)
            ax4 = fig.add_subplot(gs[1, 1])

            # Get cumulative best values based on optimization direction
            response_values = self.data[self.response_column].values
            if direction == 'maximize':
                cumulative_best = np.maximum.accumulate(response_values)
            else:  # minimize
                cumulative_best = np.minimum.accumulate(response_values)
            iterations = np.arange(1, len(cumulative_best) + 1)

            # Plot progress
            ax4.plot(iterations, cumulative_best, 'o-', color='#029E73', linewidth=2.5,
                    markersize=6, markerfacecolor='#029E73', markeredgecolor='white',
                    markeredgewidth=1.5, label='Best So Far', alpha=0.9)

            # Fill area under curve
            ax4.fill_between(iterations, cumulative_best, alpha=0.2, color='#029E73')

            # Mark current best
            ax4.scatter([len(iterations)], [cumulative_best[-1]], c='gold', s=200,
                       marker='*', edgecolors='black', linewidth=2, zorder=5,
                       label='Current Best')

            # Set Y-axis limits based on data range (with 10% margin)
            y_min = min(response_values.min(), cumulative_best.min())
            y_max = cumulative_best.max()
            y_range = y_max - y_min
            ax4.set_ylim(y_min - 0.1 * y_range, y_max + 0.05 * y_range)

            ax4.set_xlabel('Experiment Number', fontsize=10, fontweight='bold')
            ax4.set_ylabel(f'Best {self.response_column}', fontsize=10, fontweight='bold')
            ax4.set_title('Optimization Progress', fontsize=11, fontweight='bold', pad=10)
            ax4.legend(fontsize=8, framealpha=0.9, edgecolor='#CCCCCC')
            ax4.grid(True, alpha=0.15, linestyle='-', linewidth=0.5)
            ax4.set_facecolor('#FAFAFA')

            # Annotate improvement
            total_improvement = cumulative_best[-1] - cumulative_best[0]
            ax4.text(0.05, 0.05, f'Total Δ: {total_improvement:+.2f}',
                    transform=ax4.transAxes, fontsize=9, fontweight='bold',
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.8,
                    edgecolor='#CCCCCC'))

            plt.tight_layout()
            return fig
            
        except Exception as e:
            print(f"Error generating acquisition plot: {e}")
            import traceback
            traceback.print_exc()
            return None

    def export_bo_plots(self, directory, base_name="Experiment", date_str=None, file_format="png", dpi=300):
        """Export individual high-resolution BO plots for publication

        Args:
            directory: Output directory
            base_name: Base filename
            date_str: Date string (YYYYMMDD)
            file_format: File format (png, tiff, pdf, eps)
            dpi: Resolution for raster formats (PNG, TIFF)
        """
        if not self.is_initialized:
            raise ValueError("Optimizer not initialized")

        if date_str is None:
            date_str = datetime.now().strftime('%Y%m%d')

        exported_files = []

        try:
            # Get the necessary data (same as preview)
            selected_factors = self._select_most_important_factors()
            if selected_factors is None or len(selected_factors) < 2:
                return []

            factor_x_original = selected_factors[0]
            factor_y_original = selected_factors[1]
            factor_x_sanitized = self.reverse_mapping[factor_x_original]
            factor_y_sanitized = self.reverse_mapping[factor_y_original]

            # Create higher resolution grid for export
            x_min, x_max = self.factor_bounds[factor_x_original]
            y_min, y_max = self.factor_bounds[factor_y_original]
            x_range = x_max - x_min
            y_range = y_max - y_min
            x_min -= 0.05 * x_range
            x_max += 0.05 * x_range
            y_min -= 0.05 * y_range
            y_max += 0.05 * y_range

            x = np.linspace(x_min, x_max, 30)  # Higher resolution for export
            y = np.linspace(y_min, y_max, 30)
            X, Y = np.meshgrid(x, y)

            # Template params
            template_params = {}
            for factor in self.factor_columns:
                sanitized_name = self.reverse_mapping[factor]
                if factor == factor_x_original or factor == factor_y_original:
                    continue
                if factor in self.numeric_factors:
                    template_params[sanitized_name] = float(self.data[factor].median())
                else:
                    template_params[sanitized_name] = str(self.data[factor].mode()[0])

            # Build parameterizations
            parameterizations = []
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    params = template_params.copy()
                    params[factor_x_sanitized] = float(X[i, j])
                    params[factor_y_sanitized] = float(Y[i, j])
                    parameterizations.append(params)

            # Get predictions
            predictions_list = self.ax_client.get_model_predictions_for_parameterizations(
                parameterizations=parameterizations,
                metric_names=[self.response_column]
            )

            # IMPORTANT: Use dtype=float to prevent truncation when X is integer (categorical factors)
            Z_mean = np.zeros_like(X, dtype=float)
            Z_sem = np.zeros_like(X, dtype=float)
            idx = 0
            for i in range(X.shape[0]):
                for j in range(X.shape[1]):
                    pred = predictions_list[idx]
                    # Ax returns dict format: {response_name: (mean, variance)}
                    mean_variance_tuple = pred[self.response_column]
                    pred_mean = mean_variance_tuple[0]
                    variance = mean_variance_tuple[1]
                    pred_sem = np.sqrt(variance)
                    Z_mean[i, j] = pred_mean
                    Z_sem[i, j] = pred_sem
                    idx += 1

            existing_x = self.data[factor_x_original].values
            existing_y = self.data[factor_y_original].values

            # EXPORT 1: Response Surface
            fig1, ax1 = plt.subplots(1, 1, figsize=(9, 7))
            contour1 = ax1.contourf(X, Y, Z_mean, levels=20, cmap='RdYlGn', alpha=0.9)
            cbar1 = plt.colorbar(contour1, ax=ax1)
            cbar1.ax.tick_params(labelsize=11)
            cbar1.set_label(f'Predicted {self.response_column}', fontsize=12, fontweight='bold')
            contour_lines1 = ax1.contour(X, Y, Z_mean, levels=10, colors='black',
                                        alpha=0.4, linewidths=1.5)
            ax1.clabel(contour_lines1, inline=True, fontsize=9, fmt='%.1f')
            ax1.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                       edgecolors='white', linewidth=3, label='Existing Experiments',
                       zorder=5, marker='o', alpha=0.9)
            ax1.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax1.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax1.set_title('Bayesian Optimization: Predicted Response Surface',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax1.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax1.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax1.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath1 = os.path.join(directory, f'{base_name}_BO_ResponseSurface_{date_str}.{file_format}')
            fig1.savefig(filepath1, dpi=dpi, bbox_inches='tight')
            plt.close(fig1)
            exported_files.append(filepath1)

            # EXPORT 2: Acquisition Function
            fig2, ax2 = plt.subplots(1, 1, figsize=(9, 7))

            # Get current best based on optimization direction
            direction = self.response_directions.get(self.response_column, 'maximize')
            if direction == 'maximize':
                current_best = self.data[self.response_column].max()
            else:  # minimize
                current_best = self.data[self.response_column].min()

            # Proper EI formula
            Z_sem_safe = np.where(Z_sem > 1e-6, Z_sem, 1e-6)

            if direction == 'maximize':
                Z_score = (Z_mean - current_best) / Z_sem_safe
                Z_ei = (Z_mean - current_best) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)
            else:  # minimize
                Z_score = (current_best - Z_mean) / Z_sem_safe
                Z_ei = (current_best - Z_mean) * scipy_stats.norm.cdf(Z_score) + Z_sem_safe * scipy_stats.norm.pdf(Z_score)

            Z_ei = np.maximum(Z_ei, 0)
            contour2 = ax2.contourf(X, Y, Z_ei, levels=20, cmap='viridis', alpha=0.9)
            cbar2 = plt.colorbar(contour2, ax=ax2)
            cbar2.ax.tick_params(labelsize=11)
            cbar2.set_label('Expected Improvement Score', fontsize=12, fontweight='bold')
            ax2.scatter(existing_x, existing_y, c='#2E86AB', s=150,
                       edgecolors='white', linewidth=3, label='Tested', zorder=5, alpha=0.9)

            # Mark best point based on direction
            if direction == 'maximize':
                best_idx = self.data[self.response_column].idxmax()
            else:  # minimize
                best_idx = self.data[self.response_column].idxmin()
            best_x = self.data.loc[best_idx, factor_x_original]
            best_y = self.data.loc[best_idx, factor_y_original]
            ax2.scatter([best_x], [best_y], c='gold', s=300, marker='*',
                       edgecolors='black', linewidth=3, label='Current Best', zorder=6)
            ax2.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax2.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax2.set_title('Bayesian Optimization: Acquisition Function',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax2.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax2.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax2.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath2 = os.path.join(directory, f'{base_name}_BO_Acquisition_{date_str}.{file_format}')
            fig2.savefig(filepath2, dpi=dpi, bbox_inches='tight')
            plt.close(fig2)
            exported_files.append(filepath2)

            # EXPORT 3: GP Uncertainty Map
            fig3, ax3 = plt.subplots(1, 1, figsize=(9, 7))

            # Plot uncertainty (standard error) as contour map
            contour3 = ax3.contourf(X, Y, Z_sem, levels=15, cmap='YlOrRd', alpha=0.9)

            # Add contour lines for clarity
            contour_lines3 = ax3.contour(X, Y, Z_sem, levels=10, colors='black',
                                         linewidths=0.8, alpha=0.4)
            ax3.clabel(contour_lines3, inline=True, fontsize=9, fmt='%.2f')

            # Mark observed points
            ax3.scatter(existing_x, existing_y, c='black', s=120,
                       marker='o', edgecolors='white', linewidth=2.5,
                       label='Observed Points', zorder=5)

            # Mark the best point (already calculated above)
            ax3.scatter([best_x], [best_y], c='lime', s=350, marker='*',
                       edgecolors='black', linewidth=3, label='Best Point', zorder=6)

            # Add colorbar
            cbar3 = plt.colorbar(contour3, ax=ax3, orientation='vertical', pad=0.02, shrink=0.9)
            cbar3.set_label('GP Standard Error', fontsize=13, fontweight='bold')
            cbar3.ax.tick_params(labelsize=11)

            # Calculate uncertainty statistics
            mean_uncertainty = np.mean(Z_sem)
            max_uncertainty = np.max(Z_sem)
            min_uncertainty = np.min(Z_sem)
            ax3.text(0.05, 0.95, f'Mean: {mean_uncertainty:.3f}\\nMax: {max_uncertainty:.3f}\\nMin: {min_uncertainty:.3f}',
                    transform=ax3.transAxes, fontsize=11, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.9, edgecolor='#CCCCCC', linewidth=2))

            ax3.set_xlabel(factor_x_original, fontsize=13, fontweight='bold', color='#333333')
            ax3.set_ylabel(factor_y_original, fontsize=13, fontweight='bold', color='#333333')
            ax3.set_title('Model Uncertainty (GP Standard Error)',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax3.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='lower right')
            ax3.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax3.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath3 = os.path.join(directory, f'{base_name}_BO_Uncertainty_{date_str}.{file_format}')
            fig3.savefig(filepath3, dpi=dpi, bbox_inches='tight')
            plt.close(fig3)
            exported_files.append(filepath3)

            # EXPORT 4: Progress
            fig4, ax4 = plt.subplots(1, 1, figsize=(9, 7))
            response_values = self.data[self.response_column].values

            # Get cumulative best based on direction
            if direction == 'maximize':
                cumulative_best = np.maximum.accumulate(response_values)
                total_improvement = cumulative_best[-1] - cumulative_best[0]
            else:  # minimize
                cumulative_best = np.minimum.accumulate(response_values)
                total_improvement = cumulative_best[0] - cumulative_best[-1]  # Positive for improvement

            iterations = np.arange(1, len(cumulative_best) + 1)
            ax4.plot(iterations, cumulative_best, 'o-', color='#029E73', linewidth=3,
                    markersize=8, markerfacecolor='#029E73', markeredgecolor='white',
                    markeredgewidth=2, label='Best Value Found', alpha=0.9)
            ax4.fill_between(iterations, cumulative_best, alpha=0.2, color='#029E73')
            ax4.scatter([len(iterations)], [cumulative_best[-1]], c='gold', s=400,
                       marker='*', edgecolors='black', linewidth=3, zorder=5,
                       label='Current Best')
            # Set Y-axis limits based on data range (with 10% margin)
            y_min = min(response_values.min(), cumulative_best.min())
            y_max = cumulative_best.max()
            y_range = y_max - y_min
            ax4.set_ylim(y_min - 0.1 * y_range, y_max + 0.05 * y_range)
            ax4.text(0.05, 0.95, f'Total Improvement: {total_improvement:+.2f}\\nIterations: {len(iterations)}',
                    transform=ax4.transAxes, fontsize=11, fontweight='bold',
                    verticalalignment='top', bbox=dict(boxstyle='round',
                    facecolor='white', alpha=0.9, edgecolor='#CCCCCC', linewidth=2))
            ax4.set_xlabel('Experiment Number', fontsize=13, fontweight='bold', color='#333333')
            ax4.set_ylabel(f'Best {self.response_column}', fontsize=13, fontweight='bold', color='#333333')
            ax4.set_title('Optimization Progress',
                        fontsize=15, fontweight='bold', pad=15, color='#1a1a1a')
            ax4.legend(fontsize=11, framealpha=0.95, edgecolor='#CCCCCC', loc='best')
            ax4.grid(True, alpha=0.2, linestyle='-', linewidth=0.8, color='gray')
            ax4.set_facecolor('#FAFAFA')
            plt.tight_layout()
            filepath4 = os.path.join(directory, f'{base_name}_BO_Progress_{date_str}.{file_format}')
            fig4.savefig(filepath4, dpi=dpi, bbox_inches='tight')
            plt.close(fig4)
            exported_files.append(filepath4)

            return exported_files

        except Exception as e:
            print(f"Error exporting BO plots: {e}")
            import traceback
            traceback.print_exc()
            return exported_files

    def _smart_column_match(self, column_name: str) -> str:
        """
        Intelligently match Excel column name to internal factor name for BO.
        Handles any factor format from ExpModel Suite.
        
        Examples:
        "NaCl (mM)" → "nacl"
        "Detergent" → "detergent" (categorical)
        "Detergent (%)" → "detergent_concentration"
        "Buffer pH" → "buffer pH"
        "Reducing Agent" → "reducing_agent" (categorical)
        "Reducing Agent (mM)" → "reducing_agent_concentration"
        """
        # Handle None or empty column names
        if column_name is None or (isinstance(column_name, str) and not column_name.strip()):
            return None

        name = column_name.strip()
        
        # Special case: Buffer pH (keep as-is for BO)
        if "Buffer pH" in name or "buffer pH" in name:
            return "buffer pH"
        
        # Extract base name by removing units
        if '(' in name:
            base_name = name.split('(')[0].strip()
        else:
            base_name = name
        
        # Normalize
        normalized = base_name.lower().replace(' ', '_').replace('-', '_')
        
        # Handle concentration suffixes
        if "buffer_conc" in normalized or "buffer conc" in base_name.lower():
            return "buffer_concentration"
        elif "detergent" in base_name.lower():
            # Check if it has units (concentration) or is categorical (name)
            if "%" in column_name or "conc" in base_name.lower():
                return "detergent_concentration"
            else:
                return "detergent"
        elif "reducing_agent" in normalized or "reducing agent" in base_name.lower():
            # Check if it has units (concentration) or is categorical (name)
            if "mM" in column_name or "conc" in base_name.lower():
                return "reducing_agent_concentration"
            else:
                return "reducing_agent"
        
        # Default: return normalized name
        return normalized
    
    def export_bo_batch_to_files(self, n_suggestions, batch_number, excel_path,
                                 stock_concs, final_volume, buffer_ph_values,
                                 per_level_concs=None):
        """Export BO suggestions to Excel and Opentrons CSV

        Args:
            n_suggestions: Number of BO suggestions to generate
            batch_number: Batch number for this BO iteration
            excel_path: Path to existing Excel file to append to
            stock_concs: Dict of stock concentrations {factor: concentration}
            final_volume: Final volume in µL
            buffer_ph_values: List of buffer pH values used
            per_level_concs: Optional dict of per-level concentrations {factor: {level: {stock, final}}}

        Returns:
            Tuple of (xlsx_path, csv_path) or None if failed
        """
        try:
            import openpyxl
            import csv as csv_module
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Get BO suggestions (already rounded to 0.5 increments)
            suggestions = self.get_next_suggestions(n=n_suggestions)
            
            if not suggestions:
                print("ERROR: No suggestions returned from BO!")
                return None
            
            print(f"\nGenerating {len(suggestions)} BO suggestions...")
            print(f"First suggestion: {suggestions[0]}")
            
            # Collect pH values from BO suggestions (for dynamic CSV columns)
            bo_ph_values = []
            for suggestion in suggestions:
                # Check both possible pH column names
                for ph_key in ['Buffer pH', 'buffer pH']:
                    if ph_key in suggestion:
                        ph_value = float(suggestion[ph_key])  # Ensure float
                        if ph_value not in bo_ph_values:
                            bo_ph_values.append(ph_value)
                        break
            
            # Convert buffer_ph_values to floats and merge
            original_ph_floats = [float(ph) for ph in buffer_ph_values]
            all_ph_values = sorted(set(original_ph_floats + bo_ph_values))
            
            print(f"Original pH values: {original_ph_floats}")
            print(f"BO suggested pH values: {bo_ph_values}")
            print(f"Combined pH values for CSV: {all_ph_values}")
            
            # Read existing Excel to get last ID and structure
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Find the ACTUAL last row with data (not just max_row which includes empty rows)
            last_row_with_data = 1
            for row_num in range(1, ws.max_row + 1):
                # Check if ID column has data
                if ws.cell(row=row_num, column=1).value is not None:
                    last_row_with_data = row_num
            
            last_id = ws.cell(row=last_row_with_data, column=1).value
            next_id = int(last_id) + 1 if last_id else 1
            
            print(f"\nExcel structure:")
            print(f"  Headers: {headers}")
            print(f"  Excel max_row: {ws.max_row}")
            print(f"  Actual last row with data: {last_row_with_data}")
            print(f"  Last ID: {last_id}, Next ID: {next_id}")
            
            # Prepare new rows for Excel
            new_excel_rows = []
            volume_rows = []
            
            # Find factor column indices in Excel
            factor_col_indices = {}
            display_to_internal = {}
            response_col_idx = None
            
            for idx, header in enumerate(headers):
                # Skip None or empty headers
                if header is None or (isinstance(header, str) and not header.strip()):
                    continue
                if header == 'Response':
                    response_col_idx = idx
                    continue
                if header in ['ID', 'Plate_96', 'Well_96', 'Well_384', 'Source', 'Batch']:
                    continue

                # Use smart matching to map Excel column names to internal names
                # This will work for any factor exported by ExpModel Suite
                internal_name = self._smart_column_match(header)
                if internal_name:
                    factor_col_indices[internal_name] = idx
                    display_to_internal[header] = internal_name
            
            print(f"\nColumn mapping:")
            print(f"  Factor columns: {factor_col_indices}")
            print(f"  Response column index: {response_col_idx}")
            print(f"  Display to internal: {display_to_internal}")
            
            # Generate well positions (restart from A1 for new Opentrons run)
            
            # Collect unique values for categorical factors from all suggestions BEFORE processing
            detergent_values = set()
            reducing_agent_values = set()
            
            for sug in suggestions:
                # Check for detergent (try multiple possible keys)
                for det_key in ['detergent', 'Detergent']:
                    if det_key in sug:
                        det_val = str(sug[det_key]).strip()
                        if det_val:
                            detergent_values.add(det_val)
                        break
                
                # Check for reducing agent (try multiple possible keys)
                for agent_key in ['reducing_agent', 'Reducing Agent']:
                    if agent_key in sug:
                        agent_val = str(sug[agent_key]).strip()
                        if agent_val:
                            reducing_agent_values.add(agent_val)
                        break
            
            print(f"Detected categorical values:")
            print(f"  Detergents: {detergent_values}")
            print(f"  Reducing Agents: {reducing_agent_values}")
            
            for idx, suggestion in enumerate(suggestions):
                # Well position calculation
                plate_num = (idx // 96) + 1
                well_idx = idx % 96
                # Column-major order: A1, B1, C1...H1, A2, B2...H12
                row_letter = chr(65 + (well_idx % 8))  # 0-7 for A-H
                col_number = (well_idx // 8) + 1       # 0-11 for 1-12
                well_pos = f"{row_letter}{col_number}"
                
                # 384-well conversion
                row_384 = chr(65 + (well_idx // 12) * 2 + (plate_num - 1) % 2)
                col_384 = (well_idx % 12) * 2 + 1 + (plate_num - 1) // 2
                well_384 = f"{row_384}{col_384}"
                
                # Build Excel row matching existing structure
                excel_row = [None] * len(headers)
                excel_row[0] = next_id + idx  # ID
                excel_row[1] = plate_num  # Plate_96
                excel_row[2] = well_pos  # Well_96
                excel_row[3] = well_384  # Well_384
                excel_row[4] = "BO"  # Source
                excel_row[5] = batch_number  # Batch
                
                # Fill in factor values from suggestions
                # BO suggestions come with original display names
                for factor_name, value in suggestion.items():
                    # Try direct match with display name
                    if factor_name in headers:
                        col_idx = headers.index(factor_name)
                        excel_row[col_idx] = value
                    # Try matching through internal name conversion
                    elif factor_name in display_to_internal:
                        internal_name = display_to_internal[factor_name]
                        if internal_name in factor_col_indices:
                            excel_row[factor_col_indices[internal_name]] = value
                
                # Response column (empty) - use the correct index
                if response_col_idx is not None:
                    excel_row[response_col_idx] = ""
                
                new_excel_rows.append(excel_row)
                
                # Calculate volumes for Opentrons CSV
                volumes = {}
                total_volume_used = 0
                
                # Handle buffer pH (categorical - one column per pH value)
                if 'buffer pH' in factor_col_indices:
                    buffer_ph_col = factor_col_indices['buffer pH']
                    buffer_ph = str(excel_row[buffer_ph_col])
                    
                    # Initialize ALL buffer pH columns to 0 (including new BO pHs)
                    for ph in all_ph_values:
                        volumes[f"buffer_{ph}"] = 0
                    
                    # Calculate volume for the specific pH used
                    if 'buffer_concentration' in factor_col_indices:
                        buffer_conc_col = factor_col_indices['buffer_concentration']
                        buffer_conc_value = excel_row[buffer_conc_col]
                        
                        if buffer_conc_value is not None:
                            desired_conc = float(buffer_conc_value)
                            buffer_stock = stock_concs.get('buffer_concentration', 0)
                            
                            if buffer_stock > 0:
                                volume = (desired_conc * final_volume) / buffer_stock
                                volumes[f"buffer_{buffer_ph}"] = round(volume, 2)
                                total_volume_used += volumes[f"buffer_{buffer_ph}"]
                
                # Handle detergent (categorical - one column per detergent type)
                if 'detergent' in factor_col_indices:
                    detergent_col = factor_col_indices['detergent']
                    detergent_type = str(excel_row[detergent_col]).strip()

                    # Initialize all detergent columns to 0
                    for det in detergent_values:
                        det_clean = det.lower().replace(' ', '_').replace('-', '_')
                        volumes[det_clean] = 0

                    # Check for per-level concentrations first
                    per_level_det = per_level_concs.get('detergent', {}) if per_level_concs else {}

                    if detergent_type and per_level_det and detergent_type in per_level_det:
                        # Use per-level concentrations (stock and final are predefined)
                        level_data = per_level_det[detergent_type]
                        detergent_stock = level_data['stock']
                        desired_conc = level_data['final']

                        if detergent_stock > 0:
                            volume = (desired_conc * final_volume) / detergent_stock
                            det_clean = detergent_type.lower().replace(' ', '_').replace('-', '_')
                            volumes[det_clean] = round(volume, 2)
                            total_volume_used += volumes[det_clean]

                    elif detergent_type and 'detergent_concentration' in factor_col_indices:
                        # Fall back to normal mode (get concentration from BO suggestion)
                        detergent_conc_col = factor_col_indices['detergent_concentration']
                        detergent_conc_value = excel_row[detergent_conc_col]

                        if detergent_conc_value is not None:
                            desired_conc = float(detergent_conc_value)
                            detergent_stock = stock_concs.get('detergent_concentration', 0)

                            if detergent_stock > 0:
                                volume = (desired_conc * final_volume) / detergent_stock
                                det_clean = detergent_type.lower().replace(' ', '_').replace('-', '_')
                                volumes[det_clean] = round(volume, 2)
                                total_volume_used += volumes[det_clean]
                
                # Handle reducing_agent (categorical - one column per reducing agent type)
                if 'reducing_agent' in factor_col_indices:
                    agent_col = factor_col_indices['reducing_agent']
                    agent_type = str(excel_row[agent_col]).strip()

                    # Initialize all reducing agent columns to 0
                    for agent in reducing_agent_values:
                        agent_clean = agent.lower().replace(' ', '_').replace('-', '_')
                        volumes[agent_clean] = 0

                    # Check for per-level concentrations first
                    per_level_agent = per_level_concs.get('reducing_agent', {}) if per_level_concs else {}

                    if agent_type and per_level_agent and agent_type in per_level_agent:
                        # Use per-level concentrations (stock and final are predefined)
                        level_data = per_level_agent[agent_type]
                        agent_stock = level_data['stock']
                        desired_conc = level_data['final']

                        if agent_stock > 0:
                            volume = (desired_conc * final_volume) / agent_stock
                            agent_clean = agent_type.lower().replace(' ', '_').replace('-', '_')
                            volumes[agent_clean] = round(volume, 2)
                            total_volume_used += volumes[agent_clean]

                    elif agent_type and 'reducing_agent_concentration' in factor_col_indices:
                        # Fall back to normal mode (get concentration from BO suggestion)
                        agent_conc_col = factor_col_indices['reducing_agent_concentration']
                        agent_conc_value = excel_row[agent_conc_col]

                        if agent_conc_value is not None:
                            desired_conc = float(agent_conc_value)
                            agent_stock = stock_concs.get('reducing_agent_concentration', 0)

                            if agent_stock > 0:
                                volume = (desired_conc * final_volume) / agent_stock
                                agent_clean = agent_type.lower().replace(' ', '_').replace('-', '_')
                                volumes[agent_clean] = round(volume, 2)
                                total_volume_used += volumes[agent_clean]
                
                # Calculate volumes for other simple factors (NaCl, Zinc, Glycerol, etc.)
                for internal_name, col_idx in factor_col_indices.items():
                    if internal_name in ['buffer pH', 'buffer_concentration', 'detergent', 'detergent_concentration',
                                        'reducing_agent', 'reducing_agent_concentration']:
                        continue
                    
                    if internal_name in stock_concs:
                        factor_value = excel_row[col_idx]
                        
                        if factor_value is not None:
                            desired_conc = float(factor_value)
                            stock_conc = stock_concs[internal_name]
                            
                            if stock_conc > 0:
                                volume = (desired_conc * final_volume) / stock_conc
                                volumes[internal_name] = round(volume, 2)
                                total_volume_used += volumes[internal_name]
                
                # Calculate water
                water_volume = round(final_volume - total_volume_used, 2)
                volumes["water"] = water_volume
                
                volume_rows.append(volumes)
            
            print(f"\nGenerated {len(new_excel_rows)} new rows")
            print(f"First Excel row: {new_excel_rows[0]}")
            print(f"First volume row: {volume_rows[0]}")
            
            # Write to Excel at specific row numbers (not append!)
            start_row = last_row_with_data + 1
            print(f"\nWriting to Excel starting at row {start_row}")
            
            for idx, excel_row in enumerate(new_excel_rows):
                row_num = start_row + idx
                for col_idx, value in enumerate(excel_row, start=1):
                    ws.cell(row=row_num, column=col_idx, value=value)
                print(f"  Wrote row {row_num}: ID={excel_row[0]}")
            
            # Save Excel
            wb.save(excel_path)

            # Generate CSV path with standardized naming: [BaseName]_BO_Batch[N]_[Date]_Opentron.csv
            base_path = os.path.splitext(excel_path)[0]
            date_str = datetime.now().strftime('%Y%m%d')

            # Remove existing _Design_YYYYMMDD suffix from original file if present
            import re
            base_path = re.sub(r'_Design_\d{8}$', '', base_path)

            csv_path = f"{base_path}_BO_Batch{batch_number}_{date_str}_Opentron.csv"
            
            # Find which pH values are actually used in this batch
            used_ph_values = set()
            for volumes in volume_rows:
                for ph in all_ph_values:
                    buffer_key = f"buffer_{ph}"
                    if volumes.get(buffer_key, 0) > 0:
                        used_ph_values.add(ph)
            
            # Build CSV headers with categorical columns
            csv_headers = []
            
            # Add buffer pH columns (only used ones)
            if used_ph_values:
                for ph in sorted(used_ph_values):
                    csv_headers.append(f"buffer_{ph}")
            
            # Add detergent columns (all unique detergent types)
            if detergent_values:
                for det in sorted(detergent_values):
                    det_clean = det.lower().replace(' ', '_').replace('-', '_')
                    csv_headers.append(det_clean)
            
            # Add reducing agent columns (all unique reducing agent types)
            if reducing_agent_values:
                for agent in sorted(reducing_agent_values):
                    agent_clean = agent.lower().replace(' ', '_').replace('-', '_')
                    csv_headers.append(agent_clean)
            
            # Add other simple factors (skip categorical factors and their concentrations)
            for internal_name in factor_col_indices.keys():
                if internal_name not in ['buffer pH', 'buffer_concentration', 'detergent', 'detergent_concentration',
                                        'reducing_agent', 'reducing_agent_concentration']:
                    csv_headers.append(internal_name)
            
            csv_headers.append("water")
            
            # Write CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv_module.writer(f)
                writer.writerow(csv_headers)
                
                for volumes in volume_rows:
                    row = [volumes.get(h, 0) for h in csv_headers]
                    writer.writerow(row)
            
            print(f"\n✓ Successfully exported:")
            print(f"  Excel: {excel_path}")
            print(f"  CSV: {csv_path}")
            print(f"  Rows added: {len(new_excel_rows)}")
            
            return excel_path, csv_path
            
        except Exception as e:
            print(f"Error exporting BO batch: {e}")
            import traceback
            traceback.print_exc()
            return None

